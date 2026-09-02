#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""Marking an assessment nobody marked, so the rest of the pipeline can run.

The simulator stands in for steps 5 and 6 of the lifecycle -- the grader
writing a mark on the feedback sheet, and the grader copying it into their
own workbook. It writes **both** records, because they are two records and
step 7 exists to catch them disagreeing. So the thing this suite really
checks is that the two records are separable: that a run with
``discrepancies=`` produces exactly the failure `reconcile_marks` is for,
and a run without one produces none.

House convention: ``pl`` is pathlib, ``pr`` is polars.
"""

import sys

import pandas as pd
import pytest

from grader_helper import (
    allocate_graders,
    catch_grades,
    distribute_feedback_sheets,
    import_brightspace_classlist,
    ingest_completed_graderfiles,
    load_module,
    reconcile_marks,
)
from grader_helper.simulating import (
    BOUNDARY_MARKS,
    grade_cell_value,
    draw_marks,
    feedback_sheets,
    simulate_marking,
)


@pytest.fixture
def ready_to_mark(tmp_path, repo_root):
    """A module allocated and distributed, with every sheet still blank.

    Exactly where a module leader is on the morning the graders start.
    """
    sys.path.insert(0, str(repo_root / "tests"))
    from fake_module import make_fake_module

    root = tmp_path / "PS4001"
    make_fake_module(root, distributed=False, marked=False)

    module = load_module(root)
    cw1 = module.assessment("cw1")
    class_list = import_brightspace_classlist(module.classlist_path)
    allocate_graders(cw1, class_list, seed=1)
    distribute_feedback_sheets(cw1.submissions_path, cw1.rubric_path)
    return module


@pytest.fixture
def cw1(ready_to_mark):
    return ready_to_mark.assessment("cw1")


def _graders(assessment):
    return [g.initials for g in assessment.graders]


def _both_records(assessment):
    """What the two halves of the audit see, as the ML would fetch them."""
    received = catch_grades(assessment.submissions_path, assessment.grade_cell)
    reported = ingest_completed_graderfiles(
        assessment.grading_output_path, _graders(assessment), file_type="excel"
    )
    return received, reported


# ---------------------------------------------------------------------------
# The feedback sheets
# ---------------------------------------------------------------------------


def test_every_distributed_sheet_gets_a_mark(cw1):
    simulate_marking(cw1, seed=1)

    received = catch_grades(cw1.submissions_path, cw1.grade_cell)
    assert len(received) > 0
    assert received["grade"].notna().all(), "a sheet was left empty"


def test_a_student_who_submitted_twice_gets_both_sheets_marked(cw1):
    """A resubmission is two folders and two sheets, and it is a normal
    state -- nothing renames them until somebody decides which counts.
    Marking one leaves catch_grades reading an empty cell for a student who
    was marked."""
    sheets = feedback_sheets(cw1)
    resubmitted = {i: p for i, p in sheets.items() if len(p) > 1}
    assert resubmitted, "the fake module is meant to include a resubmission"

    result = simulate_marking(cw1, seed=1)

    for identifier in resubmitted:
        assert len(result.sheets[identifier]) == len(sheets[identifier])

    received = catch_grades(cw1.submissions_path, cw1.grade_cell)
    assert received["grade"].notna().all()


def test_marks_are_on_the_assessments_own_scale(cw1):
    result = simulate_marking(cw1, seed=1)

    assert all(0 <= mark <= cw1.marks_out_of for mark in result.marks.values())


def test_marks_land_on_band_edges_and_awkward_halves(cw1):
    """A cohort with no exact halves in it never tests the rounding rule,
    and one with no band edges never tests the bands."""
    result = simulate_marking(cw1, seed=1, boundaries=4)

    planted = {m for m in result.marks.values() if m in BOUNDARY_MARKS}
    assert len(planted) >= 3


def test_a_seed_makes_the_cohort_reproducible(cw1):
    first = simulate_marking(cw1, seed=7, overwrite=True).marks
    second = simulate_marking(cw1, seed=7, overwrite=True).marks

    assert first == second


def test_marks_can_be_handed_in(cw1):
    """For a test that wants a specific student on a specific mark."""
    given = {i: 61.0 for i in feedback_sheets(cw1)}

    result = simulate_marking(cw1, marks=given)

    assert set(result.marks.values()) == {61.0}


# ---------------------------------------------------------------------------
# The grader workbooks -- the second record
# ---------------------------------------------------------------------------


def test_the_marks_reach_the_grader_workbooks(cw1):
    result = simulate_marking(cw1, seed=1)

    assert set(result.workbooks) == set(_graders(cw1))
    _, reported = _both_records(cw1)
    marked = reported[reported["Mark"].notna()]
    assert len(marked) == len(result.marks)


def test_a_mark_the_grader_already_wrote_is_left_alone(cw1):
    """Filling a column that is already filled would undo real work."""
    path = cw1.grading_output_path / f"{_graders(cw1)[0]}.xlsx"
    frame = pd.read_excel(path, dtype={"Student ID": str})
    frame.loc[0, "Mark"] = 99.0
    kept = frame.loc[0, "Student ID"]
    frame.to_excel(path, index=False)

    simulate_marking(cw1, seed=1)

    after = pd.read_excel(path, dtype={"Student ID": str})
    assert after.loc[after["Student ID"] == kept, "Mark"].item() == 99.0


# ---------------------------------------------------------------------------
# The two records, and making them disagree
# ---------------------------------------------------------------------------


def test_with_no_discrepancies_the_two_records_agree(cw1):
    """The baseline. A simulator whose records never agree is no use for
    testing the happy path."""
    simulate_marking(cw1, seed=1)

    received, reported = _both_records(cw1)
    reconciliation = reconcile_marks(received, reported)

    assert reconciliation.disagreements.empty or set(
        reconciliation.disagreements["_merge"]
    ) <= {"right_only"}, "only non-submitters should differ"


def test_a_planted_discrepancy_is_what_reconciliation_catches(cw1):
    """The mistyped copy. This is the whole reason step 7 exists, and
    without a way to produce one it can never be seen working."""
    result = simulate_marking(cw1, seed=1, discrepancies=2)

    assert len(result.discrepancies) == 2

    received, reported = _both_records(cw1)
    caught = reconcile_marks(received, reported)
    both = caught.disagreements[caught.disagreements["_merge"] == "both"]

    assert set(both["Student ID"]) == set(result.discrepancies)
    for _, row in both.iterrows():
        sheet, workbook = result.discrepancies[row["Student ID"]]
        assert (row["grade"], row["Mark"]) == (sheet, workbook)


def test_the_sheet_keeps_the_mark_the_student_was_given(cw1):
    """The slip is in the copy, not in the marking. The feedback sheet is
    what the student received and has to stay that."""
    result = simulate_marking(cw1, seed=1, discrepancies=2)

    received = catch_grades(cw1.submissions_path, cw1.grade_cell)
    for identifier, (sheet, _) in result.discrepancies.items():
        assert received.loc[
            received["Student ID"] == identifier, "grade"
        ].iloc[0] == sheet


def test_the_mistyped_students_are_not_the_boundary_students(cw1):
    """Two Random(seed) objects make the same first choice, so drawing the
    boundary marks and choosing who mistypes one off separate streams put
    every discrepancy on a planted mark, every run."""
    seen = set()
    for seed in range(6):
        result = simulate_marking(
            cw1, seed=seed, boundaries=3, discrepancies=2, overwrite=True
        )
        seen.add(
            all(m in BOUNDARY_MARKS for m, _ in result.discrepancies.values())
        )

    assert False in seen, "the mistyped students track the boundary students"


def test_more_discrepancies_than_students_is_refused(cw1):
    with pytest.raises(ValueError, match="only"):
        simulate_marking(cw1, seed=1, discrepancies=500)


# ---------------------------------------------------------------------------
# Not damaging a module by accident
# ---------------------------------------------------------------------------


def test_a_sheet_that_already_has_a_mark_is_skipped(cw1):
    simulate_marking(cw1, seed=1)
    first = catch_grades(cw1.submissions_path, cw1.grade_cell)

    again = simulate_marking(cw1, seed=99)

    assert not again.sheets, "nothing should have been rewritten"
    assert again.skipped
    second = catch_grades(cw1.submissions_path, cw1.grade_cell)
    pd.testing.assert_frame_equal(
        first.sort_values("Student ID", ignore_index=True),
        second.sort_values("Student ID", ignore_index=True),
    )


def test_overwrite_marks_them_anyway(cw1):
    simulate_marking(cw1, seed=1)

    again = simulate_marking(cw1, seed=99, overwrite=True)

    assert again.sheets
    assert not again.skipped


def test_a_dry_run_writes_nothing(cw1):
    result = simulate_marking(cw1, seed=1, dry_run=True)

    assert result.dry_run
    assert result.sheets, "it still has to say what it would do"
    received = catch_grades(cw1.submissions_path, cw1.grade_cell)
    assert received.empty or received["grade"].isna().all()


def test_an_assessment_with_no_grade_cell_is_refused(ready_to_mark):
    """There is no cell to write into, and guessing one would put a number
    somewhere nothing reads."""
    cw2 = ready_to_mark.assessment("cw2")
    cw2.grade_cell = None

    with pytest.raises(ValueError, match="no grade_cell"):
        simulate_marking(cw2, seed=1)


def test_an_assessment_with_no_sheets_says_which_step_is_missing(cw1):
    for identifier, paths in feedback_sheets(cw1).items():
        for path in paths:
            path.unlink()

    with pytest.raises(ValueError, match="distribute_feedback_sheets"):
        simulate_marking(cw1, seed=1)


# ---------------------------------------------------------------------------
# draw_marks on its own
# ---------------------------------------------------------------------------


def test_marks_are_drawn_to_the_nearest_half():
    marks = draw_marks([str(i) for i in range(40)], 100, seed=3, boundaries=0)

    assert all((m * 2) % 1 == 0 for m in marks.values())


def test_the_scale_is_the_assessments_own():
    marks = draw_marks([str(i) for i in range(40)], 50, seed=3, boundaries=0)

    assert all(0 <= m <= 50 for m in marks.values())
    assert max(marks.values()) > 20, "not a cohort scaled to 100 by mistake"


def test_a_cohort_spreads_across_bands():
    """A moderation sample stratified by band needs bands to stratify."""
    from grader_helper import make_letter_grade

    marks = draw_marks([str(i) for i in range(60)], 100, seed=5)

    assert len({make_letter_grade(m) for m in marks.values()}) >= 5


# ---------------------------------------------------------------------------
# From a terminal
# ---------------------------------------------------------------------------


def test_each_assessment_gets_its_own_cohort(ready_to_mark):
    """One seed for the whole module gives every student the same mark in
    cw1 as in cw2. The total then equals the component, and a weighting bug
    looks exactly like a correct answer."""
    from grader_helper.simulating import _seed_for

    cw1 = ready_to_mark.assessment("cw1")
    cw2 = ready_to_mark.assessment("cw2")

    assert _seed_for(1, cw1) != _seed_for(1, cw2)
    # Deterministic across processes: hash() is salted per run, crc32 is not.
    assert _seed_for(1, cw1) == _seed_for(1, cw1)
    assert _seed_for(None, cw1) is None


def test_the_command_line_writes_nothing_without_write(ready_to_mark, capsys):
    from grader_helper.simulating import main

    assert main([str(ready_to_mark.root), "-a", "cw1", "--seed", "1"]) == 0

    assert "Nothing was written" in capsys.readouterr().out
    cw1 = ready_to_mark.assessment("cw1")
    received = catch_grades(cw1.submissions_path, cw1.grade_cell)
    assert received.empty or received["grade"].isna().all()


def test_the_command_line_marks_the_module(ready_to_mark, capsys):
    from grader_helper.simulating import main

    assert main([str(ready_to_mark.root), "--seed", "1", "--write"]) == 0

    out = capsys.readouterr().out
    assert "Nothing was written" not in out
    for assessment_id in ("cw1", "cw2"):
        assessment = ready_to_mark.assessment(assessment_id)
        received = catch_grades(assessment.submissions_path, assessment.grade_cell)
        assert received["grade"].notna().all(), assessment_id


def test_an_assessment_nobody_marks_by_hand_is_passed_over(ready_to_mark):
    """No grade_cell means no cell a mark is written into and read back
    from -- a quiz, or an MCQ collected straight out of Brightspace. There
    is no feedback sheet to simulate, so it is not an error, it is skipped.
    """
    from grader_helper.simulating import _assessments_to_mark

    assert "mcq" in [a.id for a in _assessments_to_mark(ready_to_mark, None)]

    # This fake module's MCQ is marked on a sheet. Take its cell away and it
    # becomes the collected kind.
    ready_to_mark.assessment("mcq").grade_cell = None

    assert [a.id for a in _assessments_to_mark(ready_to_mark, None)] == ["cw1", "cw2"]


# ---------------------------------------------------------------------------
# Pointing at folders, with no module.toml anywhere
# ---------------------------------------------------------------------------
#
# A folder extracted from a past module is a download and a set of grader
# workbooks sitting where somebody put them. There is no module.toml to load
# and no assessment layout to resolve paths against, so the same run has to
# be reachable from the paths alone.


@pytest.fixture
def loose(tmp_path, ready_to_mark):
    """The same files, rearranged into folders named by a human."""
    import shutil

    cw1 = ready_to_mark.assessment("cw1")
    folder = tmp_path / "PS4001 CW1"
    folder.mkdir()
    shutil.copytree(cw1.submissions_path, folder / "Downloads")
    shutil.copytree(cw1.grading_output_path, folder / "Grader sheets")

    assert not list(folder.rglob("module.toml"))
    return folder


def test_a_download_can_be_marked_from_its_path_alone(loose):
    from grader_helper.simulating import simulate_marking_in

    result = simulate_marking_in(loose / "Downloads", "D30", seed=1)

    assert result.sheets
    received = catch_grades(loose / "Downloads", "D30")
    assert received["grade"].notna().all()


def test_without_workbooks_there_is_only_one_record(loose):
    """Marking the sheets alone is half the job -- reconciliation then has
    nothing to compare, which is worth being obvious about."""
    from grader_helper.simulating import simulate_marking_in

    result = simulate_marking_in(loose / "Downloads", "D30", seed=1)

    assert result.workbooks == {}


def test_the_graders_are_read_off_the_folder(loose):
    """Nothing says who marks this, so the workbooks in the folder do."""
    from grader_helper.simulating import simulate_marking_in

    result = simulate_marking_in(
        loose / "Downloads", "D30", workbooks=loose / "Grader sheets", seed=1
    )

    assert set(result.workbooks) == {"KOM", "SOB"}


def test_named_graders_narrow_it(loose):
    from grader_helper.simulating import simulate_marking_in

    result = simulate_marking_in(
        loose / "Downloads", "D30",
        workbooks=loose / "Grader sheets", graders=["KOM"], seed=1,
    )

    assert set(result.workbooks) == {"KOM"}


def test_what_the_tool_wrote_is_not_a_graders_workbook(loose):
    """`grading_output` holds the collated file and the group membership as
    well, and neither is somebody's marking."""
    from grader_helper.simulating import grader_workbooks

    books = loose / "Grader sheets"
    (books / "completed_grades.xlsx").write_bytes(b"")
    (books / "distributed.xlsx").write_bytes(b"")
    (books / "~$KOM.xlsx").write_bytes(b"")

    assert set(grader_workbooks(books)) == {"KOM", "SOB"}


def test_the_scale_is_given_when_no_module_says_it(loose):
    from grader_helper.simulating import simulate_marking_in

    result = simulate_marking_in(
        loose / "Downloads", "D30", marks_out_of=50, seed=1, boundaries=0
    )

    assert all(0 <= m <= 50 for m in result.marks.values())


def test_the_two_records_still_disagree_on_purpose(loose):
    from grader_helper.simulating import simulate_marking_in

    result = simulate_marking_in(
        loose / "Downloads", "D30",
        workbooks=loose / "Grader sheets", seed=1, discrepancies=2,
    )

    received = catch_grades(loose / "Downloads", "D30")
    reported = ingest_completed_graderfiles(
        loose / "Grader sheets", ["KOM", "SOB"], file_type="excel"
    )
    both = reconcile_marks(received, reported)
    caught = both.disagreements[both.disagreements["_merge"] == "both"]

    assert set(caught["Student ID"]) == set(result.discrepancies)


def test_the_command_line_takes_folders(loose, capsys):
    from grader_helper.simulating import main

    assert main([
        "-s", str(loose / "Downloads"),
        "-w", str(loose / "Grader sheets"),
        "-c", "D30", "--seed", "1", "--write",
    ]) == 0

    assert catch_grades(loose / "Downloads", "D30")["grade"].notna().all()


def test_submissions_without_a_cell_is_refused(loose):
    """The one thing about a feedback sheet that cannot be guessed from it."""
    from grader_helper.simulating import main

    with pytest.raises(SystemExit):
        main(["-s", str(loose / "Downloads")])


def test_a_module_and_folders_together_is_refused(ready_to_mark, loose):
    """Two answers to where the submissions are."""
    from grader_helper.simulating import main

    with pytest.raises(SystemExit):
        main([
            str(ready_to_mark.root),
            "-s", str(loose / "Downloads"), "-c", "D30",
        ])


def test_neither_a_module_nor_folders_is_refused():
    from grader_helper.simulating import main

    with pytest.raises(SystemExit):
        main([])


# ---------------------------------------------------------------------------
# A sheet that will not take a mark
# ---------------------------------------------------------------------------
#
# A run over a real cohort writes eighty-odd files. Dying on the fortieth
# leaves half of them marked with no record of which half, which is worse
# than any one sheet failing.


def _merge_the_grade_cell(path, cell="D30"):
    """Merge the grade cell into a range it is not the top-left of, which is
    what a rubric that spreads the total across two columns does."""
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    column, row = cell[0], cell[1:]
    workbook.worksheets[0].merge_cells(f"{chr(ord(column) - 1)}{row}:{cell}")
    workbook.save(path)


def test_a_merged_grade_cell_is_reported_not_raised(cw1):
    sheets = feedback_sheets(cw1)
    first = next(iter(sheets.values()))[0]
    _merge_the_grade_cell(first, cw1.grade_cell)

    result = simulate_marking(cw1, seed=1)

    assert list(result.refused) == [first]
    assert "merged" in result.refused[first]
    assert "would not take a mark" in str(result)


def test_the_other_sheets_are_still_marked(cw1):
    """The point of collecting rather than raising."""
    sheets = feedback_sheets(cw1)
    first = next(iter(sheets.values()))[0]
    _merge_the_grade_cell(first, cw1.grade_cell)

    result = simulate_marking(cw1, seed=1)

    written = sum(len(paths) for paths in result.sheets.values())
    assert written == sum(len(p) for p in sheets.values()) - 1


def test_a_mark_that_never_reached_a_sheet_is_not_reported_by_the_grader(cw1):
    """A grader does not copy a mark they never wrote. Left in, the student
    would reconcile as a disagreement -- an empty sheet against a workbook
    mark -- which is a fault this tool invented, not one it found."""
    sheets = feedback_sheets(cw1)
    identifier, paths = next(iter(sheets.items()))
    _merge_the_grade_cell(paths[0], cw1.grade_cell)

    simulate_marking(cw1, seed=1)

    _, reported = _both_records(cw1)
    theirs = reported.loc[reported["Student ID"] == identifier, "Mark"]
    assert theirs.isna().all(), "a mark was reported for an unmarked sheet"


def test_a_clean_run_refuses_nothing(cw1):
    result = simulate_marking(cw1, seed=1)

    assert result.refused == {}


# ---------------------------------------------------------------------------
# A rubric that calculates its own total
# ---------------------------------------------------------------------------
#
# Every sheet in this suite had an EMPTY grade cell, because that is what
# fake_module writes. A real rubric calculates the total from criterion
# cells, so the moment Excel saves it the grade cell caches a result --
# usually 0, the sum of criteria nobody has filled in. "Holds a number"
# then means every distributed sheet looks marked, and a run that means to
# fill them all in skips every one of them while still filling the grader
# workbooks.


def _make_rubric_calculate(module, assessment_id="cw1", total=0):
    """Give the blank rubric and every distributed sheet a cached total."""
    from openpyxl import load_workbook

    assessment = module.assessment(assessment_id)
    targets = [assessment.rubric_path]
    for paths in feedback_sheets(assessment).values():
        targets.extend(paths)
    for path in targets:
        workbook = load_workbook(path)
        workbook.worksheets[0][assessment.grade_cell] = total
        workbook.save(path)
    return assessment


def test_a_calculated_zero_is_not_a_mark(ready_to_mark):
    """The one that stopped it working on a real module."""
    cw1 = _make_rubric_calculate(ready_to_mark)

    result = simulate_marking(cw1, seed=1)

    assert result.sheets, "every sheet was skipped as already marked"
    assert not result.skipped
    received = catch_grades(cw1.submissions_path, cw1.grade_cell)
    assert (received["grade"] != 0).any(), "the sheets still read as blank"


def test_a_mark_a_grader_wrote_is_still_left_alone(ready_to_mark):
    """The rule it must not break: a sheet showing something OTHER than the
    blank one is somebody's marking."""
    cw1 = _make_rubric_calculate(ready_to_mark)
    marked = next(iter(feedback_sheets(cw1).values()))[0]
    from openpyxl import load_workbook

    workbook = load_workbook(marked)
    workbook.worksheets[0][cw1.grade_cell] = 63.0
    workbook.save(marked)

    result = simulate_marking(cw1, seed=1)

    assert marked not in [p for ps in result.sheets.values() for p in ps]
    assert grade_cell_value(marked, cw1.grade_cell) == 63.0


def test_the_grader_copies_what_is_on_the_sheet_not_what_was_drawn(
    ready_to_mark,
):
    """A skipped sheet keeps the mark it already had, and that is the mark
    the grader copies. Reporting the drawn one instead would reconcile as a
    disagreement the simulator invented."""
    cw1 = _make_rubric_calculate(ready_to_mark)
    identifier, paths = next(iter(feedback_sheets(cw1).items()))
    from openpyxl import load_workbook

    workbook = load_workbook(paths[0])
    workbook.worksheets[0][cw1.grade_cell] = 63.0
    workbook.save(paths[0])

    simulate_marking(cw1, seed=1)

    _, reported = _both_records(cw1)
    theirs = reported.loc[reported["Student ID"] == identifier, "Mark"]
    assert theirs.iloc[0] == 63.0


def test_without_a_blank_any_number_still_counts(ready_to_mark):
    """Pointed at folders with no rubric named, the old rule stands -- and
    is right for a sheet this package wrote, whose grade cell is empty."""
    from grader_helper.simulating import simulate_marking_in

    cw1 = _make_rubric_calculate(ready_to_mark)

    result = simulate_marking_in(cw1.submissions_path, cw1.grade_cell, seed=1)

    assert not result.sheets
    assert result.skipped


# ---------------------------------------------------------------------------
# Saying what it sees
# ---------------------------------------------------------------------------
#
# Four separate guesses about why a real module's sheets were not being
# marked all turned out to be about what was IN the grade cell -- which
# nobody can see from here and everybody can see from there.


def test_explain_lists_every_sheet_and_its_decision(cw1):
    from grader_helper.simulating import explain_sheets

    frame = explain_sheets(cw1.submissions_path, cw1.grade_cell, cw1.rubric_path)

    assert len(frame) == sum(len(p) for p in feedback_sheets(cw1).values())
    assert set(frame["decision"]) == {"write"}
    assert list(frame.columns) == [
        "identifier", "sheet", "folder", "value", "blank", "decision"
    ]


def test_explain_shows_why_a_sheet_would_be_skipped(ready_to_mark):
    from grader_helper.simulating import explain_sheets

    cw1 = ready_to_mark.assessment("cw1")
    marked = next(iter(feedback_sheets(cw1).values()))[0]
    from openpyxl import load_workbook

    workbook = load_workbook(marked)
    workbook.worksheets[0][cw1.grade_cell] = 63.0
    workbook.save(marked)

    frame = explain_sheets(cw1.submissions_path, cw1.grade_cell, cw1.rubric_path)
    row = frame.loc[frame["sheet"] == marked.name].iloc[0]

    assert row["value"] == 63.0
    assert "already marked" in row["decision"]


def test_explain_says_nothing_found_rather_than_raising(tmp_path):
    """It is a diagnostic. Refusing to run is the opposite of its job."""
    from grader_helper.simulating import explain_sheets

    empty = tmp_path / "submissions"
    empty.mkdir()

    assert explain_sheets(empty, "D30").empty


def test_a_macro_enabled_sheet_is_found(cw1):
    """catch_grades reads .xlsm, .xlsb and .xls as well. A rubric saved as
    one was findable by the reader and invisible to the writer."""
    import shutil

    original = next(iter(feedback_sheets(cw1).values()))[0]
    macro = original.with_suffix(".xlsm")
    shutil.copy2(original, macro)
    original.unlink()

    found = feedback_sheets(cw1)

    assert macro in [p for paths in found.values() for p in paths]


@pytest.mark.parametrize(
    "stem",
    ["Feedback sheet 24439711", "feedback sheet 24439711",
     "FEEDBACK SHEET 24439711"],
)
def test_the_prefix_match_ignores_case(stem):
    """The constant is lower case and the filename is not, which reads like
    a bug every time this file is scanned. It is not one -- the stem is
    lower-cased before the comparison -- and this is what says so."""
    from grader_helper.simulating import _identifier

    assert _identifier(stem) == "24439711"


def test_the_identifier_keeps_its_own_case():
    from grader_helper.simulating import _identifier

    assert _identifier("Feedback sheet Team 3") == "Team 3"


def test_explain_and_write_together_are_refused(ready_to_mark):
    """Leaving --explain on a --write command was a silent no-op that
    printed a report looking exactly like a successful run. That is the one
    way to conclude the writer is broken when it has not run at all."""
    from grader_helper.simulating import main

    with pytest.raises(SystemExit):
        main([str(ready_to_mark.root), "-a", "cw1", "--explain", "--write"])

    cw1 = ready_to_mark.assessment("cw1")
    received = catch_grades(cw1.submissions_path, cw1.grade_cell)
    assert received.empty or received["grade"].isna().all()
