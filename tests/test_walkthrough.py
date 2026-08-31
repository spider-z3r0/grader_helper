#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""The walkthrough notebook must actually run.

`docs/running-locally.md` has described the walkthrough as "already verified
end to end" since it was written, and that was true when someone last ran
it by hand. This makes it a fact the suite checks: `marimo.App.run()`
executes every cell and raises whatever a cell raises, so a library change
that breaks the notebook fails here rather than in front of a colleague.

The notebook is driven into `tmp_path` through `GRADER_HELPER_SCRATCH`,
which exists for exactly this. One test, not several, because running it
builds a whole module on disk and there is no reason to do that twice.

House convention: ``pl`` is pathlib, ``pr`` is polars.
"""

import importlib.util
import sys

import pytest

pytest.importorskip("marimo", reason="marimo is a dev dependency")


@pytest.fixture
def walkthrough(tmp_path, monkeypatch, repo_root):
    """Run every cell of the notebook, and return what it defined."""
    monkeypatch.setenv("GRADER_HELPER_SCRATCH", str(tmp_path))
    # The notebook does sys.path.insert(0, "tests"), which is relative to the
    # working directory -- the same requirement running-locally.md gives a
    # reader ("run marimo edit from the repo root").
    monkeypatch.chdir(repo_root)
    # And the repo root ahead of whatever else is on the path, which is what
    # running from there gives you.
    monkeypatch.syspath_prepend(str(repo_root))

    # By the time any test runs, `grader_helper` is already in sys.modules,
    # and in a bare `pytest tests/test_walkthrough.py` it is bound to the
    # repo directory's shim __init__.py rather than to the package -- the
    # repo directory is itself called grader_helper, and its parent is on
    # sys.path. The shim star-re-exports the public API, so
    # `from grader_helper import catch_grades` still works and the breakage
    # is invisible until something imports a *submodule*: fake_module's
    # `from grader_helper.dataframe_operations import ...`, which the shim
    # has no path for.
    #
    # Dropping the binding makes the notebook import the real package, the
    # way it does for someone running marimo from the repo root. delitem
    # rather than `del` so it is put back afterwards and the rest of the
    # session is unaffected. test_import.py guards the same hazard, but in a
    # subprocess, which is why it does not catch this one.
    for name in [
        n for n in sys.modules
        if n == "grader_helper" or n.startswith("grader_helper.")
    ]:
        monkeypatch.delitem(sys.modules, name)

    path = repo_root / "notebooks" / "grading_walkthrough.py"
    spec = importlib.util.spec_from_file_location("grading_walkthrough", path)
    notebook = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(notebook)

    _, definitions = notebook.app.run()
    return definitions


def test_the_walkthrough_runs_and_collects_the_quizzes(walkthrough):
    """One assertion per thing a reader is being shown.

    The marks are the load-bearing one: `fake_module` builds the eleven
    exports backwards from the mark each student should end up with, so
    matching that column means the notebook read the exports, applied the
    module's own rules and reconciled against the class list -- not merely
    that it ran without raising.
    """
    fake = walkthrough["FAKE"]
    quizzes = walkthrough["A3"]
    marks = walkthrough["quiz_marks"]

    # The module built with quizzes rather than the MCQ.
    assert [a.id for a in walkthrough["MODULE"].assessments] == [
        "cw1", "cw2", "quizzes",
    ]
    assert quizzes.raw_column == "Quizzes (10)"

    # Read off the file, not passed in by the notebook.
    assert (quizzes.pass_mark, quizzes.free_passes) == (80.0, 1)

    collected = marks.set_index("Student ID")["Quizzes (10)"]
    assert collected.to_dict() == (
        fake.expected.set_index("Student ID")["mcq"].to_dict()
    )

    # The two edges the notebook shows by name.
    assert collected["23304309"] == 0, "the non-participant must not be given a free pass"
    assert collected["23304308"] == 10, "nine passes plus the free pass, capped"

    # The module, not just the assessments: the notebook now ends by
    # collating all three and banding the totals.
    #
    # Not compared against fake.expected -- the notebook's own grader cells
    # write synthetic marks into the feedback sheets, so expected is the
    # truth for the quiz column and nothing else. What is checked instead is
    # internal consistency: every enrolled student reaches the sheet, the
    # departmental columns are there, the quiz marks survived the collation,
    # and every letter grade is the one its own total earns.
    graded = walkthrough["module_sheet"].set_index("Student ID")

    assert set(graded.index) == set(fake.expected["Student ID"])
    assert list(graded.columns) == [
        "Name",
        "Coursework 1 (100)", "Coursework 1 (40)",
        "Coursework 2 (100)", "Coursework 2 (50)",
        "Quizzes (10)",
        "Total % Grade", "Letter Grade",
    ]
    assert graded["Quizzes (10)"].to_dict() == collected.to_dict()

    from grader_helper import make_letter_grade

    assert graded["Letter Grade"].to_dict() == {
        student: make_letter_grade(total)
        for student, total in graded["Total % Grade"].items()
    }

    # And the progress each section records. All three, because cw1 went
    # without a record_progress cell until someone ran the notebook to the
    # end and read its own status output back.
    recorded = walkthrough["ModuleFile"].load(fake.root).module
    assert {
        a.id: a.status.grades_collected for a in recorded.assessments
    } == {"cw1": True, "cw2": True, "quizzes": True}


def test_the_walkthrough_writes_both_departmental_sheets(walkthrough):
    """The two workbooks the notebook now ends with.

    PS4001 is the template's own shape, so building it must give the
    department's file back. PS4002 is the shape it has no room for, and is
    the reason the builder exists -- so what is checked there is the two
    things a module leader would otherwise have re-pointed by hand.
    """
    from openpyxl import load_workbook

    template = load_workbook(walkthrough["TEMPLATE"])["GradeTemplate"]
    first = load_workbook(walkthrough["sheet_path"])["GradeTemplate"]
    second = load_workbook(walkthrough["sheet_2_path"])["GradeTemplate"]

    def headers(sheet):
        return [
            sheet.cell(29, column).value
            for column in range(1, sheet.max_column + 1)
            if sheet.cell(29, column).value
        ]

    # PS4001 fits the template's shape -- two 100-mark pieces and a ten-mark
    # one -- so its sheet is the department's own layout, column for column.
    # Only the third assessment's *name* differs, because the walkthrough
    # builds the module with weekly quizzes in the MCQ's slot.
    assert len(headers(first)) == len(headers(template))
    assert headers(first)[:6] == headers(template)[:6]
    assert headers(first)[7:] == headers(template)[7:]
    assert headers(first)[6] == "Quizzes (10)"
    assert first["D30"].value == template["D30"].value
    assert first["F30"].value == template["F30"].value
    assert first["H30"].value == template["H30"].value
    assert first["I30"].value == template["I30"].value
    # ...with the marks in and the samples gone.
    assert first["A30"].value != "Sample1"
    assert first["C30"].value is not None

    # PS4002 does not fit. Six assessment columns where the template has five,
    # and the two MCQs are on different scales -- out of 100 and out of 10 --
    # because both happen in practice.
    assert headers(second) == [
        "Name", "Student ID",
        "Coursework 1 (100)", "Coursework 1 (30)",
        "MCQ 1 (100)", "MCQ 1 (35)",
        "MCQ 2 (10)", "MCQ 2 (35)",
        "Total % Grade", "Letter Grade", "Comments",
    ]

    # One weighting scales down, the other scales up, and neither simplifies
    # to a single divisor.
    assert second["F30"].value == "=E30/100*35"
    assert second["H30"].value == "=G30/10*35"

    # Every assessment reaches the total -- the failure the builder exists to
    # prevent is one of them quietly missing from it.
    assert second["I30"].value == "=ROUND(SUM(D30,F30,H30),0)"

    # The letter grade reads the total where it actually is, and the
    # distribution reads the letter grade where it actually is.
    assert second["J30"].value.startswith("=IF(ROUND(I30,2)>0,")
    assert second["H6"].value == '=COUNTIF(J30:J530,"A1")'

    # The descriptives cover every column, which is the A23 block the notes
    # name as the thing that goes wrong by hand.
    for column in range(3, 10):  # C..I: the assessment block plus the total
        assert second.cell(23, column).value is not None, column
        assert second.cell(25, column).value is not None, column
    assert second.cell(23, 10).value is None, "nothing spills into Letter Grade"


def test_the_second_module_totals_what_the_fixture_expects(walkthrough):
    """PS4002's marks survive collation, weighting and banding.

    The MCQs are handed in through `marks=` rather than read off disk, and
    they are marked out of 10 while being worth 35 -- a scale-up, which is
    the case that silently lost its weighted column before
    `collate_module_marks` stopped inferring the column's name.
    """
    second = walkthrough["SECOND"]
    sheet = walkthrough["module_2_sheet"].set_index("Student ID")
    expected = second.expected.set_index("Student ID")

    assert set(sheet.index) == set(expected.index)
    assert sheet["Total % Grade"].to_dict() == expected["Total % Grade"].to_dict()
    assert sheet["Letter Grade"].to_dict() == expected["Letter Grade"].to_dict()

    # Both weighted columns exist: one scaling down, one scaling up.
    assert sheet["MCQ 1 (35)"].to_dict() == (sheet["MCQ 1 (100)"] * 0.35).to_dict()
    assert sheet["MCQ 2 (35)"].to_dict() == (sheet["MCQ 2 (10)"] * 3.5).to_dict()

    # Joyce sat nothing and scored nothing, in this module as in the other.
    assert sheet.loc["23304309", "Letter Grade"] == "NG"


def test_the_third_module_collates_from_three_sources(walkthrough):
    """PS4003's four assessments arrive three different ways, in one call.

    The coursework is read off its feedback sheets, the quizzes out of
    Brightspace's exports, and the MCQ and exam are handed in. Only the last
    two are passed to `collate_module_marks`; the other two it has to find.
    `collate_module_marks` chooses per assessment, by asking what each one
    *has*, and a module with a single source is no evidence that it does.
    """
    third = walkthrough["THIRD"]
    module = walkthrough["MODULE_3"]
    marks = walkthrough["module_3_marks"].set_index("Student ID")
    expected = third.expected.set_index("Student ID")

    assert [a.id for a in module.assessments] == ["cw1", "quizzes", "mcq", "exam"]
    assert len(third.quiz_exports["quizzes"]) == 10, "ten quizzes for ten marks"

    # Ten quizzes and no free pass, so the mark is the number passed. PS4001
    # sets eleven and forgives one; both are read off module.toml.
    quizzes = module.assessment("quizzes")
    assert (quizzes.marks_out_of, quizzes.free_passes) == (10, 0)
    assert quizzes.weighted_column is None, "ten marks worth ten need no weighting"

    # Neither of these was handed in.
    assert marks["Coursework 1 (100)"].dropna().to_dict() == (
        expected["cw1"].dropna().to_dict()
    )
    assert marks["Quizzes (10)"].to_dict() == expected["quizzes"].to_dict()

    sheet = walkthrough["module_3_sheet"].set_index("Student ID")
    assert sheet["Total % Grade"].to_dict() == expected["Total % Grade"].to_dict()
    assert sheet["Letter Grade"].to_dict() == expected["Letter Grade"].to_dict()
    assert sheet.loc["23304309", "Letter Grade"] == "NG"


def test_the_third_module_sheet_carries_every_component(walkthrough):
    """Seven assessment columns, and the raw one still reaches the total."""
    from openpyxl import load_workbook

    worksheet = load_workbook(walkthrough["sheet_3_path"])["GradeTemplate"]
    headers = [
        worksheet.cell(29, column).value
        for column in range(1, worksheet.max_column + 1)
        if worksheet.cell(29, column).value
    ]

    assert headers == [
        "Name", "Student ID",
        "Coursework 1 (100)", "Coursework 1 (30)",
        "Quizzes (10)",
        "MCQ (100)", "MCQ (20)",
        "Exam (100)", "Exam (40)",
        "Total % Grade", "Letter Grade", "Comments",
    ]
    # E30 is the quizzes' raw column and reaches the total directly, between
    # two weighted ones.
    assert worksheet["J30"].value == "=ROUND(SUM(D30,E30,G30,I30),0)"
    assert worksheet["G30"].value == "=F30/5", "100 worth 20 divides exactly"
    assert worksheet["K30"].value.startswith("=IF(ROUND(J30,2)>0,")
    assert worksheet["H6"].value == '=COUNTIF(K30:K530,"A1")'


def test_the_walkthrough_moderates_the_third_module(walkthrough):
    """The pack the notebook builds, and the seed that justifies it.

    The load-bearing assertion is the last one. A draw nobody can reproduce
    cannot answer "why was this student moderated?", and the manifest is
    what carries the answer.
    """
    moderation = walkthrough["moderation"]
    pack = walkthrough["pack"]

    # One per band, no non-participants, and the borderline cases as well.
    bands = set(moderation.selected["Letter Grade"])
    assert "NG" not in bands, "a student who submitted nothing has nothing to moderate"
    assert bands, "the draw selected nobody"

    reasons = " ".join(moderation.selected["Selected Because"])
    assert "drawn" in reasons
    assert "borderline" in reasons, "the notebook asks for borderline='include'"

    # Only the coursework has anything to copy; the other three assessments
    # have no submissions folder and are skipped rather than failing.
    assert set(pack.copied) == {"cw1"}
    assert pack.manifest.is_file()

    from grader_helper import read_moderation_manifest, sample_for_moderation

    manifest = read_moderation_manifest(pack.root)
    assert len(manifest) == len(moderation.selected)
    assert set(manifest["Seed"]) == {moderation.seed}

    # And the recorded seed really does reproduce the draw.
    again = sample_for_moderation(
        walkthrough["module_3_sheet"],
        n=1,
        borderline="include",
        seed=int(manifest["Seed"].iloc[0]),
    )
    assert again.selected["Student ID"].tolist() == (
        moderation.selected["Student ID"].tolist()
    )


def test_a_moderation_pack_spans_every_marked_assessment(walkthrough):
    """PS4001 has two marked courseworks, and both reach the moderator.

    PS4003 has only one assessment with a download, so it cannot show this --
    and a pack that quietly held one assessment's work when the module has two
    would look complete. The junk a real download carries must not reach the
    moderator either.
    """
    pack = walkthrough["ps4001_pack"]
    sample = walkthrough["ps4001_sample"]

    assert set(pack.copied) == {"cw1", "cw2"}, (
        "both marked courseworks must be copied; the quizzes have no "
        "submission folders and are rightly absent"
    )
    assert pack.copied["cw1"] == pack.copied["cw2"]

    # Each band folder holds a sub-folder per assessment, named for it.
    for _, student in sample.selected.iterrows():
        band = pack.root / str(student["Letter Grade"])
        assert band.is_dir()
    assessments = {
        path.name
        for path in pack.root.rglob("*")
        if path.is_dir() and path.parent.parent == pack.root
    }
    assert assessments == {"Coursework 1", "Coursework 2"}

    # A band whose sampled student submitted nothing still appears, with a
    # note saying so rather than as an empty folder or no folder at all.
    for student, band in zip(
        sample.selected["Student ID"], sample.selected["Letter Grade"]
    ):
        assert (pack.root / str(band)).is_dir(), f"band {band} missing from the pack"
    unserved = {student for student, _ in pack.missing}
    for student in unserved:
        assert list(pack.root.rglob(f"NOTHING SUBMITTED - {student}.txt")), (
            f"{student} was sampled, submitted nothing, and the pack does not say so"
        )

    # The fixture's download carries a __MACOSX folder and an index.html,
    # which is what a real one looks like. Neither parses as a submission.
    copied_folders = [
        path.name
        for path in pack.root.rglob("*")
        if path.is_dir() and path.parent.name.startswith("Coursework")
    ]
    assert copied_folders, "nothing was copied at all"
    assert not any("MACOSX" in name for name in copied_folders)
    assert all(" - " in name for name in copied_folders)
