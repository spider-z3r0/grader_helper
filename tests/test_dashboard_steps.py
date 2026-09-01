#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""Driving a whole assessment from the app's own buttons.

The dashboard's steps are functions rather than code inside a button guard
precisely so this can exist: a test cannot click, but it can call exactly
what a click calls. So this runs a coursework the way a module leader
does -- allocate, distribute, wait for the marking, collect and reconcile,
rename the folders back -- against a real module on disk, and then collects
a term of quizzes.

What it is really checking is that the app is not merely a page that renders.
Each step has to leave the right artefacts behind, and the status it records
has to survive a reload, because the leader closes the page between the
distribution and the marks coming back.

House convention: ``pl`` is pathlib, ``pr`` is polars.
"""

import importlib.util
import pathlib as pl
import shutil
import sys

import pandas as pd
import pytest

pytest.importorskip("marimo", reason="marimo is a dev dependency")
from openpyxl import load_workbook  # noqa: E402

from grader_helper import extract_studentid_grade  # noqa: E402
from grader_helper.models import ModuleFile  # noqa: E402

sys.path.insert(0, "tests")


@pytest.fixture
def module_on_disk(tmp_path, repo_root):
    """A whole fake module, before any of the work has been done."""
    sys.path.insert(0, str(repo_root / "tests"))
    from fake_module import make_fake_module

    root = tmp_path / "PS4001"
    make_fake_module(root, distributed=False, marked=False, quizzes=True)
    return root


@pytest.fixture
def dashboard(monkeypatch, repo_root):
    """The notebook, run against a folder, returning the names it defined."""

    def _run(folder):
        monkeypatch.setenv("GRADER_HELPER_MODULE", str(folder))
        monkeypatch.chdir(repo_root)
        monkeypatch.syspath_prepend(str(repo_root))
        for name in [
            n for n in sys.modules
            if n == "grader_helper" or n.startswith("grader_helper.")
        ]:
            monkeypatch.delitem(sys.modules, name)

        path = repo_root / "notebooks" / "module_dashboard.py"
        spec = importlib.util.spec_from_file_location("module_dashboard", path)
        notebook = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(notebook)
        return notebook.app.run()[1]

    return _run


def mark_the_sheets(assessment, graders):
    """Stand in for the graders, who are the slow part of a real module.

    Two separate acts, because that is what the audit exists to compare: the
    mark is written into the feedback sheet the student receives, and then
    *copied by hand* into the grader's own workbook.
    """
    written = {}
    for index, sheet in enumerate(
        sorted(assessment.submissions_path.glob("*/Feedback sheet *.xlsx"))
    ):
        workbook = load_workbook(sheet)
        workbook.active[assessment.grade_cell] = 40 + (index * 7) % 55
        workbook.save(sheet)
        written[sheet.stem.split(" ")[-1]] = sheet

    for grader in graders:
        grader_file = assessment.grading_output_path / f"{grader}.xlsx"
        allocated = pd.read_excel(grader_file, dtype={"Student ID": str})
        marks = []
        for student in allocated["Student ID"]:
            sheet = written.get(student)
            found = (
                extract_studentid_grade(sheet, assessment.grade_cell)
                if sheet
                else None
            )
            marks.append(found[1] if found else None)
        allocated["Mark"] = marks
        allocated.to_excel(grader_file, index=False)


@pytest.fixture
def leader_managed_module(module_on_disk):
    """The same module, with cw1 turned into a leader-managed group piece.

    Edited as text rather than through `ModuleFile.save`, which only updates
    keys already in the file -- adding a key to an existing `[[assessment]]`
    is a hand edit, and this is what the hand edit looks like.
    """
    import pandas as pd

    toml = module_on_disk / "module.toml"
    text = toml.read_text(encoding="utf-8")
    text = text.replace(
        'id = "cw1"\ntype = "coursework"',
        'id = "cw1"\ntype = "coursework"\ngroup = true\n'
        'group_source = "module_leader"',
        1,
    )
    toml.write_text(text, encoding="utf-8")

    # The leader's own sheets, which is the only place these groups exist.
    ids = pd.read_excel(module_on_disk / "classlist.xlsx", dtype=str)
    ids = [str(i).replace("#", "") for i in ids["Username"]]
    sheets = module_on_disk / "assessments" / "cw1" / "groups"
    sheets.mkdir(parents=True, exist_ok=True)
    half = len(ids) // 2
    pd.DataFrame({"Student ID": ids[:half]}).to_excel(
        sheets / "Team 1.xlsx", index=False
    )
    pd.DataFrame({"Student ID": ids[half:]}).to_excel(
        sheets / "Team 2.xlsx", index=False
    )
    return module_on_disk


def test_a_leader_managed_group_module_still_reads_its_class_list(
    leader_managed_module, dashboard
):
    """Brightspace never knew about these groups, so the class list has no
    group column -- and asking for one would refuse a file that is correct."""
    names = dashboard(leader_managed_module)

    class_list = names["class_list"]
    assert class_list is not None, "the class list must still read"
    assert "Group" not in class_list.columns


def test_a_leader_managed_group_assessment_allocates_from_the_buttons(
    leader_managed_module, dashboard
):
    """The whole point of the wiring: the groups are collected, whole teams
    go to one marker, and every student still gets a row to be marked on."""
    names = dashboard(leader_managed_module)
    cw1 = names["found"].module.assessment("cw1")

    master, workbooks, allocation = names["allocate_marking"](
        cw1, names["class_list"]
    )

    assert cw1.group_membership_path.exists()
    assert (allocation.groupby("Group")["grader"].nunique() == 1).all()

    written = pd.concat(
        pd.read_excel(path, dtype={"Student ID": str})
        for path in workbooks.values()
    )
    assert len(written) == master.students
    assert "Student ID" in written.columns

    assert ModuleFile.load(leader_managed_module).module.assessment(
        "cw1"
    ).status.graders_allocated, "the flag has to survive a reload of the file"


def test_a_coursework_runs_from_the_buttons(module_on_disk, dashboard):
    """Allocate, distribute, mark, collect, reconcile, rename back."""
    names = dashboard(module_on_disk)
    module = names["found"].module
    cw1 = module.assessment("cw1")
    class_list = names["class_list"]
    graders = [g.initials for g in cw1.graders]

    assert class_list is not None, "the class list must be read off module.toml"

    # --- 1. allocate ------------------------------------------------------
    master, workbooks, allocation = names["allocate_marking"](cw1, class_list)

    assert (cw1.folder_path / "distributed.xlsx").exists()
    assert sorted(p.stem for p in workbooks.values()) == sorted(graders)
    assert set(allocation["grader"]) <= set(graders)
    assert ModuleFile.load(module_on_disk).module.assessment(
        "cw1"
    ).status.graders_allocated, "the flag has to survive a reload of the file"

    # --- 1a. one student submitted twice ----------------------------------
    # Real cohorts do this, and nothing downstream will run until it is
    # resolved: two folders for one student cannot both be renamed for
    # marking. Which attempt counts is the leader's call, so the app asks.
    plan = names["resolve_resubmissions"](cw1, "latest")

    assert plan.removed, "the fake module has a resubmitter; it should be found"
    assert all(folder.exists() for folder in plan.removed), "a plan deletes nothing"

    names["resolve_resubmissions"](cw1, "latest", apply=True)

    assert not any(folder.exists() for folder in plan.removed)

    # --- 2. distribute ----------------------------------------------------
    distribution, log = names["distribute_sheets"](cw1, class_list)

    assert (cw1.submissions_path / "folder_rename_log.csv").exists()
    assert len(log) > 0
    assert list(cw1.submissions_path.glob("*/Feedback sheet *.xlsx"))

    # The fixture ships a `__MACOSX` folder, because unzipping on a Mac makes
    # one and a real download has whatever the leader's machine put there.
    # It is genuinely unrecognised, so the flag is withheld: a run that
    # matched every student but one has not finished, and the tick would hide
    # the one it missed.
    assert distribution.unmatched == ["__MACOSX"]
    assert not ModuleFile.load(module_on_disk).module.assessment(
        "cw1"
    ).status.sheets_distributed, "an unmatched folder must withhold the flag"

    # Clear it the way a leader would, and the same step records.
    shutil.rmtree(cw1.submissions_path / "__MACOSX")
    distribution, log = names["distribute_sheets"](cw1, class_list)

    assert not distribution.unmatched
    assert ModuleFile.load(module_on_disk).module.assessment(
        "cw1"
    ).status.sheets_distributed

    # --- the graders do their part ----------------------------------------
    mark_the_sheets(cw1, graders)

    # --- 3. collect and reconcile ----------------------------------------
    collation, audit = names["collect_marks"](cw1)

    assert (cw1.grading_output_path / "completed_grades.xlsx").exists()

    # Not `audit.agree`: a real cohort has a student who never submitted. They
    # are on the class list, so a grader was allocated them and they reach the
    # collated file with no feedback sheet to read -- which is normal, and is
    # the whole reason the kinds are told apart rather than counted.
    assert audit.transcription_slips.empty, (
        "a mark on a student's sheet differs from the one the department "
        f"would be sent: {audit.transcription_slips}"
    )
    assert audit.not_allocated.empty
    assert len(audit.not_submitted) == len(audit.disagreements)
    assert ModuleFile.load(module_on_disk).module.assessment(
        "cw1"
    ).status.grades_collected

    # --- 4. put the folders back ------------------------------------------
    renamed = sorted(p.name for p in cw1.submissions_path.iterdir() if p.is_dir())
    restoration = names["rename_back"](cw1)
    restored = sorted(p.name for p in cw1.submissions_path.iterdir() if p.is_dir())

    assert restored != renamed, "the folders were not renamed back"
    assert set(restored) == set(log["Original Name"]), (
        "they must go back to the exact name Brightspace gave, or the "
        "re-upload does not match"
    )
    assert restoration


def test_the_quizzes_run_from_their_own_button(module_on_disk, dashboard):
    """The other path: no allocation, no sheets, nothing to reconcile."""
    names = dashboard(module_on_disk)
    quizzes = names["found"].module.assessment("quizzes")

    marks = names["collect_quizzes"](quizzes, names["class_list"])

    assert len(marks) == len(names["class_list"]), (
        "every student on the class list needs a row, including those who sat "
        "no quiz -- a missing row takes a component out of a module total"
    )
    assert marks[quizzes.raw_column].between(0, quizzes.marks_out_of).all()
    assert ModuleFile.load(module_on_disk).module.assessment(
        "quizzes"
    ).status.grades_collected


def test_a_step_refuses_to_replace_what_it_already_wrote(module_on_disk, dashboard):
    """Re-running allocation is not free: it changes who marks whom.

    So it refuses unless the page's tick says otherwise, rather than quietly
    reassigning a cohort somebody has already started marking.
    """
    names = dashboard(module_on_disk)
    cw1 = names["found"].module.assessment("cw1")
    class_list = names["class_list"]

    names["allocate_marking"](cw1, class_list)

    with pytest.raises(FileExistsError):
        names["allocate_marking"](cw1, class_list)

    names["allocate_marking"](cw1, class_list, replace=True)


def test_distribution_never_replaces_a_feedback_sheet(module_on_disk, dashboard):
    """A sheet already in a student's folder may carry a mark.

    There is no tick that changes this, so running distribution twice must
    leave the marks alone rather than handing back a blank sheet.
    """
    names = dashboard(module_on_disk)
    cw1 = names["found"].module.assessment("cw1")
    class_list = names["class_list"]

    names["allocate_marking"](cw1, class_list)
    names["resolve_resubmissions"](cw1, "latest", apply=True)
    names["distribute_sheets"](cw1, class_list)
    mark_the_sheets(cw1, [g.initials for g in cw1.graders])

    sheet = sorted(cw1.submissions_path.glob("*/Feedback sheet *.xlsx"))[0]
    before = load_workbook(sheet).active[cw1.grade_cell].value

    names["distribute_sheets"](cw1, class_list)

    assert load_workbook(sheet).active[cw1.grade_cell].value == before


@pytest.fixture
def marked_module(tmp_path, repo_root):
    """A module whose marking is already done, ready to be collated.

    The departmental template is the department's own workbook, which lives
    in this repo. A real module keeps its own copy, so this puts one in the
    module folder and names it -- by editing the file directly, because
    ModuleFile.save only ever updates keys that are already there.
    """
    sys.path.insert(0, str(repo_root / "tests"))
    from fake_module import make_fake_module

    root = tmp_path / "PS4001"
    make_fake_module(root, distributed=True, marked=True, quizzes=True)

    shutil.copy(
        repo_root / "Dept grade sheet Template 2026.xlsx", root / "template.xlsx"
    )
    module_file = root / "module.toml"
    module_file.write_text(
        module_file.read_text(encoding="utf-8").replace(
            "[paths]", '[paths]\ndepartmental_template = "template.xlsx"', 1
        ),
        encoding="utf-8",
    )
    return root


def test_the_module_level_steps_run_from_the_buttons(marked_module, dashboard):
    """Collate, the departmental sheet, the moderation pack, SI's upload."""
    names = dashboard(marked_module)
    module = names["found"].module
    class_list = names["class_list"]

    # --- 5. collate -------------------------------------------------------
    marks, sheet = names["collate_the_module"](module, class_list, "feedback")

    assert len(sheet) == len(class_list), "every enrolled student reaches the sheet"
    assert "Total % Grade" in sheet.columns and "Letter Grade" in sheet.columns
    assert set(module.grade_sheet_columns) <= set(marks.columns)

    # --- 6. the departmental sheet ---------------------------------------
    path, written = names["write_the_departmental_sheet"](
        module, sheet, module.root / "PS4001 grades.xlsx"
    )

    assert path.exists()
    assert written.written == len(sheet)
    assert ModuleFile.load(marked_module).module.status.departmental_sheet_written

    # --- 7. the moderation pack ------------------------------------------
    sample = names["draw_the_sample"](sheet, 1, "include")
    pack = names["build_the_pack"](module, sample, module.root / "Moderation")

    assert pack.manifest.is_file()
    assert len(sample.selected) > 0
    assert ModuleFile.load(marked_module).module.status.moderation_pack_built

    # --- 8. SI's upload ---------------------------------------------------
    upload = names["fill_si_upload"](
        module, sheet, module.root / "PS4001_upload.CSV"
    )

    assert upload.filled > 0
    assert not upload.not_enrolled
    assert ModuleFile.load(marked_module).module.status.si_file_written


def test_the_sample_can_be_drawn_again_from_its_own_seed(marked_module, dashboard):
    """What makes a draw defensible months later.

    The pack records the seed it drew with, so "why these students?" has an
    answer that can be re-run rather than asserted.
    """
    names = dashboard(marked_module)
    module = names["found"].module
    _, sheet = names["collate_the_module"](module, names["class_list"], "feedback")

    first = names["draw_the_sample"](sheet, 1, "include")
    again = names["draw_the_sample"](sheet, 1, "include", seed=first.seed)

    assert list(again.selected["Student ID"]) == list(first.selected["Student ID"])


# ---------------------------------------------------------------------------
# The class list, and the steps that wait on it
# ---------------------------------------------------------------------------
#
# Every marking step is blocked until the class list reads, so a wrong path
# in module.toml left the page with nothing to do and no way forward except
# hand-editing the file -- which is the thing the app is meant to replace.


def _point_classlist_at(root, name):
    """Rewrite [paths].classlist, as a hand edit would."""
    toml = root / "module.toml"
    text = toml.read_text(encoding="utf-8")
    import re

    text = re.sub(r'classlist = "[^"]*"', f'classlist = "{name}"', text)
    toml.write_text(text, encoding="utf-8")
    return toml


def test_a_missing_class_list_does_not_stop_the_page(module_on_disk, dashboard):
    """It is the ordinary state before the export has been downloaded."""
    _point_classlist_at(module_on_disk, "not there.csv")

    names = dashboard(module_on_disk)

    assert names["class_list"] is None
    assert names["class_list_path"].name == "not there.csv"
    assert names["classlist_picker"] is not None, "and there is a way forward"


def test_a_class_list_that_will_not_parse_does_not_crash_the_page(
    module_on_disk, dashboard
):
    """import_brightspace_classlist returns None rather than raising for a
    file it cannot make sense of. Reading len(None) took the whole page down
    with a TypeError, which is the one thing a step may not do."""
    (module_on_disk / "rubbish.csv").write_text("not,a,class,list\n1,2,3,4\n")
    _point_classlist_at(module_on_disk, "rubbish.csv")

    names = dashboard(module_on_disk)

    assert names["class_list"] is None


def test_a_good_class_list_still_reads(module_on_disk, dashboard):
    names = dashboard(module_on_disk)

    assert names["class_list"] is not None
    assert len(names["class_list"]) > 0


def test_a_class_list_can_be_recorded_in_the_module_file(module_on_disk, dashboard):
    _point_classlist_at(module_on_disk, "not there.csv")
    names = dashboard(module_on_disk)

    written = names["remember_class_list"](module_on_disk / "classlist.xlsx")

    assert written == pl.Path("classlist.xlsx")
    assert 'classlist = "classlist.xlsx"' in (
        module_on_disk / "module.toml"
    ).read_text(encoding="utf-8")
    # And the module now loads with it.
    assert ModuleFile.load(module_on_disk).module.classlist_path.exists()


def test_a_class_list_outside_the_module_folder_is_refused(
    module_on_disk, dashboard, tmp_path
):
    """module.toml stores nothing absolute -- these folders live under
    OneDrive, where the absolute path differs per machine."""
    stray = tmp_path / "elsewhere.csv"
    stray.write_text("Username,Last Name,First Name\n#1,A,a\n")
    names = dashboard(module_on_disk)

    with pytest.raises(ValueError, match="outside the module folder"):
        names["remember_class_list"](stray)


def test_a_classlist_line_is_added_when_the_file_has_none(
    module_on_disk, dashboard
):
    """A form that cannot record its answer asks the same question every
    time. The key goes in above the table's trailing comments, so the
    comment introducing the next section stays with it."""
    toml = module_on_disk / "module.toml"
    text = toml.read_text(encoding="utf-8")
    import re

    toml.write_text(re.sub(r'classlist = "[^"]*"\n', "", text), encoding="utf-8")
    names = dashboard(module_on_disk)

    names["remember_class_list"](module_on_disk / "classlist.xlsx")

    after = toml.read_text(encoding="utf-8")
    assert 'classlist = "classlist.xlsx"' in after
    assert ModuleFile.load(module_on_disk).module.classlist_path.exists()
    for line in text.splitlines():
        if line.startswith("#"):
            assert line in after, f"comment lost: {line}"


# ---------------------------------------------------------------------------
# Collecting the groups from the page
# ---------------------------------------------------------------------------


def test_the_groups_can_be_collected_from_the_buttons(
    leader_managed_module, dashboard
):
    """The step where a mistyped id shows up, before the graders have
    workbooks rather than after."""
    names = dashboard(leader_managed_module)
    cw1 = names["found"].module.assessment("cw1")

    membership, attached = names["collect_groups"](cw1, names["class_list"])

    assert cw1.group_membership_path.exists()
    assert membership.frame["Group"].nunique() == 2
    assert "Group" in attached.columns


def test_a_student_left_off_the_sheets_is_named_here(
    leader_managed_module, dashboard
):
    """Rather than at allocation, which is after the graders have workbooks."""
    names = dashboard(leader_managed_module)
    # Imported AFTER the run, not before. The dashboard fixture drops
    # grader_helper from sys.modules so the notebook imports the real
    # package rather than the repo-root shim, which means a class imported
    # up here is a different object from the one the notebook raises, and
    # pytest.raises would not match it.
    from grader_helper import MissingGroupError

    cw1 = names["found"].module.assessment("cw1")
    (cw1.group_sheets_path / "Team 2.xlsx").unlink()

    with pytest.raises(MissingGroupError):
        names["collect_groups"](cw1, names["class_list"])


def test_allocation_says_when_there_are_no_group_sheets(
    leader_managed_module, dashboard
):
    """Up front, rather than as a traceback from inside the step."""
    names = dashboard(leader_managed_module)
    cw1 = names["found"].module.assessment("cw1")
    for sheet in cw1.group_sheets_path.iterdir():
        sheet.unlink()
    cw1.group_sheets_path.rmdir()

    again = dashboard(leader_managed_module)

    assert "no sheets at" in again["blocking"]("group sheets")[0]


def test_an_individual_assessment_has_no_group_sheets_to_miss(
    module_on_disk, dashboard
):
    names = dashboard(module_on_disk)

    assert names["blocking"]("group sheets") == []


def test_a_picked_class_list_wins_over_the_remembered_one(module_on_disk, dashboard):
    """A test cannot click a file browser, so this drives the choice the
    browser feeds. Picked has to win: a wrong path in module.toml is the
    whole reason the control exists."""
    names = dashboard(module_on_disk)
    choose = names["class_list_choice"]

    assert choose(pl.Path("picked.csv"), pl.Path("remembered.csv")) == (
        pl.Path("picked.csv"), "picked here",
    )
    assert choose(None, pl.Path("remembered.csv")) == (
        pl.Path("remembered.csv"), "from module.toml",
    )
    assert choose(None, None) == (None, "from module.toml")


def test_reading_a_class_list_from_a_path_given(module_on_disk, dashboard):
    names = dashboard(module_on_disk)

    frame, note = names["read_class_list"](
        module_on_disk / "classlist.xlsx", names["found"].module
    )

    assert frame is not None and len(frame) > 0
    assert "students" in note


# ---------------------------------------------------------------------------
# What is actually in the submissions folder
# ---------------------------------------------------------------------------
#
# "folders there: 0" was true of a folder that is not there, one with nothing
# in it, a download still in its zip, and a download unzipped one level too
# deep. Four situations wanting four different things done, reported
# identically -- and every step that reads submissions iterates the
# directories immediately inside it, so the difference is the difference
# between working and silently doing nothing.


@pytest.fixture
def describe(module_on_disk, dashboard):
    return dashboard(module_on_disk)["submissions_state"]


def test_a_real_download_is_counted(describe, module_on_disk):
    subs = module_on_disk / "assessments" / "cw1" / "submissions"

    how_many, trouble = describe(subs)

    assert how_many > 0
    assert trouble == ""


def test_a_folder_that_is_not_there_says_so(describe, tmp_path):
    how_many, trouble = describe(tmp_path / "nothing")

    assert how_many == 0
    assert "not there" in trouble


def test_an_empty_folder_is_not_a_missing_one(describe, tmp_path):
    empty = tmp_path / "submissions"
    empty.mkdir()

    how_many, trouble = describe(empty)

    assert how_many == 0
    assert "empty" in trouble
    assert "not there" not in trouble


def test_a_download_still_in_its_zip_is_named(describe, tmp_path):
    subs = tmp_path / "submissions"
    subs.mkdir()
    (subs / "PS4034 Assignment 1 Download.zip").write_bytes(b"PK")

    how_many, trouble = describe(subs)

    assert how_many == 0
    assert "still zipped" in trouble
    assert "PS4034 Assignment 1 Download.zip" in trouble


def test_loose_files_with_no_folders_are_named(describe, tmp_path):
    subs = tmp_path / "submissions"
    subs.mkdir()
    (subs / "essay.docx").write_bytes(b"")

    how_many, trouble = describe(subs)

    assert how_many == 0
    assert "loose file" in trouble


def test_a_download_unzipped_one_level_too_deep_is_named(describe, tmp_path):
    """Extracting the zip makes a folder named for the download, with the
    student folders inside it. Everything downstream looks one level up from
    where they are and finds nothing, without complaining."""
    subs = tmp_path / "submissions"
    wrapper = subs / "PS4034 Assignment 1 Download Aug 5, 2026"
    (wrapper / "27236-46025 - 23304308 Angood - 05 March 2026 612 PM").mkdir(
        parents=True
    )

    how_many, trouble = describe(subs)

    assert "one level too deep" in trouble
    assert "PS4034 Assignment 1 Download Aug 5, 2026" in trouble


def test_one_student_is_not_mistaken_for_a_nested_download(describe, tmp_path):
    """A cohort of one, or the last folder left after resolving the rest.
    A student folder holds files, not folders."""
    subs = tmp_path / "submissions"
    only = subs / "27236-46025 - 23304308 Angood - 05 March 2026 612 PM"
    only.mkdir(parents=True)
    (only / "essay.docx").write_bytes(b"")

    how_many, trouble = describe(subs)

    assert (how_many, trouble) == (1, "")


def test_a_folder_with_no_student_folders_blocks_distribution(
    module_on_disk, dashboard
):
    """And says so about the folder that is there, rather than claiming
    there is no folder -- which sent the reader looking in the wrong place."""
    subs = module_on_disk / "assessments" / "cw1" / "submissions"
    import shutil

    shutil.rmtree(subs)
    subs.mkdir()

    names = dashboard(module_on_disk)
    reasons = names["blocking"]("submissions")

    assert reasons and "no student folders" in reasons[0]
    assert "there is no submissions folder" not in reasons[0]


# ---------------------------------------------------------------------------
# Group sheets kept in one file
# ---------------------------------------------------------------------------
#
# `group_sheets` defaults to a folder called `groups/`, and a module written
# before that key existed has not got it at all -- so the page looked in an
# empty folder and said there were no sheets, about an assessment whose
# sheets were sitting in one file beside it.


@pytest.fixture
def one_groups_file(leader_managed_module):
    """The same groups, as one workbook at the assessment root."""
    import pandas as pd

    root = leader_managed_module
    cw1_folder = root / "assessments" / "cw1"
    ids = pd.read_excel(root / "classlist.xlsx", dtype=str)["Username"]
    ids = [str(i).replace("#", "") for i in ids]
    half = len(ids) // 2
    pd.DataFrame(
        {
            "Student Id": ids,
            "Group": ["Team 1"] * half + ["Team 2"] * (len(ids) - half),
        }
    ).to_excel(cw1_folder / "groups.xlsx", index=False)

    for sheet in (cw1_folder / "groups").iterdir():
        sheet.unlink()
    (cw1_folder / "groups").rmdir()
    return root


def test_groups_can_be_collected_from_one_file(one_groups_file, dashboard):
    names = dashboard(one_groups_file)
    cw1 = names["found"].module.assessment("cw1")
    groups_file = one_groups_file / "assessments" / "cw1" / "groups.xlsx"

    membership, attached = names["collect_groups"](
        cw1, names["class_list"], source=groups_file
    )

    assert membership.frame["Group"].nunique() == 2
    assert attached is not None


def test_where_the_sheets_are_can_be_remembered(one_groups_file, dashboard):
    """A form that cannot record its answer asks the same question every
    time -- and this key was not in the file to be updated in place."""
    names = dashboard(one_groups_file)
    cw1 = names["found"].module.assessment("cw1")
    groups_file = one_groups_file / "assessments" / "cw1" / "groups.xlsx"

    written = names["remember_group_sheets"](cw1, groups_file)

    assert written == pl.Path("groups.xlsx")
    reloaded = ModuleFile.load(one_groups_file).module.assessment("cw1")
    assert reloaded.group_sheets == "groups.xlsx"
    assert reloaded.group_sheets_path.is_file()


def test_group_sheets_outside_the_assessment_folder_are_refused(
    one_groups_file, dashboard, tmp_path
):
    """`group_sheets` is relative to the assessment's own folder, and
    module.toml stores nothing absolute."""
    names = dashboard(one_groups_file)
    cw1 = names["found"].module.assessment("cw1")
    stray = tmp_path / "groups.xlsx"
    stray.write_bytes(b"")

    with pytest.raises(ValueError, match="outside"):
        names["remember_group_sheets"](cw1, stray)


def test_a_remembered_groups_file_then_collects_on_its_own(
    one_groups_file, dashboard
):
    """End of the path: recorded, re-read off the disk, collected with no
    source given."""
    names = dashboard(one_groups_file)
    cw1 = names["found"].module.assessment("cw1")
    names["remember_group_sheets"](
        cw1, one_groups_file / "assessments" / "cw1" / "groups.xlsx"
    )

    again = dashboard(one_groups_file)
    membership, _ = again["collect_groups"](
        again["found"].module.assessment("cw1"), again["class_list"]
    )

    assert membership.frame["Group"].nunique() == 2
