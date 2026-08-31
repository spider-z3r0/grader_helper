#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""Building the departmental grade sheet for a module of any shape.

The load-bearing test here is `test_rebuilds_the_committed_template`. The
builder generalises the department's formulas to blocks of any width, and the
only way to know the generalisation is faithful is to point it at the shape the
department wrote by hand and require the template back. Everything else in this
file tests a shape nobody has a golden copy of, so it is only trustworthy
because that one passes.

House convention: ``pl`` is pathlib, ``pr`` is polars.
"""

import pathlib as pl
import statistics

import pandas as pd
import pytest
from openpyxl import load_workbook

from grader_helper import (
    Assessment,
    DepartmentalLayout,
    Module,
    Person,
    build_departmental_sheet,
    excel_round,
    sort_order_columns,
    write_departmental_sheet,
)
from grader_helper.file_operations.departmental_layout import (
    FIRST_DATA_ROW,
    HEADER_ROW,
    LAST_DATA_ROW,
    MEAN_ROW,
    N_ROW,
    SD_ROW,
)

TEMPLATE_NAME = "Dept grade sheet Template 2026.xlsx"

#: Rows 30-49 hold the sample data, which the builder clears on purpose.
LAST_SAMPLE_ROW = 50


def _module(*assessments: Assessment, code: str = "PS4001") -> Module:
    return Module(
        code=code,
        name="A module",
        year="2025/26",
        leader=Person(first_name="Kevin", last_name="O Malley", initials="KOM"),
        assessments=list(assessments),
    )


def _coursework(id: str, name: str, marks_out_of: float, weight: float) -> Assessment:
    return Assessment(
        id=id, type="coursework", name=name, marks_out_of=marks_out_of, weight=weight
    )


@pytest.fixture
def template(repo_root) -> pl.Path:
    path = repo_root / TEMPLATE_NAME
    assert path.is_file(), f"the departmental template is missing from {repo_root}"
    return path


@pytest.fixture
def template_shaped_module() -> Module:
    """The module the department's template was written for, by hand."""
    return _module(
        _coursework("cw1", "Coursework 1", 100, 40),
        _coursework("cw2", "Coursework 2", 100, 50),
        Assessment(id="mcq", type="mcq", name="MCQ", marks_out_of=10, weight=10),
    )


@pytest.fixture
def four_assessment_module() -> Module:
    """A shape the template has no room for: four pieces, two of them weighted."""
    return _module(
        _coursework("cw1", "Coursework 1", 100, 30),
        _coursework("cw2", "Coursework 2", 100, 30),
        Assessment(id="quiz", type="quiz", name="Quizzes", marks_out_of=10, weight=10),
        Assessment(id="exam", type="exam", name="Exam", marks_out_of=100, weight=30),
    )


def _value(cell):
    """A cell's formula text, whether it is a plain formula or an array one."""
    return getattr(cell.value, "text", cell.value)


def _header_row(worksheet) -> list:
    return [
        worksheet.cell(row=HEADER_ROW, column=column).value
        for column in range(1, worksheet.max_column + 1)
    ]


# --------------------------------------------------------------------------
# The guard the rest of the file rests on.


def test_rebuilds_the_committed_template(template, template_shaped_module, tmp_path):
    """A module of the template's shape produces the template back.

    Cell for cell, over the header row, all 501 ruled rows, the descriptives
    and the grade distribution. The one deliberate difference is that the sample rows are cleared, which is
    the point of building a sheet.
    """
    built = build_departmental_sheet(
        template_shaped_module, template, tmp_path / "built.xlsx"
    )
    want = load_workbook(template)["GradeTemplate"]
    got = load_workbook(built)["GradeTemplate"]

    assert _header_row(got) == _header_row(want)

    for row in (*range(FIRST_DATA_ROW, LAST_DATA_ROW + 1, 97), LAST_DATA_ROW):
        for column in range(1, 11):
            expected = _value(want.cell(row, column))
            actual = _value(got.cell(row, column))
            if row <= LAST_SAMPLE_ROW and actual is None:
                continue  # sample data, cleared on purpose
            assert actual == expected, f"row {row}, column {column}"

    for row in (MEAN_ROW, SD_ROW, N_ROW):
        for column in range(1, 11):
            assert _value(got.cell(row, column)) == _value(want.cell(row, column)), (
                f"descriptives differ at row {row}, column {column}"
            )

    for row in range(6, 18):
        for column in (7, 8, 9):
            assert _value(got.cell(row, column)) == _value(want.cell(row, column)), (
                f"band/distribution block differs at row {row}, column {column}"
            )

    for row in (HEADER_ROW, 51, 200, LAST_DATA_ROW):
        for column in range(1, 11):
            assert (
                got.cell(row, column).number_format
                == want.cell(row, column).number_format
            ), f"number format differs at row {row}, column {column}"


def test_the_weighting_divides_exactly_where_it_can(template_shaped_module):
    """``=E30/2``, not ``=E30/100*50`` -- and the difference is a whole mark.

    The tidier-looking long form is the wrong choice, which is worth a test
    because it is not obvious. ``x/2`` is exact in binary floating point;
    ``x/100*50`` is two roundings and is not. The gap is ~1e-14, but the total
    is ``ROUND(SUM(...),0)``, so a sum landing on an exact half falls the other
    way.
    """
    layout = DepartmentalLayout.for_module(template_shaped_module)
    cw1, cw2, _mcq = layout.module.assessments
    assert layout.weighting_formula(cw2, 30) == "=E30/2"
    assert layout.weighting_formula(cw1, 30) == "=C30/100*40"  # 100/40 = 2.5


def test_the_long_weighting_form_would_have_moved_marks():
    """The evidence for the rule above, kept so it cannot be undone quietly.

    A reviewer looking at ``=E30/2`` may reasonably want it written as
    ``=E30/100*50`` to show the weight. This is what that costs.
    """
    moved = [
        (cw2, excel_round(cw2 / 2), excel_round(cw2 / 100 * 50))
        for cw2 in [x / 2 for x in range(0, 201)]
        if excel_round(cw2 / 2) != excel_round(cw2 / 100 * 50)
    ]
    assert moved, "if this is empty the two forms agree and the rule is pointless"
    assert (29.0, 15.0, 14.0) in moved


def test_the_sheet_and_a_prepared_frame_agree_on_column_order(
    four_assessment_module,
):
    """The layout and `sort_order_columns` must not be two lists that match.

    A frame is ordered by `sort_order_columns` and the sheet is laid out by
    `DepartmentalLayout`. If those ever drift, marks are written into the
    wrong columns and every one of them still looks like a mark. They share
    the same tuples, and this is what says so.
    """
    layout = DepartmentalLayout.for_module(four_assessment_module)
    frame_order = sort_order_columns(layout.headers, four_assessment_module)
    assert frame_order == list(layout.headers)


# --------------------------------------------------------------------------
# The shapes the template has no golden copy of.


def test_every_assessment_reaches_the_total(template, four_assessment_module, tmp_path):
    """Four assessments, four terms in the total.

    The failure this exists to catch is the quiet one: a fourth assessment
    written into a sheet whose total sums three columns produces a mark that is
    30 points low and looks entirely plausible.
    """
    built = build_departmental_sheet(
        four_assessment_module, template, tmp_path / "four.xlsx"
    )
    worksheet = load_workbook(built)["GradeTemplate"]
    headers = _header_row(worksheet)

    assert headers[:2] == ["Name", "Student ID"]
    assert headers[2:9] == [
        "Coursework 1 (100)",
        "Coursework 1 (30)",
        "Coursework 2 (100)",
        "Coursework 2 (30)",
        "Quizzes (10)",
        "Exam (100)",
        "Exam (30)",
    ]
    assert headers[9:12] == ["Total % Grade", "Letter Grade", "Comments"]

    # D and F weight the courseworks, G is the quiz (marked on its own weight),
    # I weights the exam. Those four are what the total must sum.
    assert _value(worksheet["J30"]) == "=ROUND(SUM(D30,F30,G30,I30),0)"
    assert _value(worksheet["D30"]) == "=C30/100*30"
    assert _value(worksheet["I30"]) == "=H30/100*30"
    assert _value(worksheet["G30"]) is None, "a quiz marked on its weight is raw only"


def test_the_letter_grade_column_follows_the_block(
    template, four_assessment_module, tmp_path
):
    """The Letter Grade formula and its distribution counts move together.

    One of the two things the module leader has to re-point by hand, and the
    one that fails most visibly -- a distribution counting an empty column
    reports every student as NG.
    """
    built = build_departmental_sheet(
        four_assessment_module, template, tmp_path / "four.xlsx"
    )
    worksheet = load_workbook(built)["GradeTemplate"]

    assert worksheet["K29"].value == "Letter Grade"
    grade = _value(worksheet["K30"])
    assert grade.startswith('=IF(ROUND(J30,2)>0,'), grade
    assert '$A$8' in grade and '$B$17' in grade, "bands must still come from A8:B17"
    assert grade.endswith(',"NG")'), "a total of zero is NG, not F"

    counts = [_value(worksheet.cell(row, 8)) for row in range(6, 17)]
    assert counts[0] == '=COUNTIF(K30:K530,"A1")'
    assert counts[-1] == '=COUNTIF(K30:K530,"NG")'
    assert all("K30:K530" in formula for formula in counts)


def test_the_descriptives_cover_every_column(
    template, four_assessment_module, tmp_path
):
    """Mean, SD and N under all seven assessment columns and the total.

    The A23 block is the other hand-edit that goes wrong, and it fails quietly:
    a mean computed over six of seven components is a perfectly reasonable
    number, and nothing on the sheet says which one is missing.
    """
    built = build_departmental_sheet(
        four_assessment_module, template, tmp_path / "four.xlsx"
    )
    worksheet = load_workbook(built)["GradeTemplate"]

    covered = range(3, 11)  # C..I, the assessment block, plus J the total
    for column in [*covered, 10]:
        assert _value(worksheet.cell(MEAN_ROW, column)) is not None, column
        assert _value(worksheet.cell(SD_ROW, column)) is not None, column
        assert _value(worksheet.cell(N_ROW, column)) is not None, column

    assert _value(worksheet.cell(MEAN_ROW, 10)) == '=AVERAGEIF(J30:J530, "> 0")'
    # Nothing spills into the letter grade or comments columns.
    for column in (11, 12):
        assert _value(worksheet.cell(MEAN_ROW, column)) is None


def test_n_counts_a_raw_column_not_a_formula_column(
    template, four_assessment_module, tmp_path
):
    """N must count marks, and only a raw column knows how many there are.

    Every weighted and total cell carries a formula in all 501 rows, so
    counting one returns 501 whatever the cohort. The template's own N row
    counts the raw column to its left, and so does this.
    """
    built = build_departmental_sheet(
        four_assessment_module, template, tmp_path / "four.xlsx"
    )
    worksheet = load_workbook(built)["GradeTemplate"]

    assert _value(worksheet.cell(N_ROW, 3)) == "=COUNT(C30:C530)"
    assert _value(worksheet.cell(N_ROW, 4)) == "=COUNT(C30:C530)", "weighted counts raw"
    assert _value(worksheet.cell(N_ROW, 8)) == "=COUNT(H30:H530)"
    assert _value(worksheet.cell(N_ROW, 9)) == "=COUNT(H30:H530)", "weighted counts raw"
    assert _value(worksheet.cell(N_ROW, 10)) == "=COUNT(H30:H530)", "total counts raw"


def test_a_narrower_module_leaves_nothing_of_the_old_block(template, tmp_path):
    """Two assessments, and the template's extra columns are really gone.

    A sheet narrower than the template is the case where clearing matters: a
    stale 'MCQ (10)' header with the department's formulas under it would be
    read as a real component by anyone opening the file.
    """
    module = _module(
        _coursework("cw1", "Essay", 100, 50),
        _coursework("cw2", "Report", 100, 50),
    )
    built = build_departmental_sheet(module, template, tmp_path / "two.xlsx")
    worksheet = load_workbook(built)["GradeTemplate"]

    assert _header_row(worksheet)[:9] == [
        "Name",
        "Student ID",
        "Essay (100)",
        "Essay (50)",
        "Report (100)",
        "Report (50)",
        "Total % Grade",
        "Letter Grade",
        "Comments",
    ]
    assert _value(worksheet["G30"]) == "=ROUND(SUM(D30,F30),0)"
    for row in (HEADER_ROW, 30, 200, LAST_DATA_ROW):
        for column in range(10, 13):
            assert _value(worksheet.cell(row, column)) is None, (row, column)
    assert "K" not in worksheet.column_dimensions or not (
        worksheet.column_dimensions["K"].width
    )


def test_a_module_with_nothing_to_weight_has_one_column_each(template, tmp_path):
    """Two pieces each marked on its own contribution: no weighted columns.

    The two-numbers rule says a piece marked out of what it is worth needs one
    column, and the total then sums the raw marks directly.
    """
    module = _module(
        Assessment(id="a", type="mcq", name="Test A", marks_out_of=50, weight=50),
        Assessment(id="b", type="mcq", name="Test B", marks_out_of=50, weight=50),
    )
    built = build_departmental_sheet(module, template, tmp_path / "flat.xlsx")
    worksheet = load_workbook(built)["GradeTemplate"]

    assert _header_row(worksheet)[:7] == [
        "Name",
        "Student ID",
        "Test A (50)",
        "Test B (50)",
        "Total % Grade",
        "Letter Grade",
        "Comments",
    ]
    assert _value(worksheet["E30"]) == "=ROUND(SUM(C30,D30),0)"
    assert _value(worksheet.cell(N_ROW, 4)) == "=COUNT(D30:D530)"


def test_samples_are_gone_but_formulas_are_not(
    template, template_shaped_module, tmp_path
):
    built = build_departmental_sheet(
        template_shaped_module, template, tmp_path / "built.xlsx"
    )
    worksheet = load_workbook(built)["GradeTemplate"]

    names = [
        worksheet.cell(row, 1).value
        for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1)
    ]
    assert not any(names), "the template's sample students are still in the sheet"
    assert _value(worksheet["D30"]) == "=C30/100*40"
    assert _value(worksheet["H530"]) == "=ROUND(SUM(D530,F530,G530),0)"


def test_the_other_sheets_are_untouched(template, four_assessment_module, tmp_path):
    """Only GradeTemplate is rebuilt; moderation and guidance are left alone."""
    built = build_departmental_sheet(
        four_assessment_module, template, tmp_path / "four.xlsx"
    )
    want = load_workbook(template)
    got = load_workbook(built)

    assert got.sheetnames == want.sheetnames
    for name in want.sheetnames:
        if name == "GradeTemplate":
            continue
        a, b = want[name], got[name]
        assert b.max_row == a.max_row and b.max_column == a.max_column, name
        for row in a.iter_rows():
            for cell in row:
                assert _value(b[cell.coordinate]) == _value(cell), (name, cell.coordinate)


# --------------------------------------------------------------------------
# Refusals.


def test_refuses_to_overwrite_a_sheet_that_may_hold_marks(
    template, template_shaped_module, tmp_path
):
    destination = tmp_path / "built.xlsx"
    build_departmental_sheet(template_shaped_module, template, destination)
    with pytest.raises(FileExistsError, match="already exists"):
        build_departmental_sheet(template_shaped_module, template, destination)
    build_departmental_sheet(
        template_shaped_module, template, destination, overwrite=True
    )


def test_refuses_a_module_with_no_assessments(template, tmp_path):
    module = _module()
    with pytest.raises(ValueError, match="no assessments"):
        build_departmental_sheet(module, template, tmp_path / "empty.xlsx")


def test_refuses_a_missing_template(template_shaped_module, tmp_path):
    with pytest.raises(FileNotFoundError):
        build_departmental_sheet(
            template_shaped_module, tmp_path / "nope.xlsx", tmp_path / "out.xlsx"
        )


def test_refuses_a_sheet_without_the_named_tab(
    template, template_shaped_module, tmp_path
):
    with pytest.raises(KeyError, match="Marks"):
        build_departmental_sheet(
            template_shaped_module, template, tmp_path / "out.xlsx", sheet="Marks"
        )


# --------------------------------------------------------------------------
# Writing marks into a sheet.


@pytest.fixture
def marks() -> pd.DataFrame:
    """Three students, including one who has a mark for only part of the module."""
    return pd.DataFrame(
        {
            "Name": ["Ahearn, Niamh", "Byrne, Sean", "Okafor, Chidi"],
            "Student ID": ["00123456", "23304308", "56170559"],
            "Coursework 1 (100)": [74.0, 53.5, 61.0],
            "Coursework 2 (100)": [66.5, 67.0, None],
            "MCQ (10)": [7.0, 7.0, 4.0],
        }
    )


@pytest.fixture
def built_sheet(template, template_shaped_module, tmp_path) -> pl.Path:
    return build_departmental_sheet(
        template_shaped_module, template, tmp_path / "grades.xlsx"
    )


def test_writes_the_marks_and_leaves_the_formulas_alone(
    built_sheet, template_shaped_module, marks
):
    """Three values a row, and the sheet's own arithmetic untouched.

    openpyxl does not evaluate formulas, so what is checked here is that the
    marks landed and the department's formulas are still formulas. That they
    compute the right answer needs a real Excel, which is what the `excel`
    marker is for.
    """
    write_departmental_sheet(marks, template_shaped_module, built_sheet)
    worksheet = load_workbook(built_sheet)["GradeTemplate"]

    assert worksheet["A30"].value == "Ahearn, Niamh"
    assert worksheet["B30"].value == "00123456"
    assert worksheet["C30"].value == 74.0
    assert worksheet["E30"].value == 66.5
    assert worksheet["G30"].value == 7.0

    assert _value(worksheet["D30"]) == "=C30/100*40"
    assert _value(worksheet["F30"]) == "=E30/2"
    assert _value(worksheet["H30"]) == "=ROUND(SUM(D30,F30,G30),0)"
    assert _value(worksheet["I30"]).startswith("=IF(ROUND(H30,2)>0,")

    # Nothing below the cohort, and the formulas still waiting there.
    assert worksheet["A33"].value is None
    assert _value(worksheet["H33"]) == "=ROUND(SUM(D33,F33,G33),0)"


def test_a_student_id_keeps_its_leading_zeros(built_sheet, template_shaped_module, marks):
    """'00123456' must not come back as 123456.

    A column of digit strings that goes into Excel as a number loses its
    leading zeros, and every later merge against the class list then fails on a
    dtype mismatch -- which is the loud version. The quiet version is a student
    who cannot be matched at all.
    """
    write_departmental_sheet(marks, template_shaped_module, built_sheet)
    worksheet = load_workbook(built_sheet)["GradeTemplate"]
    assert worksheet["B30"].value == "00123456"
    assert isinstance(worksheet["B30"].value, str)


def test_a_missing_mark_stays_blank_rather_than_becoming_zero(
    built_sheet, template_shaped_module, marks
):
    """An unmarked component is empty, not nil.

    A zero is a mark a student earned. A blank is work that has not been
    marked, and the difference is a third of the module.
    """
    write_departmental_sheet(marks, template_shaped_module, built_sheet)
    worksheet = load_workbook(built_sheet)["GradeTemplate"]
    assert worksheet["E32"].value is None
    assert worksheet["C32"].value == 61.0


def test_a_second_run_does_not_strand_the_first_cohort(
    built_sheet, template_shaped_module, marks
):
    """Rewriting with fewer students leaves nobody behind.

    Without clearing, the tail of the previous run stays in the sheet, is
    counted by N, averaged by the descriptives and given a letter grade. It
    looks exactly like part of the cohort.
    """
    write_departmental_sheet(marks, template_shaped_module, built_sheet)
    write_departmental_sheet(marks.head(1), template_shaped_module, built_sheet)
    worksheet = load_workbook(built_sheet)["GradeTemplate"]

    assert worksheet["A30"].value == "Ahearn, Niamh"
    for row in range(31, LAST_DATA_ROW + 1):
        assert worksheet.cell(row, 1).value is None, row
        assert worksheet.cell(row, 3).value is None, row


def test_writes_to_a_destination_without_touching_the_source(
    built_sheet, template_shaped_module, marks, tmp_path
):
    out = write_departmental_sheet(
        marks, template_shaped_module, built_sheet, tmp_path / "final.xlsx"
    )
    assert out.path == tmp_path / "final.xlsx"
    assert out.written == len(marks), "the count is the evidence record() reads"
    assert load_workbook(out.path)["GradeTemplate"]["A30"].value == "Ahearn, Niamh"
    assert load_workbook(built_sheet)["GradeTemplate"]["A30"].value is None


def test_refuses_an_assessment_the_sheet_has_no_column_for(
    built_sheet, four_assessment_module, marks
):
    """A module the sheet was not laid out for is named, not half-written.

    This is the refusal the whole design turns on: writing three of four
    assessments produces a total short by the fourth's weight, and there is
    nothing on the face of the sheet to say so.
    """
    with pytest.raises(ValueError, match="no column for"):
        write_departmental_sheet(marks, four_assessment_module, built_sheet)


def test_refuses_a_sheet_column_the_module_does_not_account_for(
    built_sheet, marks, tmp_path
):
    """The same failure from the other side, and just as quiet.

    A sheet with an MCQ column written for a module that has no MCQ leaves that
    column empty. It still reaches the total, contributing zero, so every
    student is ten points light.
    """
    module = _module(
        _coursework("cw1", "Coursework 1", 100, 40),
        _coursework("cw2", "Coursework 2", 100, 60),
    )
    frame = marks.drop(columns=["MCQ (10)"]).rename(
        columns={"Coursework 2 (100)": "Coursework 2 (100)"}
    )
    with pytest.raises(ValueError, match="does not account for"):
        write_departmental_sheet(frame, module, built_sheet)


def test_refuses_a_cohort_bigger_than_the_sheet(
    built_sheet, template_shaped_module, marks
):
    """501 ruled rows, and row 531 is outside every formula on the sheet."""
    too_many = pd.concat([marks] * 200, ignore_index=True)
    assert len(too_many) > 501
    with pytest.raises(ValueError, match="the sheet rules rows"):
        write_departmental_sheet(too_many, template_shaped_module, built_sheet)


def test_refuses_a_frame_without_the_marks(built_sheet, template_shaped_module, marks):
    with pytest.raises(ValueError, match="missing the marks for"):
        write_departmental_sheet(
            marks.drop(columns=["MCQ (10)"]), template_shaped_module, built_sheet
        )


def test_refuses_an_empty_frame(built_sheet, template_shaped_module):
    with pytest.raises(ValueError, match="empty"):
        write_departmental_sheet(pd.DataFrame(), template_shaped_module, built_sheet)


# --------------------------------------------------------------------------
# The one check that needs a real Excel.


@pytest.mark.excel
def test_excel_computes_what_we_compute(built_sheet, template_shaped_module, marks):
    """The sheet's own total and letter grade agree with ours.

    Everything above asserts that the right *formulas* landed in the right
    cells; openpyxl does not evaluate them, so nothing so far shows they
    compute the right answer. This opens the workbook in Excel, lets it
    recalculate, and compares what it produces against
    `prepare_data_for_departmental_template`.

    That comparison is the whole point of the design. The sheet is the source
    of truth, our arithmetic exists to be checked against it, and this is the
    only test where the two actually meet. It is skipped without Excel, which
    means Linux CI never runs it -- so a failure here is a failure on the
    module leader's own machine, which is where it matters.
    """
    xw = pytest.importorskip("xlwings")
    from grader_helper import (
        calculate_weighted_score,
        prepare_data_for_departmental_template,
    )

    # A complete cohort. A student missing a component is a different question
    # -- Excel's SUM reads a blank as zero, so the sheet gives them a total as
    # though they had scored nil -- and it is recorded in the notes as a gap
    # rather than smuggled into this test.
    complete = marks.dropna().reset_index(drop=True)
    write_departmental_sheet(complete, template_shaped_module, built_sheet)

    ours = complete.copy()
    for assessment in template_shaped_module.assessments:
        if assessment.weighted_column:
            calculate_weighted_score(
                ours, assessment.raw_column, assessment.weight_fraction()
            )
    ours = prepare_data_for_departmental_template(ours, template_shaped_module)

    app = xw.App(visible=False)
    try:
        book = app.books.open(str(built_sheet))
        sheet = book.sheets["GradeTemplate"]
        app.calculate()
        totals = [sheet.range(f"H{30 + i}").value for i in range(len(complete))]
        letters = [sheet.range(f"I{30 + i}").value for i in range(len(complete))]
        # The descriptives, and the SD in particular. It is a dynamic-array
        # formula (`_xlfn.STDEV.S` over `_xlfn._xlws.FILTER`), the builder
        # regenerates it, and openpyxl can only confirm the text is right --
        # whether Excel still evaluates it in a rebuilt file, rather than
        # showing #NAME?, is only answerable here.
        descriptives = {
            row: sheet.range(f"C{row}").value for row in (MEAN_ROW, SD_ROW, N_ROW)
        }
        book.close()
    finally:
        app.quit()

    assert totals == [float(t) for t in ours["Total % Grade"]], (
        "Excel's total and ours disagree. The sheet wins -- ours is the one to "
        "fix."
    )
    assert letters == list(ours["Letter Grade"])

    marks_awarded = list(complete["Coursework 1 (100)"])
    assert descriptives[N_ROW] == len(marks_awarded)
    assert descriptives[MEAN_ROW] == pytest.approx(statistics.mean(marks_awarded))
    assert descriptives[SD_ROW] == pytest.approx(statistics.stdev(marks_awarded)), (
        "the SD is a dynamic-array formula; a rebuilt sheet must still "
        "evaluate it rather than returning an error string"
    )


# --------------------------------------------------------------------------
# The two templates in circulation.


@pytest.fixture
def template_with_an_unreachable_top_band(template, tmp_path) -> pl.Path:
    """The other 2026 template: A1 closed with '<' instead of '<='.

    Two copies of the departmental template exist and differ by exactly this
    character, in the letter-grade column and nowhere else. Built here from
    the committed one so the difference is stated in code rather than a second
    binary being committed to prove a point.
    """
    from openpyxl import load_workbook as _load

    book = _load(template)
    worksheet = book["GradeTemplate"]
    for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
        cell = worksheet.cell(row, 9)
        cell.value = cell.value.replace("<=$B$17", "<$B$17")
    path = tmp_path / "other template.xlsx"
    book.save(path)
    return path


def test_refuses_a_template_whose_top_band_excludes_a_perfect_score(
    template_with_an_unreachable_top_band, template_shaped_module, tmp_path
):
    """A total of exactly 100 must not come out as NG.

    Every band but the last is closed by the one above it; A1 has nothing
    above it, so it has to close with '<='. With '<' a total of 100 matches
    no band and falls through the nested IF to the final "NG" -- full marks
    recorded as no participation.

    The builder regenerates this formula, so without the guard it would emit
    the correct form silently and overrule the department's file without
    saying so. It refuses instead.
    """
    with pytest.raises(ValueError, match="falls through"):
        build_departmental_sheet(
            template_shaped_module,
            template_with_an_unreachable_top_band,
            tmp_path / "out.xlsx",
        )


def test_the_committed_template_grades_a_perfect_score_as_a1(template):
    """The band table and the formula beside it have to agree.

    Read off the committed file rather than asserted about our own code: row
    17 says A1 runs to 100 inclusive, and the formula must include 100 too.
    """
    worksheet = load_workbook(template)["GradeTemplate"]

    assert worksheet.cell(17, 1).value == 80
    assert worksheet.cell(17, 2).value == 100
    assert worksheet.cell(17, 3).value == "A1"
    assert "<=$B$17" in worksheet["I30"].value

    from grader_helper import make_letter_grade

    assert make_letter_grade(100) == "A1", "our code must agree with the sheet"


# --------------------------------------------------------------------------
# A raw column in the middle of the block.


@pytest.fixture
def mixed_module() -> Module:
    """PS4003's shape: an unweighted assessment between two weighted ones.

    Ten quiz marks worth ten need no weighted column, so the raw column has to
    reach the total itself while its neighbours reach it through theirs. That
    is the arrangement most likely to be got wrong by hand -- summing only the
    weighted columns drops it, summing every column double-counts the marks
    that were weighted.
    """
    return _module(
        _coursework("cw1", "Coursework 1", 100, 30),
        Assessment(id="quizzes", type="quiz", name="Quizzes",
                   marks_out_of=10, weight=10, pass_mark=80.0),
        Assessment(id="mcq", type="mcq", name="MCQ", marks_out_of=100, weight=20),
        Assessment(id="exam", type="exam", name="Exam", marks_out_of=100, weight=40),
    )


def test_an_unweighted_column_reaches_the_total_directly(
    template, mixed_module, tmp_path
):
    built = build_departmental_sheet(mixed_module, template, tmp_path / "mixed.xlsx")
    worksheet = load_workbook(built)["GradeTemplate"]

    assert _header_row(worksheet)[2:9] == [
        "Coursework 1 (100)", "Coursework 1 (30)",
        "Quizzes (10)",
        "MCQ (100)", "MCQ (20)",
        "Exam (100)", "Exam (40)",
    ]
    # D, G, I are weighted; E is the quizzes' raw column, reaching the total
    # on its own because ten marks worth ten need no weighting.
    assert _value(worksheet["J30"]) == "=ROUND(SUM(D30,E30,G30,I30),0)"
    assert _value(worksheet["F30"]) is None, "a raw column holds no formula"


def test_the_exact_divisor_form_is_used_where_it_applies(mixed_module):
    """100 marks worth 20 divides exactly, so it gets `=F30/5`.

    The coursework and exam do not -- 100 divides neither 30 nor 40 -- so they
    take the long form. One module, both shapes.
    """
    layout = DepartmentalLayout.for_module(mixed_module)
    cw1, _quizzes, mcq, exam = layout.module.assessments

    assert layout.weighting_formula(mcq, 30) == "=F30/5"
    assert layout.weighting_formula(cw1, 30) == "=C30/100*30"
    assert layout.weighting_formula(exam, 30) == "=H30/100*40"


def test_n_counts_the_right_source_around_an_unweighted_column(
    template, mixed_module, tmp_path
):
    """The 'nearest raw column at or before' rule, with a raw one mid-block."""
    built = build_departmental_sheet(mixed_module, template, tmp_path / "mixed.xlsx")
    worksheet = load_workbook(built)["GradeTemplate"]

    assert _value(worksheet.cell(N_ROW, 3)) == "=COUNT(C30:C530)"
    assert _value(worksheet.cell(N_ROW, 4)) == "=COUNT(C30:C530)"
    assert _value(worksheet.cell(N_ROW, 5)) == "=COUNT(E30:E530)", "quizzes count themselves"
    assert _value(worksheet.cell(N_ROW, 6)) == "=COUNT(F30:F530)"
    assert _value(worksheet.cell(N_ROW, 7)) == "=COUNT(F30:F530)"
    assert _value(worksheet.cell(N_ROW, 10)) == "=COUNT(H30:H530)", "total counts the last raw"
