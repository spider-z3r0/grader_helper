#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""Fill in marks that nobody marked, so the rest of the pipeline can be run.

Everything downstream of allocation needs marks to exist: reconciliation
needs two records to compare, collation needs something to collate, the
moderation sample needs a spread of bands to stratify, and the departmental
sheet needs numbers to put in it. Producing those by hand means opening
forty workbooks, and producing them by marking real work means marking real
work.

So this is the stand-in for **steps 5 and 6 of the lifecycle** -- the grader
writing a mark on the feedback sheet, and the grader copying that mark into
their own workbook. Two writes, not one, because they are two records and
the whole point of step 7 is that they can disagree. A simulator that only
ever wrote them together could never exercise the control that exists to
catch the copy going wrong, so ``discrepancies=`` deliberately mistypes a
few.

**This writes into real files.** It is a testing aid, meant for a scratch
copy of a module. Two things protect you and neither is clever:

- a feedback sheet whose grade cell already holds a number is **skipped**,
  so a part-marked assessment is not quietly overwritten;
- the command line **shows the plan and writes nothing** unless you pass
  ``--write``.

Writing a number into the grade cell **replaces whatever formula is there**.
That is unavoidable -- a real rubric calculates its total from criterion
cells this knows nothing about -- and it is the main reason to point this at
a copy.

House convention::

    import pathlib as pl        # pl is pathlib
    import polars as pr         # pr is polars

From a terminal::

    uv run simulate-marking ~/scratch/PS4001
    uv run simulate-marking ~/scratch/PS4001 -a cw1 --write
    uv run simulate-marking ~/scratch/PS4001 -a cw1 --explain

or, for files that are not laid out as a module -- a download and the grader
workbooks sitting in folders of their own::

    uv run simulate-marking -s "PS4001 CW1/Downloads" -c D30 \
                            -w "PS4001 CW1/Grader sheets" --write
"""

import argparse
import pathlib as pl
import random
import zlib
from typing import Mapping, NamedTuple, Sequence

import pandas as pd
from openpyxl import load_workbook

from .dataframe_operations.make_letter_grade import GRADE_BANDS
from .models import Assessment, Module, load_module

#: Workbook formats a feedback sheet may be saved in. The same set
#: `catch_grades` reads, so what can be written can be read back.
SHEET_SUFFIXES = (".xlsx", ".xlsm", ".xlsb", ".xls")

#: What `distribute_feedback_sheets` names the sheets it copies in --
#: "Feedback sheet 24439711.xlsx" -- with the identifier after it a student
#: id or a group label like "Team 3".
#:
#: Lower case because it is compared against a lower-cased stem, so the
#: match ignores case. Written out here because a lower-case constant beside
#: a capitalised filename reads like a bug on every scan of this file.
SHEET_PREFIX = "feedback sheet "

#: The column a grader writes their mark into, matching `allocating` and
#: `collating` -- this stands in for the grader, so it writes where they do.
MARK_COLUMN = "Mark"

#: Roughly where a psychology coursework cohort actually sits, as fractions
#: of whatever the piece is marked out of. Not uniform: a uniform cohort
#: puts a tenth of the class in every band, which makes the moderation
#: sample look easy and the grade distribution look nothing like one.
MEAN_FRACTION = 0.58
SD_FRACTION = 0.15

#: Marks worth planting deliberately, as percentages, because they are the
#: ones the arithmetic can get wrong and random draws almost never hit:
#:
#: * every band edge, where one mark either way is a different letter grade;
#: * the halves that Python and Excel round in opposite directions --
#:   ``round(64.5)`` is 64 (half to even) and Excel's is 65 (half away from
#:   zero), and 64 is a B3 where 65 is a B2. That rule has a house note of
#:   its own, and a cohort with no exact halves in it never tests it;
#: * zero, which the departmental sheet grades NG rather than F.
BOUNDARY_MARKS: tuple[float, ...] = tuple(
    sorted(
        {band.lower for band in GRADE_BANDS}
        | {54.5, 64.5, 74.5, 39.5, 49.5}
        | {0.0}
    )
)


class SimulatedMarking(NamedTuple):
    """What a simulated marking run did, or would do.

    Returned rather than printed, so the command line can render it and a
    test can assert on it.
    """

    #: ``{identifier: mark}`` -- the mark written on each feedback sheet.
    marks: dict[str, float]
    #: Feedback sheets written, by identifier. A list each, because a
    #: student who submitted twice has two sheets and both are theirs.
    sheets: dict[str, list[pl.Path]]
    #: Sheets left alone because they already carried a mark.
    skipped: dict[str, list[pl.Path]]
    #: Grader workbooks filled in, and how many rows each got.
    workbooks: dict[str, int]
    #: ``{identifier: (on the sheet, in the workbook)}`` where the two were
    #: deliberately made to disagree -- the mistyped copy step 7 catches.
    discrepancies: dict[str, tuple[float, float]]
    #: Sheets that would not take a mark, and why. Collected rather than
    #: raised: a run over a real cohort writes eighty-odd files, and dying on
    #: the fortieth leaves half of them marked with no record of which half.
    refused: dict[pl.Path, str]
    #: True when nothing was written.
    dry_run: bool

    def __str__(self) -> str:
        what = "would write" if self.dry_run else "wrote"
        written = sum(len(paths) for paths in self.sheets.values())
        skipped = sum(len(paths) for paths in self.skipped.values())
        parts = [
            f"{what} {written} feedback sheet(s) for {len(self.sheets)} student(s)",
            f"{skipped} already marked",
            f"{len(self.workbooks)} grader workbook(s)",
        ]
        if self.discrepancies:
            parts.append(f"{len(self.discrepancies)} planted discrepancy(ies)")
        if self.refused:
            parts.append(f"**{len(self.refused)} would not take a mark**")
        return ", ".join(parts)


def _identifier(stem: str) -> str | None:
    """The student id or group label a feedback sheet is named for.

    Case-insensitive on the prefix and case-preserving on what follows:
    "Feedback sheet 24439711" and "FEEDBACK SHEET Team 3" both work, and the
    identifier comes back as it was written.
    """
    if not stem.lower().startswith(SHEET_PREFIX):
        return None
    found = stem[len(SHEET_PREFIX):].strip()
    return found or None


def feedback_sheets_in(submissions: pl.Path) -> dict[str, list[pl.Path]]:
    """Every distributed feedback sheet under a folder, by identifier.

    The folder form. `feedback_sheets` is the same thing for an assessment
    that knows where its submissions are.

    **A list per identifier, not one path.** A student who submitted twice
    has two folders and therefore two feedback sheets, and that is a normal
    state -- resolving resubmissions is a judgement call, so a download sits
    in it until somebody makes one. Keeping only the last would leave the
    other sheet unmarked, and `catch_grades` reads every sheet it finds, so
    the result is a warning about an empty cell on a student who was marked.

    Raises
    ------
    NotADirectoryError
        If the submissions folder is not there, which means the download has
        not been unzipped into it yet.
    """
    submissions = pl.Path(submissions)
    if not submissions.is_dir():
        raise NotADirectoryError(
            f"{submissions} is not a directory. This should be the unzipped "
            "folder of submissions, with a feedback sheet in each student's "
            "or group's folder."
        )

    found: dict[str, list[pl.Path]] = {}
    for path in sorted(submissions.rglob("*")):
        # The same extensions catch_grades reads. A rubric saved as .xlsm
        # was findable by the reader and invisible to the writer, which is a
        # difference nothing should have.
        if path.suffix.lower() not in SHEET_SUFFIXES:
            continue
        if path.name.startswith("~$"):
            continue
        identifier = _identifier(path.stem)
        if identifier is not None:
            found.setdefault(identifier, []).append(path)
    return found


def feedback_sheets(assessment: Assessment) -> dict[str, list[pl.Path]]:
    """Every distributed feedback sheet for an assessment, by identifier."""
    return feedback_sheets_in(assessment.submissions_path)


#: Files `grading_output` holds that are not a grader's workbook. Everything
#: else in there is named for the grader it belongs to, which is what makes
#: the folder readable without being told who the graders are.
NOT_A_GRADER_WORKBOOK = ("completed_grades", "distributed", "group_membership")


def grader_workbooks(
    folder: pl.Path, graders: "Sequence[str] | None" = None
) -> dict[str, pl.Path]:
    """The grader workbooks in a folder, by the grader they are named for.

    Given ``graders``, only those, and a missing one is simply absent rather
    than an error -- a grader who has not been allocated anything yet has no
    workbook, and that is not this function's business. Given none, every
    workbook in the folder, which is how a folder can be pointed at without
    being told who marks it.
    """
    folder = pl.Path(folder)
    if not folder.is_dir():
        return {}

    if graders is not None:
        found = {g: folder / f"{g}.xlsx" for g in graders}
        return {g: path for g, path in found.items() if path.exists()}

    return {
        path.stem: path
        for path in sorted(folder.glob("*.xlsx"))
        if not path.name.startswith("~$")
        and path.stem.lower() not in NOT_A_GRADER_WORKBOOK
    }


def grade_cell_value(path: pl.Path, cell: str):
    """The cached value in a workbook's grade cell, or None.

    Cached, because that is what ``catch_grades`` reads too. A formula that
    Excel has never evaluated has no cached result and comes back None; one
    it has comes back as whatever it last computed.
    """
    workbook = None
    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
        return workbook.worksheets[0][cell].value
    except Exception:
        return None
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass


def _is_a_mark(value, blank=None) -> bool:
    """Whether a grade cell holds a mark somebody put there.

    **"Holds a number" is not enough.** A real rubric *calculates* its total,
    so the moment Excel saves it the grade cell caches a result -- usually
    ``0``, the sum of criteria nobody has filled in yet. Every distributed
    sheet then looks marked, and a run that means to fill them all in skips
    every one of them, quietly, while still filling the grader workbooks. So
    the test is not "is there a number" but "is it a *different* number from
    the one the blank sheet has".

    Without a blank to compare -- pointed at folders, with no rubric named --
    it falls back to any number, which is the old rule and is right for a
    sheet written by this package, whose blank grade cell is empty.
    """
    if value is None or isinstance(value, bool) or not isinstance(value, (int, float)):
        return False
    if blank is not None and value == blank:
        return False
    return True


def draw_marks(
    identifiers: Sequence[str],
    marks_out_of: float,
    *,
    seed: int | None = None,
    boundaries: int = 3,
    rng: "random.Random | None" = None,
) -> dict[str, float]:
    """A plausible mark for each identifier, rounded to the nearest half.

    Parameters
    ----------
    identifiers
        Who is being marked -- student ids, or group labels.
    marks_out_of
        The assessment's own scale. The distribution and the planted
        boundary marks are both expressed as fractions of it, so a piece
        marked out of 50 gets sensible numbers without being told anything
        else.
    seed
        For a reproducible cohort. Worth using: a bug found against one set
        of marks is much easier to chase when the marks come back.
    rng
        An existing random source, used instead of ``seed``. How
        ``simulate_marking`` keeps drawing marks and choosing who mistypes
        one off a single stream -- two ``Random(seed)`` objects make the
        same first choice, so the mistyped students were exactly the
        boundary students, every run.
    boundaries
        How many of the marks to place on a band edge or an awkward half
        rather than drawing them. See :data:`BOUNDARY_MARKS`.

    Returns
    -------
    dict
        ``{identifier: mark}``.
    """
    rng = rng if rng is not None else random.Random(seed)
    ordered = list(identifiers)

    planted: dict[str, float] = {}
    if boundaries > 0 and ordered:
        chosen = rng.sample(ordered, min(boundaries, len(ordered)))
        for i, identifier in enumerate(chosen):
            percent = BOUNDARY_MARKS[i % len(BOUNDARY_MARKS)]
            planted[identifier] = round(percent / 100 * marks_out_of, 2)

    marks: dict[str, float] = {}
    for identifier in ordered:
        if identifier in planted:
            marks[identifier] = planted[identifier]
            continue
        drawn = rng.gauss(MEAN_FRACTION * marks_out_of, SD_FRACTION * marks_out_of)
        clamped = min(max(drawn, 0.0), marks_out_of)
        # To the nearest half: what a grader actually writes, and it puts
        # exact halves in front of the rounding rule on purpose.
        marks[identifier] = round(clamped * 2) / 2
    return marks


def _write_sheet(path: pl.Path, cell: str, mark: float) -> None:
    """Put a mark in the grade cell.

    Raises
    ------
    ValueError
        If the cell cannot take a value, with the reason in words. The one
        that actually happens is a **merged** grade cell: openpyxl will only
        write to the top-left of a merged range, and a rubric that merges the
        total across two columns puts the grade cell somewhere else in it.
        Worth naming, because "'MergedCell' object attribute 'value' is
        read-only" is not a sentence about a feedback sheet.
    """
    workbook = load_workbook(path)
    sheet = workbook.worksheets[0]
    target = sheet[cell]

    if type(target).__name__ == "MergedCell":
        anchor = next(
            (
                str(rng).split(":")[0]
                for rng in sheet.merged_cells.ranges
                if target.coordinate in rng
            ),
            None,
        )
        raise ValueError(
            f"{cell} is inside a merged range, so only {anchor} can be "
            f"written. Either put the mark cell at {anchor} in module.toml, "
            "or unmerge it in the blank feedback sheet."
        )

    target.value = mark
    workbook.save(path)


def _key_column(frame: pd.DataFrame) -> str:
    """What a grader workbook's rows are keyed by.

    ``Student ID`` for the per-student shape, ``Group`` for the per-group one
    a Brightspace-managed group assessment gets -- the same split
    `allocate_graders` writes them in.
    """
    for column in ("Student ID", "Group"):
        if column in frame.columns:
            return column
    raise KeyError(
        "A grader workbook has neither a 'Student ID' nor a 'Group' column, "
        f"so its rows cannot be matched to a mark. Columns present: "
        f"{list(frame.columns)}. allocate_graders writes one or the other."
    )


def simulate_marking_in(
    submissions: pl.Path,
    grade_cell: str,
    *,
    workbooks: pl.Path | None = None,
    graders: "Sequence[str] | None" = None,
    blank: pl.Path | None = None,
    marks_out_of: float = 100,
    seed: int | None = None,
    boundaries: int = 3,
    discrepancies: int = 0,
    overwrite: bool = False,
    dry_run: bool = False,
    marks: Mapping[str, float] | None = None,
) -> SimulatedMarking:
    """
    Mark a folder of submissions as if a grader had, on both records.

    **The folder form.** Nothing here needs a ``module.toml`` or the
    assessment layout -- point it at an unzipped download and, if you have
    them, at the folder holding the grader workbooks.
    :func:`simulate_marking` is the same thing for an assessment that
    already knows where those two folders are.

    Parameters
    ----------
    submissions : pathlib.Path
        The unzipped download, with a feedback sheet in each student's or
        group's folder.
    grade_cell : str
        The cell in the feedback sheet the mark goes in -- the same cell
        ``catch_grades`` reads back.
    workbooks : pathlib.Path, optional
        The folder holding the grader workbooks. Left out, only the feedback
        sheets are marked, which is the *first* of the two records and on
        its own gives reconciliation nothing to compare.
    graders : sequence of str, optional
        Whose workbooks to fill. Left out, every workbook in ``workbooks``
        is taken to be a grader's -- see :func:`grader_workbooks`.
    blank : pathlib.Path, optional
        The blank feedback sheet. A distributed sheet whose grade cell still
        shows what this one shows has not been marked, however much of a
        number it is. Without it, any number counts as a mark -- see
        :func:`_is_a_mark`.
    marks_out_of : float
        The scale to draw on. The distribution and the planted boundary
        marks are both fractions of it.
    seed, boundaries, discrepancies, overwrite, dry_run, marks
        As :func:`simulate_marking`.

    Returns
    -------
    SimulatedMarking
        What was written, skipped and planted.

    Examples
    --------
    ::

        import pathlib as pl        # pl is pathlib

        simulate_marking_in(
            pl.Path("~/marking/PS4001/cw1/submissions").expanduser(),
            "D30",
            workbooks=pl.Path("~/marking/PS4001/cw1/grading_output").expanduser(),
            seed=1,
            discrepancies=2,
        )
    """
    sheets = feedback_sheets_in(submissions)
    if not sheets:
        raise ValueError(
            f"No feedback sheets under {submissions}. They are named "
            "'Feedback sheet <id>.xlsx' and distribute_feedback_sheets puts "
            "one in each folder; nothing can be marked until it has run."
        )

    # One random stream for the whole run. Two Random(seed) objects make the
    # same first choice, which had the mistyped students coming out exactly
    # equal to the boundary students on every seed.
    rng = random.Random(seed)
    drawn = (
        dict(marks)
        if marks is not None
        else draw_marks(list(sheets), marks_out_of, boundaries=boundaries, rng=rng)
    )

    if discrepancies > len(drawn):
        raise ValueError(
            f"Asked for {discrepancies} discrepancy(ies) but there are only "
            f"{len(drawn)} mark(s) to spoil."
        )

    # What the blank sheet shows, so a sheet still showing it is unmarked.
    blank_value = grade_cell_value(blank, grade_cell) if blank else None

    written: dict[str, list[pl.Path]] = {}
    skipped: dict[str, list[pl.Path]] = {}
    refused: dict[pl.Path, str] = {}
    #: What ended up on each student's sheet -- the drawn mark where one was
    #: written, and whatever was already there where it was not. This, not
    #: the draw, is what a grader copies into their workbook.
    on_sheet: dict[str, float] = {}

    for identifier, paths in sheets.items():
        for path in paths:
            existing = grade_cell_value(path, grade_cell)
            if not overwrite and _is_a_mark(existing, blank_value):
                skipped.setdefault(identifier, []).append(path)
                on_sheet.setdefault(identifier, existing)
                continue
            if not dry_run:
                try:
                    _write_sheet(path, grade_cell, drawn[identifier])
                except Exception as exc:
                    # One unwritable sheet is not a reason to abandon the
                    # other eighty, and which ones failed is the thing worth
                    # knowing afterwards.
                    refused[path] = f"{type(exc).__name__}: {exc}"
                    continue
            written.setdefault(identifier, []).append(path)
            on_sheet[identifier] = drawn[identifier]

    # The mistyped copy: the sheet says one thing, the workbook another.
    # Planted here rather than before the writing so that only marks that
    # actually reached a sheet can be mistyped -- a slip against a sheet
    # nothing was written to is a disagreement this tool invented rather
    # than one it found.
    reported = dict(on_sheet)
    planted: dict[str, tuple[float, float]] = {}
    spoilable = sorted(written)
    if discrepancies > len(spoilable):
        discrepancies = len(spoilable)
    for identifier in rng.sample(spoilable, discrepancies):
        # A transposition-sized slip, and never off the end of the scale.
        slip = rng.choice([-10, -9, -2, -1, 1, 2, 9, 10])
        wrong = min(max(on_sheet[identifier] + slip, 0.0), marks_out_of)
        if wrong == on_sheet[identifier]:
            wrong = min(on_sheet[identifier] + 1, marks_out_of)
        reported[identifier] = wrong
        planted[identifier] = (on_sheet[identifier], wrong)

    filled = (
        _fill_grader_workbooks(
            grader_workbooks(workbooks, graders), reported, dry_run=dry_run
        )
        if workbooks is not None
        else {}
    )

    return SimulatedMarking(
        marks=drawn,
        sheets=written,
        skipped=skipped,
        workbooks=filled,
        discrepancies=planted,
        refused=refused,
        dry_run=dry_run,
    )


def simulate_marking(
    assessment: Assessment,
    *,
    seed: int | None = None,
    boundaries: int = 3,
    discrepancies: int = 0,
    overwrite: bool = False,
    dry_run: bool = False,
    marks: Mapping[str, float] | None = None,
) -> SimulatedMarking:
    """
    Mark an assessment as if a grader had, on both records.

    Writes a mark into every distributed feedback sheet's grade cell, and
    the same mark into the ``Mark`` column of the grader workbook that
    student or group belongs to -- which is what a grader does, in that
    order, and what step 7 then reconciles.

    The assessment supplies the two folders, the cell and the scale;
    :func:`simulate_marking_in` is the same run against folders named
    directly, for a set of files that is not laid out as a module.

    Parameters
    ----------
    assessment
        A bound assessment, i.e. one reached through a module loaded with
        ``load_module``. It must have a ``grade_cell``.
    seed
        For a reproducible cohort.
    boundaries
        How many marks to place on band edges and awkward halves instead of
        drawing them.
    discrepancies
        How many students to have the grader **mistype** into their
        workbook. This is the failure `reconcile_marks` exists to catch, and
        without it every reconciliation trivially passes.
    overwrite
        Mark sheets that already carry a number. Off by default, so a
        part-marked assessment survives a careless run.
    dry_run
        Work out what would happen and write nothing.
    marks
        Use these marks instead of drawing any. ``{identifier: mark}``.

    Returns
    -------
    SimulatedMarking
        What was written, skipped and planted.

    Raises
    ------
    ValueError
        If the assessment has no ``grade_cell``, or if more discrepancies
        are asked for than there are marks to spoil.
    NotADirectoryError
        If the submissions folder is not there.

    Examples
    --------
    ::

        import pathlib as pl        # pl is pathlib

        module = load_module(pl.Path("~/scratch/PS4001/module.toml"))
        result = simulate_marking(module.assessment("cw1"), seed=1,
                                  discrepancies=2)
        print(result)
    """
    if not isinstance(assessment, Assessment):
        raise ValueError(
            "assessment must be an Assessment, reached through a module "
            "loaded with load_module() -- an unbound one does not know where "
            "its own files are. Point simulate_marking_in() at the folders "
            "instead if these files are not laid out as a module."
        )
    if assessment.grade_cell is None:
        raise ValueError(
            f"Assessment {assessment.id!r} has no grade_cell, so there is no "
            "cell to write a mark into. Set grade_cell in module.toml -- it "
            "is the same cell catch_grades reads back."
        )

    return simulate_marking_in(
        assessment.submissions_path,
        assessment.grade_cell,
        workbooks=assessment.grading_output_path,
        graders=[g.initials for g in assessment.graders],
        blank=assessment.rubric_path,
        marks_out_of=assessment.marks_out_of,
        seed=seed,
        boundaries=boundaries,
        discrepancies=discrepancies,
        overwrite=overwrite,
        dry_run=dry_run,
        marks=marks,
    )


def _fill_grader_workbooks(
    workbooks: Mapping[str, pl.Path],
    marks: Mapping[str, float],
    *,
    dry_run: bool = False,
) -> dict[str, int]:
    """Copy the marks into each grader's workbook, as the grader would."""
    filled: dict[str, int] = {}
    for grader, path in workbooks.items():
        frame = pd.read_excel(path, dtype={"Student ID": str})
        key = _key_column(frame)
        keys = frame[key].astype(str)
        values = keys.map(lambda k: marks.get(k))
        if MARK_COLUMN in frame.columns:
            # Only fill what the grader has not already written.
            frame[MARK_COLUMN] = frame[MARK_COLUMN].where(
                frame[MARK_COLUMN].notna(), values
            )
        else:
            frame[MARK_COLUMN] = values

        filled[grader] = int(values.notna().sum())
        if not dry_run:
            frame.to_excel(path, index=False)
    return filled


# ---------------------------------------------------------------------------
# From a terminal
# ---------------------------------------------------------------------------


def explain_sheets(
    submissions: pl.Path, grade_cell: str, blank: pl.Path | None = None
) -> pd.DataFrame:
    """What the simulator sees, sheet by sheet, and what it would do.

    Written because four separate guesses about why a real module's sheets
    were not being marked all turned out to be about what was *in* the grade
    cell -- which is a thing nobody can see from here and everybody can see
    from there. One column per question worth asking:

    ``identifier``   who the sheet is named for
    ``value``        what its grade cell holds now, as `catch_grades` reads it
    ``blank``        what the blank rubric holds, if there is one
    ``decision``     ``write`` or ``skip``, and skip says why

    Returns
    -------
    pandas DataFrame
        One row per feedback sheet found. Empty when none were.
    """
    blank_value = grade_cell_value(blank, grade_cell) if blank else None

    rows = []
    for identifier, paths in feedback_sheets_in(submissions).items():
        for path in paths:
            value = grade_cell_value(path, grade_cell)
            marked = _is_a_mark(value, blank_value)
            rows.append(
                {
                    "identifier": identifier,
                    "sheet": path.name,
                    "folder": path.parent.name,
                    "value": value,
                    "blank": blank_value,
                    "decision": (
                        "skip -- already marked" if marked else "write"
                    ),
                }
            )
    return pd.DataFrame(
        rows,
        columns=["identifier", "sheet", "folder", "value", "blank", "decision"],
    )


def _assessments_to_mark(module: Module, wanted: str | None) -> list[Assessment]:
    if wanted is not None:
        return [module.assessment(wanted)]
    # Everything a human marks on a sheet. A quiz has no feedback sheet and
    # nobody marks it, which is why its section of the walkthrough is four
    # cells against the coursework's ten.
    return [a for a in module.assessments if a.grade_cell is not None]


def _seed_for(seed: int | None, assessment: Assessment) -> int | None:
    """One seed per assessment, derived from the one the caller gave.

    Handing every assessment the same seed gives every student the same mark
    in cw1 as in cw2, which is reproducible and useless: the module total
    then equals the component, and a weighting bug looks exactly like a
    correct answer. crc32 of the id rather than ``hash``, which is salted
    per process, so a seed still means the same cohort tomorrow.
    """
    if seed is None:
        return None
    return seed + zlib.crc32(assessment.id.encode("utf-8"))


def _run_on_folders(args, parser) -> int:
    """The folder form: no module.toml, just the paths given."""
    if args.cell is None:
        parser.error(
            "--submissions needs --cell: the cell a mark goes in is the one "
            "thing about a feedback sheet that cannot be guessed from it, "
            "and it has to be the same cell catch_grades reads back."
        )

    graders = (
        [g.strip() for g in args.graders.split(",") if g.strip()]
        if args.graders
        else None
    )
    print(f"{args.submissions}  (cell {args.cell})")
    if args.explain:
        if args.write:
            print(
                "  --explain writes nothing, and --write is being ignored. "
                "Run it again without --explain to mark anything."
            )
        _explain(explain_sheets(args.submissions, args.cell, args.blank))
        return 0
    if args.workbooks is None:
        print(
            "  no --workbooks: marking the feedback sheets only, so there is "
            "one record and nothing to reconcile it against"
        )

    try:
        result = simulate_marking_in(
            args.submissions,
            args.cell,
            workbooks=args.workbooks,
            graders=graders,
            blank=args.blank,
            marks_out_of=args.marks_out_of,
            seed=args.seed,
            boundaries=args.boundaries,
            discrepancies=args.discrepancies,
            overwrite=args.overwrite,
            dry_run=not args.write,
        )
    except (ValueError, NotADirectoryError) as exc:
        print(f"  {exc}")
        return 1

    _report(result)
    if not args.write:
        print("\nNothing was written. Pass --write to do it for real.")
    return 0


def _explain(frame: pd.DataFrame, limit: int = 12) -> None:
    """Print what explain_sheets found, without needing pandas to be pretty."""
    if frame.empty:
        print(
            "  No feedback sheets found. They are named "
            "'Feedback sheet <id>.xlsx' and sit one per student folder."
        )
        return

    counts = frame["decision"].value_counts().to_dict()
    print(f"  {len(frame)} sheet(s): {counts}")
    for _, row in frame.head(limit).iterrows():
        print(
            f"    {row['identifier']:<14} {row['sheet']:<34} "
            f"cell={row['value']!r:<10} blank={row['blank']!r:<10} "
            f"{row['decision']}"
        )
    if len(frame) > limit:
        print(f"    ... and {len(frame) - limit} more")


def _report(result: SimulatedMarking, label: str = "") -> None:
    """One run, on stdout. The planted discrepancies are listed because
    they are the point: they are what reconciliation should then find."""
    print(f"  {label}{result}")
    for identifier, (sheet, workbook) in sorted(result.discrepancies.items()):
        print(f"      planted: {identifier} sheet {sheet} vs workbook {workbook}")
    for path, why in sorted(result.refused.items()):
        print(f"      refused: {path.parent.name}/{path.name} -- {why}")


def main(argv: "Sequence[str] | None" = None) -> int:
    parser = argparse.ArgumentParser(
        # No prog=: argparse takes it from argv[0], so the usage line names
        # however it was actually invoked -- `simulate-marking` for the
        # console script, the -m form for the module.
        description=(
            "Fill in marks nobody marked, so the rest of the pipeline can be "
            "run. Writes into real files -- point it at a scratch copy."
        ),
    )
    parser.add_argument(
        "module", type=pl.Path, nargs="?", default=None,
        help="A module folder, or its module.toml. Leave it out and use "
             "--submissions to work on folders that are not laid out as a "
             "module.",
    )
    parser.add_argument(
        "-a", "--assessment", default=None,
        help="Which assessment. Default: every one with a grade_cell.",
    )

    folders = parser.add_argument_group(
        "pointing at folders",
        "For a set of files that is not a module -- an unzipped download and, "
        "if you have them, the grader workbooks beside it.",
    )
    folders.add_argument(
        "-s", "--submissions", type=pl.Path, default=None,
        help="The unzipped download, with a feedback sheet in each folder.",
    )
    folders.add_argument(
        "-w", "--workbooks", type=pl.Path, default=None,
        help="The folder of grader workbooks. Without it only the feedback "
             "sheets are marked, and reconciliation has nothing to compare.",
    )
    folders.add_argument(
        "-c", "--cell", default=None,
        help="The cell in the feedback sheet the mark goes in, e.g. D30. "
             "Required with --submissions.",
    )
    folders.add_argument(
        "--graders", default=None,
        help="Whose workbooks to fill, comma separated (KOM,SOB). Default: "
             "every workbook in --workbooks.",
    )
    folders.add_argument(
        "--blank", type=pl.Path, default=None,
        help="The blank feedback sheet. A distributed sheet still showing "
             "what it shows has not been marked -- without it, any number in "
             "the grade cell counts as a mark.",
    )
    folders.add_argument(
        "--marks-out-of", type=float, default=100,
        help="The scale to draw marks on. Default 100.",
    )
    parser.add_argument(
        "--write", action="store_true",
        help="Actually write. Without it you get the plan and nothing else.",
    )
    parser.add_argument("--seed", type=int, default=None, help="For a reproducible cohort.")
    parser.add_argument(
        "--boundaries", type=int, default=3,
        help="Marks placed on band edges and awkward halves. Default 3.",
    )
    parser.add_argument(
        "-d", "--discrepancies", type=int, default=0,
        help="Marks the grader mistypes into their workbook, for reconciliation.",
    )
    parser.add_argument(
        "--overwrite", action="store_true",
        help="Mark sheets that already carry a number.",
    )
    parser.add_argument(
        "--explain", action="store_true",
        help="List every feedback sheet found, what its grade cell holds, "
             "and whether it would be written or skipped. Writes nothing.",
    )
    args = parser.parse_args(argv)

    if (args.module is None) == (args.submissions is None):
        parser.error(
            "Give either a module folder or --submissions, not both and not "
            "neither. A module knows where its own submissions are; "
            "--submissions is for files that are not laid out as one."
        )

    if args.submissions is not None:
        return _run_on_folders(args, parser)

    module = load_module(args.module)
    print(f"{module.code} -- {module.name}  ({module.root})")

    if args.explain:
        if args.write:
            print(
                "\n  --explain writes nothing, and --write is being ignored. "
                "Run it again without --explain to mark anything."
            )
        for assessment in _assessments_to_mark(module, args.assessment):
            print(f"\n  {assessment.id}  (grade cell {assessment.grade_cell})")
            print(f"  submissions: {assessment.submissions_path}")
            print(f"  blank sheet: {assessment.rubric_path}")
            _explain(
                explain_sheets(
                    assessment.submissions_path,
                    assessment.grade_cell,
                    assessment.rubric_path,
                )
            )
        return 0

    for assessment in _assessments_to_mark(module, args.assessment):
        try:
            result = simulate_marking(
                assessment,
                seed=_seed_for(args.seed, assessment),
                boundaries=args.boundaries,
                discrepancies=args.discrepancies,
                overwrite=args.overwrite,
                dry_run=not args.write,
            )
        except (ValueError, NotADirectoryError) as exc:
            print(f"  {assessment.id}: skipped -- {exc}")
            continue

        _report(result, label=f"{assessment.id}: ")

    if not args.write:
        print("\nNothing was written. Pass --write to do it for real.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
