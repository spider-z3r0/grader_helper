#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""Reshape the departmental grade sheet for whatever assessments a module has.

The department's template is one module's shape -- two courseworks and an MCQ
-- written out by hand. Any other shape has been reshaped by hand too, and
that is where marks go wrong: the descriptives at A23 and the Letter Grade
column both have to be re-pointed at the new block, and nothing complains when
they are not. A summary that quietly omits the last assessment, or a
distribution counting an empty column, looks exactly like a correct one.

So this builds the block instead. It **writes formulas, never values**: the
weighting, the total, the letter grade, the descriptives and the distribution
all go in as Excel formulas, transcribed from the template's own, so the sheet
still does its own arithmetic and stays the thing our numbers are checked
against. The band table, the QPV column and the other four sheets are read but
never touched.

The guard that this transcription is faithful is
``tests/test_departmental_sheet.py``: given a module of the template's own
shape, the generated sheet has to reproduce the committed template cell for
cell. If the generalisation is wrong, that fails.

House convention: ``pl`` is pathlib, ``pr`` is polars.

    >>> import pathlib as pl
    >>> import polars as pr
"""

import pathlib as pl
import re
from copy import copy

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.worksheet import Worksheet

from ..models import Module
from .departmental_layout import (
    BAND_LETTER_COLUMN,
    COMMENTS_COLUMN,
    DISTRIBUTION_COUNT_COLUMN,
    DISTRIBUTION_GRADE_COLUMN,
    FIRST_BAND_ROW,
    FIRST_DATA_ROW,
    FIRST_DISTRIBUTION_ROW,
    HEADER_ROW,
    ID_COLUMN,
    LAST_BAND_ROW,
    LAST_DATA_ROW,
    LAST_DISTRIBUTION_ROW,
    LETTER_COLUMN,
    MEAN_ROW,
    N_ROW,
    NAME_COLUMN,
    SD_ROW,
    SHEET_NAME,
    TOTAL_COLUMN,
    DepartmentalLayout,
)

#: A grade-sheet header, e.g. ``Coursework 1 (100)``. The stem is what pairs a
#: raw column with its weighted one, which is how the template's own columns
#: are classified so their styling can be reused.
_HEADER = re.compile(r"^(?P<stem>.+?)\s*\((?P<number>[\d.]+)\)$")

#: Roles a column can play, and the styling each one carries.
_RAW, _WEIGHTED, _TOTAL, _LETTER, _COMMENTS = (
    "raw",
    "weighted",
    "total",
    "letter",
    "comments",
)


def build_departmental_sheet(
    module: Module,
    template: pl.Path | str,
    destination: pl.Path | str,
    sheet: str = SHEET_NAME,
    overwrite: bool = False,
) -> pl.Path:
    """
    Write a copy of the departmental template laid out for this module.

    Rebuilds the assessment block from the module's assessments and re-points
    everything that reads it -- the total, the letter grade, the descriptives
    at A23 and the grade distribution -- then clears the template's sample
    rows. The result is an empty sheet of the right shape, ready for
    `write_departmental_sheet` to put marks into.

    Args:
    module (Module): The module whose assessments define the block.
    template (pl.Path | str): The department's workbook, read and not modified.
    destination (pl.Path | str): Where the reshaped copy is written.
    sheet (str): The tab to rebuild. Defaults to 'GradeTemplate'.
    overwrite (bool): Replace `destination` if it exists. Defaults to False,
        because an existing sheet of that name may already hold marks.

    Returns:
    pl.Path: The path written.

    Note:
        Only formulas are written, never computed values. The sheet keeps
        doing its own arithmetic off its own band table, so it remains the
        source of truth that `prepare_data_for_departmental_template` is
        checked against, rather than a transcript of our answers.

        A module of the template's own shape -- two courseworks and an MCQ --
        produces the template back, which is what makes the reshaping
        trustworthy for the shapes it has never seen.

    Raises:
    FileNotFoundError: If `template` does not exist.
    FileExistsError: If `destination` exists and `overwrite` is False.
    KeyError: If the workbook has no sheet called `sheet`.
    ValueError: If the module declares no assessments, or the template's
        layout cannot be read.

    Example:
        >>> import pathlib as pl
        >>> module = load_module("module.toml")
        >>> build_departmental_sheet(
        ...     module,
        ...     pl.Path("Dept grade sheet Template 2026.xlsx"),
        ...     module.root / "PS4001 grades.xlsx",
        ... )
    """
    template = pl.Path(template)
    destination = pl.Path(destination)

    if not template.is_file():
        raise FileNotFoundError(f"No departmental template at {template}")
    if destination.exists() and not overwrite:
        raise FileExistsError(
            f"{destination} already exists, and it may already hold marks. "
            "Pass overwrite=True if you are sure you want to replace it."
        )

    layout = DepartmentalLayout.for_module(module)

    workbook = load_workbook(template)
    if sheet not in workbook.sheetnames:
        raise KeyError(
            f"{template.name} has no sheet called {sheet!r}. It holds: "
            f"{workbook.sheetnames}"
        )
    worksheet = workbook[sheet]

    band_letters = _read_band_letters(worksheet)
    grades = _read_distribution_grades(worksheet)
    styles = _template_styles(worksheet)
    widths = _template_widths(worksheet)

    previous_last_column = max(worksheet.max_column, layout.last_column)

    _clear_block(worksheet, previous_last_column)
    _write_headers(worksheet, layout, styles)
    _write_row_formulas(worksheet, layout, band_letters, styles)
    _write_descriptives(worksheet, layout, styles)
    _write_distribution(worksheet, layout, grades)
    _set_widths(worksheet, layout, widths, previous_last_column)

    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination


# --------------------------------------------------------------- reading the
# template. Everything the builder reuses is read off the file rather than
# copied into this module, so a template the department revises carries its
# revisions through.


def _read_band_letters(worksheet: Worksheet) -> list[str]:
    """The grade letters from the band table, ascending from row 8."""
    letters = [
        worksheet.cell(row=row, column=BAND_LETTER_COLUMN).value
        for row in range(FIRST_BAND_ROW, LAST_BAND_ROW + 1)
    ]
    if not all(isinstance(letter, str) and letter for letter in letters):
        raise ValueError(
            f"The band table at rows {FIRST_BAND_ROW}-{LAST_BAND_ROW} does not "
            f"hold a grade letter in every row of column {BAND_LETTER_COLUMN}: "
            f"{letters}. The letter-grade formula is generated from it, so it "
            "cannot be built from an incomplete table."
        )
    return letters  # type: ignore[return-value]


def _read_distribution_grades(worksheet: Worksheet) -> list[tuple[int, str]]:
    """The grades the distribution block counts, with the row each sits on."""
    found = []
    for row in range(FIRST_DISTRIBUTION_ROW, LAST_DISTRIBUTION_ROW + 1):
        grade = worksheet.cell(row=row, column=DISTRIBUTION_GRADE_COLUMN).value
        if isinstance(grade, str) and grade:
            found.append((row, grade))
    if not found:
        raise ValueError(
            "The distribution block names no grades, so its counts cannot be "
            f"rebuilt. Expected them in column {DISTRIBUTION_GRADE_COLUMN}, "
            f"rows {FIRST_DISTRIBUTION_ROW}-{LAST_DISTRIBUTION_ROW}."
        )
    return found


def _template_roles(worksheet: Worksheet) -> dict[str, int]:
    """Which template column plays each role, so its styling can be reused.

    Classified from the template's own header row by the two-numbers rule --
    a stem appearing twice is a raw column followed by its weighted one, a
    stem appearing once is raw and needs no weighting. That is the same rule
    `Assessment` derives its column names by, so a template the department
    reshapes is still read correctly.
    """
    headers = [
        worksheet.cell(row=HEADER_ROW, column=column).value
        for column in range(1, worksheet.max_column + 1)
    ]
    try:
        total_index = headers.index(TOTAL_COLUMN) + 1
    except ValueError:
        raise ValueError(
            f"The template's header row {HEADER_ROW} has no {TOTAL_COLUMN!r} "
            f"column, so its layout cannot be read. It holds: {headers}"
        ) from None

    roles = {
        NAME_COLUMN: 1,
        ID_COLUMN: 2,
        _TOTAL: total_index,
        _LETTER: total_index + 1,
        _COMMENTS: total_index + 2,
    }

    # The assessment block sits between the id and the total. Walk it counting
    # how often each stem has been seen: the second sighting is the weighted
    # column, every first sighting is a raw one.
    seen: set[str] = set()
    for column in range(3, total_index):
        header = headers[column - 1]
        if not isinstance(header, str):
            continue
        match = _HEADER.match(header)
        stem = match.group("stem") if match else header
        role = _WEIGHTED if stem in seen else _RAW
        seen.add(stem)
        roles.setdefault(role, column)

    if _RAW not in roles:
        raise ValueError(
            f"The template's header row {HEADER_ROW} has no assessment "
            f"columns between {ID_COLUMN!r} and {TOTAL_COLUMN!r}, so there is "
            "no marks column to take styling from."
        )
    return roles


def _body_exemplar_row(worksheet: Worksheet) -> int:
    """The first ruled row the department left empty, styling taken from there.

    Not row 30, and this is worth being exact about. The template styles its
    sample rows 30-49 with a number format of ``0`` and the 481 untouched rows
    below them with ``0.00``. The samples are the odd ones out, and wrongly so:
    ``E30`` holds ``66.5`` and a format of ``0`` displays it as ``67``. A mark
    that reads as a whole number when it is not is precisely the confusion this
    package exists to prevent, so the blank rows -- the department's own
    formatting for a row nobody has touched -- are the exemplar.

    Some sample rows also carry a highlight fill, which on a real student would
    read as a flag from the module leader.
    """
    named = [
        row
        for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1)
        if worksheet.cell(row=row, column=1).value not in (None, "")
    ]
    return min(max(named) + 1, LAST_DATA_ROW) if named else FIRST_DATA_ROW


def _template_styles(worksheet: Worksheet) -> dict[str, object]:
    """One exemplar cell style per role, taken from the template itself.

    Copying the department's own styling is both less code and more faithful
    than restating their borders and number formats here.
    """
    roles = _template_roles(worksheet)
    body_row = _body_exemplar_row(worksheet)
    styles: dict[str, object] = {}

    for role, column in roles.items():
        styles[role] = copy(worksheet.cell(row=body_row, column=column)._style)

    if _WEIGHTED not in styles:
        # A template with nothing to weight still has to style a weighted
        # column if the module needs one. The raw styling with two decimals is
        # the sheet's own distinction between a mark and a contribution.
        weighted = copy(worksheet.cell(row=body_row, column=roles[_RAW])._style)
        styles[_WEIGHTED] = weighted

    # The header row's first cell carries a left border the rest do not, so
    # both are kept.
    styles["header_first"] = copy(worksheet.cell(row=HEADER_ROW, column=1)._style)
    styles["header"] = copy(worksheet.cell(row=HEADER_ROW, column=2)._style)

    # The descriptives are styled per row, not per column.
    for row, key in ((MEAN_ROW, "mean"), (SD_ROW, "sd"), (N_ROW, "count")):
        styles[key] = copy(worksheet.cell(row=row, column=roles[_RAW])._style)

    return styles


def _template_widths(worksheet: Worksheet) -> dict[str, object]:
    """The template's column widths, ready to lay over a block of any width.

    Widths are not a property of a column's *role* -- the template's own
    assessment block runs 19.57, 18, 18, 18, default, which is what dragging a
    boundary leaves behind rather than a rule anyone stated. So the block's
    widths are kept in order and reapplied positionally, and a block longer
    than the template's continues with the last real width it had. The
    identifying and derived columns, which do have stable roles, are kept by
    role.
    """
    roles = _template_roles(worksheet)

    def width(column: int) -> float | None:
        dimension = worksheet.column_dimensions.get(get_column_letter(column))
        return dimension.width if dimension is not None and dimension.width else None

    block = [width(column) for column in range(3, roles[_TOTAL])]
    return {
        NAME_COLUMN: width(roles[NAME_COLUMN]),
        ID_COLUMN: width(roles[ID_COLUMN]),
        _TOTAL: width(roles[_TOTAL]),
        _LETTER: width(roles[_LETTER]),
        _COMMENTS: width(roles[_COMMENTS]),
        "block": block,
    }


# ------------------------------------------------------------------- writing


def _clear_block(worksheet: Worksheet, last_column: int) -> None:
    """Empty everything the rebuild owns, so nothing stale survives it.

    The whole block goes, not just the cells about to be rewritten: a module
    with fewer assessments than the template produces a *narrower* sheet, and
    the columns falling off the right-hand end would otherwise keep their old
    headers and formulas and be read as real.

    Styling is reset with them. The template highlights some of its sample
    rows, and a highlight left behind on a real student reads as a flag from
    the module leader.
    """
    for row in range(HEADER_ROW, LAST_DATA_ROW + 1):
        for column in range(1, last_column + 1):
            cell = worksheet.cell(row=row, column=column)
            cell.value = None
            cell.style = "Normal"

    # The descriptives keep their labels in column A and lose everything else.
    for row in (MEAN_ROW, SD_ROW, N_ROW):
        for column in range(2, last_column + 1):
            cell = worksheet.cell(row=row, column=column)
            cell.value = None
            cell.style = "Normal"


def _write_headers(
    worksheet: Worksheet, layout: DepartmentalLayout, styles: dict
) -> None:
    for index, header in enumerate(layout.headers, start=1):
        cell = worksheet.cell(row=HEADER_ROW, column=index)
        cell.value = header
        cell._style = copy(styles["header_first" if index == 1 else "header"])


def _role_of(layout: DepartmentalLayout, header: str) -> str:
    """Which styling a generated column takes."""
    if header in (NAME_COLUMN, ID_COLUMN):
        return header
    if header == TOTAL_COLUMN:
        return _TOTAL
    if header == LETTER_COLUMN:
        return _LETTER
    if header == COMMENTS_COLUMN:
        return _COMMENTS
    if header in layout.raw_headers:
        return _RAW
    return _WEIGHTED


def _write_row_formulas(
    worksheet: Worksheet,
    layout: DepartmentalLayout,
    band_letters: list[str],
    styles: dict,
) -> None:
    """The per-student formulas, in every ruled row.

    The template carries them in all 501 rows whether or not a student sits
    there, and that is reproduced: a formula waiting in row 400 is what lets a
    late addition be typed in without anyone having to know how the sheet
    works.
    """
    name_index = layout.index_for(NAME_COLUMN)
    id_index = layout.index_for(ID_COLUMN)

    for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
        for index, header in enumerate(layout.headers, start=1):
            cell = worksheet.cell(row=row, column=index)
            cell._style = copy(styles[_role_of(layout, header)])
            if index in (name_index, id_index) or header in layout.raw_headers:
                continue  # a human's to fill, or write_departmental_sheet's
            if header == TOTAL_COLUMN:
                cell.value = layout.total_formula(row)
            elif header == LETTER_COLUMN:
                cell.value = layout.letter_grade_formula(row, band_letters)
            elif header == COMMENTS_COLUMN:
                continue

        for assessment in layout.module.assessments:
            if assessment.weighted_column is None:
                continue
            worksheet.cell(
                row=row, column=layout.index_for(assessment.weighted_column)
            ).value = layout.weighting_formula(assessment, row)


def _write_descriptives(
    worksheet: Worksheet, layout: DepartmentalLayout, styles: dict
) -> None:
    """Mean, SD and N under every assessment column and the total.

    One formula per column, which is exactly why this goes wrong by hand: add
    an assessment and the summary needs another three cells that nothing
    reminds you about, and a mean over five of six components is a perfectly
    plausible number.
    """
    worksheet.cell(row=MEAN_ROW, column=1).value = "Mean"
    worksheet.cell(row=SD_ROW, column=1).value = "SD"
    worksheet.cell(row=N_ROW, column=1).value = "N"

    for header in layout.summary_headers:
        column = layout.index_for(header)

        mean = worksheet.cell(row=MEAN_ROW, column=column)
        mean.value = layout.mean_formula(header)
        mean._style = copy(styles["mean"])

        sd_formula = layout.sd_formula(header)
        sd = worksheet.cell(row=SD_ROW, column=column)
        sd.value = ArrayFormula(ref=sd.coordinate, text=sd_formula)
        sd._style = copy(styles["sd"])

        count = worksheet.cell(row=N_ROW, column=column)
        count.value = layout.count_formula(header)
        count._style = copy(styles["count"])


def _write_distribution(
    worksheet: Worksheet, layout: DepartmentalLayout, grades: list[tuple[int, str]]
) -> None:
    """Re-point the grade counts at wherever the Letter Grade column landed."""
    for row, grade in grades:
        worksheet.cell(
            row=row, column=DISTRIBUTION_COUNT_COLUMN
        ).value = layout.distribution_formula(grade)


def _set_widths(
    worksheet: Worksheet,
    layout: DepartmentalLayout,
    widths: dict[str, object],
    previous_last_column: int,
) -> None:
    block: list = list(widths["block"])  # type: ignore[arg-type]
    trailing = next((w for w in reversed(block) if w), None)
    block_start = layout.index_for(layout.headers[2])

    for index, header in enumerate(layout.headers, start=1):
        if header in (NAME_COLUMN, ID_COLUMN):
            width = widths.get(header)
        elif header in (TOTAL_COLUMN, LETTER_COLUMN, COMMENTS_COLUMN):
            width = widths.get(_role_of(layout, header))
        else:
            offset = index - block_start
            width = block[offset] if offset < len(block) else trailing
        letter = get_column_letter(index)
        if width:
            worksheet.column_dimensions[letter].width = width
        else:
            worksheet.column_dimensions.pop(letter, None)

    # Anything to the right of the new block is not ours any more. The
    # template carries a width on an empty column past the comments, which
    # would otherwise show up as a stray wide column in a narrower sheet.
    for index in range(layout.last_column + 1, previous_last_column + 2):
        worksheet.column_dimensions.pop(get_column_letter(index), None)
