#!/usr/bin/env python
# -*- coding: utf-8 -*-

import logging
import re
from ..dependencies import xw, pythoncom, pd, pl
from openpyxl import load_workbook


_A1_RE = re.compile(r"^([A-Za-z]+)(\d+)$")


def _read_openpyxl_value(path: pl.Path, cell: str):
    """Return cached (last-saved) value of `cell` using openpyxl; None if absent."""
    wb = None
    try:
        wb = load_workbook(filename=str(path), data_only=True, read_only=True)
        ws = wb.worksheets[0]
        return ws[cell].value
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass


def _read_calamine_value(path: pl.Path, cell: str):
    """Use pandas+calamine (good for .xlsb/.xls). Returns cached value; no recalc."""
    m = _A1_RE.match(cell)
    if not m:
        return None
    col_letters, row_str = m.groups()
    row_idx = int(row_str)
    try:
        df = pd.read_excel(
            str(path),
            engine="calamine",       # requires pandas+calamine installed
            sheet_name=0,
            usecols=col_letters,     # restrict to the column
            header=None,
            nrows=1,
            skiprows=row_idx - 1,    # zero-based
        )
        return df.iat[0, 0]
    except Exception:
        return None


def _read_xlwings_value(path: pl.Path, cell: str):
    """Force recalc via Excel/COM and read live value."""
    app = None
    com_inited = False
    try:
        pythoncom.CoInitialize()
        com_inited = True
        app = xw.App(visible=False, add_book=False)
        app.display_alerts = False
        app.screen_updating = False
        try:
            # xlCalculationAutomatic = -4105
            app.api.Calculation = -4105
        except Exception:
            pass

        wb = app.books.open(str(path))
        try:
            # Full recalc to refresh cached formula results if needed
            try:
                app.api.CalculateFull()
            except Exception:
                pass
            sht = wb.sheets[0]
            return sht[cell].value
        finally:
            wb.close()
    finally:
        if app is not None:
            try:
                app.quit()
            except Exception:
                pass
        if com_inited:
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass


def extract_studentid_grade(file_path: pl.Path, cell: str, *, allow_xlwings_fallback: bool = True):
    """
    Read a single cell from a feedback workbook and return (student_id, value).

    Strategy:
      1) If .xlsx/.xlsm -> openpyxl (cached value, fast)
      2) If .xlsb/.xls   -> pandas+calamine (cached value)
      3) If value is None or a formula-like string and fallback allowed -> xlwings (recalc)
    """
    try:
        # Infer student id from filename: last space-separated token of stem
        student_id = pl.Path(file_path).stem.split(" ")[-1]

        suffix = file_path.suffix.lower()
        val = None

        if suffix in {".xlsx", ".xlsm"}:
            val = _read_openpyxl_value(file_path, cell)
        elif suffix in {".xlsb", ".xls"}:
            val = _read_calamine_value(file_path, cell)
        else:
            # Unknown format → try openpyxl anyway; if fails, xlwings below
            val = _read_openpyxl_value(file_path, cell)

        # If we got a usable value (not a formula string), return it
        if val is not None and not (isinstance(val, str) and val.startswith("=")):
            return student_id, val

        # Optional: fallback to live Excel recalc for stale/missing formula results
        if allow_xlwings_fallback:
            val2 = _read_xlwings_value(file_path, cell)
            if val2 is not None:
                return student_id, val2

        # If everything failed, log and signal skip
        logging.warning(f"No value read from {file_path} @ {cell} (suffix={suffix})")
        return None

    except Exception as e:
        logging.error(f"Error processing file '{file_path}': {e}")
        return None

# def extract_studentid_grade(file_path: pl.Path, cell: str):
#     """
#     Open an .xlsx feedback file with Excel (xlwings/COM), read `cell`, and
#     return (student_id, grade). Student ID = last space-separated token of stem.
#     """
#     app = None
#     com_inited = False
#     try:
#         pythoncom.CoInitialize()
#         com_inited = True
#
#         # Single hidden Excel instance per call; alerts off to avoid modals
#         app = xw.App(visible=False, add_book=False)
#         app.display_alerts = False
#         app.screen_updating = False
#
#         wb = app.books.open(str(file_path))
#         try:
#             sheet = wb.sheets[0]
#             student_id = pl.Path(file_path).stem.split(" ")[-1]
#             grade = sheet[cell].value
#         finally:
#             wb.close()
#
#         return student_id, grade
#
#     except FileNotFoundError:
#         logging.error(f"File not found: {file_path}")
#         return None
#     except IndexError as e:
#         logging.error(f"Error extracting student ID from filename '{file_path}': {e}")
#         return None
#     except Exception as e:
#         logging.error(f"Error processing file '{file_path}': {e}")
#         return None
#     finally:
#         if app is not None:
#             try:
#                 app.quit()
#             except Exception:
#                 pass
#         if com_inited:
#             try:
#                 pythoncom.CoUninitialize()
#             except Exception:
#                 pass
