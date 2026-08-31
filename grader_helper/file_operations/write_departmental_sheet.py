#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""Put a module's marks into the departmental grade sheet.

Writes **three kinds of cell and no others**: the name, the student id, and
each assessment's mark as awarded. The weighted columns, the total and the
letter grade are the sheet's own formulas and are left alone -- filling them in
would replace the department's arithmetic with ours, which is backwards. The
sheet is the source of truth precisely because anyone can open it and see it
compute its own answer; our copy of the calculation exists to be checked
against it, not to pre-empt it.

Columns are found by matching the sheet's header row against each assessment's
``raw_column``. That needs no new convention: row 29 of the template literally
reads ``Coursework 1 (100)`` and ``MCQ (10)``, which is what
``Assessment.raw_column`` produces from ``marks_out_of``.

The refusals both guard the same failure -- a total that is quietly missing a
component and therefore looks like a real mark:

* an assessment the sheet has no column for is named and refused, rather than
  being dropped;
* a column the sheet totals that the module does not account for is named and
  refused, because leaving it empty deflates every student's total.

House convention: ``pl`` is pathlib, ``pr`` is polars.

    >>> import pathlib as pl
    >>> import polars as pr
"""

import pathlib as pl
from typing import NamedTuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..dependencies import pd
from ..models import Module
from .departmental_layout import (
    FIRST_DATA_ROW,
    HEADER_ROW,
    ID_COLUMN,
    LAST_DATA_ROW,
    NAME_COLUMN,
    SHEET_NAME,
    TOTAL_COLUMN,
)

#: How many students the sheet has room for.
CAPACITY = LAST_DATA_ROW - FIRST_DATA_ROW + 1


class DepartmentalWrite(NamedTuple):
    """What `write_departmental_sheet` did.

    A bare path said the function had run; it did not say whether anything
    reached the sheet. `written` is what lets a caller -- and
    `ModuleFile.record` -- tell a finished step from one that merely
    completed.
    """

    #: The workbook written.
    path: pl.Path
    #: How many students' rows were filled in.
    written: int

    def __str__(self) -> str:
        return f"{self.path.name}: {self.written} students"


def write_departmental_sheet(
    df: pd.DataFrame,
    module: Module,
    workbook: pl.Path | str,
    destination: pl.Path | str | None = None,
    sheet: str = SHEET_NAME,
) -> DepartmentalWrite:
    """
    Write a module's marks into a departmental grade sheet.

    Args:
    df (pd.DataFrame): The marks, holding 'Name', 'Student ID' and one column
        per assessment named by its `raw_column`. Anything else in the frame --
        the weighted columns, the total, the letter grade -- is ignored,
        because the sheet computes those itself.
    module (Module): The module whose assessments say which columns to write.
    workbook (pl.Path | str): The grade sheet to write into. Usually one
        `build_departmental_sheet` has just laid out for this module.
    destination (pl.Path | str | None): Where to save. Defaults to None,
        meaning save `workbook` in place.
    sheet (str): The tab to write. Defaults to 'GradeTemplate'.

    Returns:
    DepartmentalWrite: The path written and how many students reached it.

    Note:
        Student ids are written as text. A column of digit strings that goes
        into Excel as a number comes back with its leading zeros gone --
        '00123456' becomes 123456 -- and every later merge against the class
        list then fails on a dtype mismatch.

        The ruled rows are cleared before writing, so a second run with a
        smaller cohort cannot leave a previous student stranded below the last
        row of the new one.

    Raises:
    FileNotFoundError: If `workbook` does not exist.
    KeyError: If the workbook has no sheet called `sheet`.
    ValueError: If the frame is empty or missing 'Name'/'Student ID'; if the
        sheet has no column for one of the module's assessments; if the sheet
        totals a column the module does not account for; or if there are more
        students than the sheet has rows.

    Example:
        >>> marks = collate_module_marks(module)
        >>> write_departmental_sheet(marks, module, module.root / "grades.xlsx")
    """
    if not isinstance(df, pd.DataFrame):
        raise ValueError("df must be a pandas DataFrame")
    if not isinstance(module, Module):
        raise ValueError(
            "module must be a Module. Load it with load_module(), which reads "
            "module.toml and tells this function which columns to write."
        )
    if df.empty:
        raise ValueError("DataFrame is empty, so there are no marks to write")

    workbook_path = pl.Path(workbook)
    if not workbook_path.is_file():
        raise FileNotFoundError(f"No grade sheet at {workbook_path}")

    missing = [c for c in (NAME_COLUMN, ID_COLUMN) if c not in df.columns]
    if missing:
        raise ValueError(
            f"DataFrame is missing columns: {', '.join(missing)}. The sheet "
            "identifies every row by name and student id."
        )

    book = load_workbook(workbook_path)
    if sheet not in book.sheetnames:
        raise KeyError(
            f"{workbook_path.name} has no sheet called {sheet!r}. It holds: "
            f"{book.sheetnames}"
        )
    worksheet = book[sheet]

    headers = _header_map(worksheet)
    _check_the_sheet_fits_the_module(headers, module, workbook_path)

    missing_marks = [
        a.raw_column for a in module.assessments if a.raw_column not in df.columns
    ]
    if missing_marks:
        raise ValueError(
            f"DataFrame is missing the marks for: {', '.join(missing_marks)}.\n"
            f"Module {module.code} declares {len(module.assessments)} "
            "assessment(s). Collate them with collate_module_marks(module) "
            "before writing the sheet."
        )

    if len(df) > CAPACITY:
        raise ValueError(
            f"{len(df)} students, but the sheet rules rows {FIRST_DATA_ROW} to "
            f"{LAST_DATA_ROW}, which is {CAPACITY}. Writing past the last "
            "ruled row would put marks outside every formula on the sheet -- "
            "the total, the descriptives and the grade distribution all stop "
            f"at row {LAST_DATA_ROW}. Extend the sheet by hand first."
        )

    written = [NAME_COLUMN, ID_COLUMN, *(a.raw_column for a in module.assessments)]
    _clear(worksheet, [headers[column] for column in written])

    for offset, (_, student) in enumerate(df.iterrows()):
        row = FIRST_DATA_ROW + offset
        worksheet.cell(row=row, column=headers[NAME_COLUMN]).value = _text(
            student[NAME_COLUMN]
        )
        worksheet.cell(row=row, column=headers[ID_COLUMN]).value = _text(
            student[ID_COLUMN]
        )
        for assessment in module.assessments:
            worksheet.cell(
                row=row, column=headers[assessment.raw_column]
            ).value = _mark(student[assessment.raw_column])

    saved = pl.Path(destination) if destination is not None else workbook_path
    saved.parent.mkdir(parents=True, exist_ok=True)
    book.save(saved)
    return DepartmentalWrite(path=saved, written=len(df))


def _header_map(worksheet: Worksheet) -> dict[str, int]:
    """The sheet's header row, as header text to column number."""
    headers: dict[str, int] = {}
    for column in range(1, worksheet.max_column + 1):
        value = worksheet.cell(row=HEADER_ROW, column=column).value
        if isinstance(value, str) and value.strip():
            headers.setdefault(value.strip(), column)
    if not headers:
        raise ValueError(
            f"Row {HEADER_ROW} of the sheet holds no column headers, so there "
            "is nowhere to write. Is this a departmental grade sheet?"
        )
    return headers


def _check_the_sheet_fits_the_module(
    headers: dict[str, int], module: Module, workbook_path: pl.Path
) -> None:
    """Refuse a sheet and a module that describe different assessments.

    Both directions matter and both produce the same silent failure. An
    assessment with no column is simply not written, and the total is short by
    its weight. A column the sheet totals that the module knows nothing about
    is never filled in, contributes zero, and the total is short by that
    instead. Either way every student gets a number that looks like a mark.
    """
    for column in (NAME_COLUMN, ID_COLUMN, TOTAL_COLUMN):
        if column not in headers:
            raise ValueError(
                f"{workbook_path.name} has no {column!r} column in row "
                f"{HEADER_ROW}, so it is not a departmental grade sheet this "
                "package can write."
            )

    homeless = [a for a in module.assessments if a.raw_column not in headers]
    if homeless:
        raise ValueError(
            "The sheet has no column for: "
            f"{', '.join(a.raw_column for a in homeless)}.\n"
            f"Module {module.code} needs {', '.join(module.grade_sheet_columns)}, "
            f"and {workbook_path.name} offers "
            f"{', '.join(_assessment_headers(headers))}.\n"
            "Lay a sheet out for this module with build_departmental_sheet() "
            "rather than writing a total that is missing a component."
        )

    unaccounted = [
        header
        for header in _assessment_headers(headers)
        if header not in module.grade_sheet_columns
    ]
    if unaccounted:
        raise ValueError(
            f"{workbook_path.name} totals columns this module does not "
            f"account for: {', '.join(unaccounted)}.\n"
            f"Module {module.code} declares {', '.join(module.grade_sheet_columns)}. "
            "Left empty, those columns contribute zero to every student's "
            "total, which is a wrong mark that looks like a real one. Either "
            "add the assessment to module.toml or lay the sheet out again "
            "with build_departmental_sheet()."
        )


def _assessment_headers(headers: dict[str, int]) -> list[str]:
    """The sheet's assessment block: everything between the id and the total."""
    id_column = headers[ID_COLUMN]
    total_column = headers[TOTAL_COLUMN]
    return [
        header
        for header, column in headers.items()
        if id_column < column < total_column
    ]


def _clear(worksheet: Worksheet, columns: list[int]) -> None:
    """Empty the columns about to be written, over every ruled row.

    Only the columns this function owns. The weighted, total and letter-grade
    columns hold the sheet's formulas in all 501 rows and must survive; so must
    anything a human has typed into Comments.
    """
    for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
        for column in columns:
            worksheet.cell(row=row, column=column).value = None


def _text(value) -> str | None:
    """A name or an id, as text, with missing values left blank."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    return str(value).strip()


def _mark(value) -> float | None:
    """A mark, as a number, with missing values left blank.

    A blank is not a zero. A student with no mark for one component must leave
    the cell empty so the total is visibly incomplete, rather than scoring nil
    on work that may simply not be marked yet.
    """
    if value is None or pd.isna(value):
        return None
    return float(value)
