import marimo

app = marimo.App(width="medium")

with app.setup():
    import os
    import pathlib as pl
    import sys

    import marimo as mo
    import pandas as pd
    from openpyxl import load_workbook

    sys.path.insert(0, "tests")          # so `fake_module` is importable
    from fake_module import (
        make_fake_module,
        make_second_module,
        make_third_module,
    )

    from grader_helper import (
        alphabetise_folders,
        assign_graders_individual,
        catch_grades,
        distribute_feedback_sheets,
        extract_studentid_grade,
        build_departmental_sheet,
        build_moderation_pack,
        collate_module_marks,
        collect_quiz_marks,
        import_brightspace_classlist,
        ingest_completed_graderfiles,
        prepare_data_for_departmental_template,
        read_quiz,
        save_distributed_graders,
        flag_borderline,
        read_moderation_manifest,
        sample_for_moderation,
        save_grader_sheets,
        scan_multiple_subs,
        write_departmental_sheet,
        write_si_marks,
    )
    from grader_helper.file_operations.brightspace_name_folders import (
        brightspace_name_folders,
    )
    from grader_helper.models import ModuleFile

    # Somewhere short and disposable. NOT under OneDrive -- the generated
    # folder names are long, and make_fake_module overwrites feedback sheets
    # every time it runs.
    #
    # The environment variable is only so the test suite can run this whole
    # notebook into a temporary directory instead of your home folder. Ignore
    # it; the default is the path you want.
    ROOT = pl.Path(
        os.environ.get(
            "GRADER_HELPER_SCRATCH", pl.Path.home() / "grader_helper_scratch"
        )
    ) / "PS4001"

    # The second module, which exists to show a shape the department's
    # template has no room for. Same scratch directory, beside PS4001.
    ROOT_2 = ROOT.parent / "PS4002"

    # And a third, with four assessments collected three different ways.
    ROOT_3 = ROOT.parent / "PS4003"

    # The department's own workbook, which lives in the repo. Located from
    # this file rather than the working directory, so the notebook works
    # wherever marimo is started from.
    TEMPLATE = (
        pl.Path(__file__).resolve().parent.parent
        / "Dept grade sheet Template 2026.xlsx"
    )

    # The assessments this walkthrough drives.
    ASSESSMENT_1_ID = "cw1"
    ASSESSMENT_2_ID = "cw2"
    ASSESSMENT_3_ID = "quizzes"


@app.cell
def intro():
    mo.md(
        """
        # Grading walkthrough

        One assessment, from an unzipped Brightspace download to a reconciled
        set of marks and folders renamed ready to go back up.

        The marking itself is a two-person process, and the last step is the
        control that makes it trustworthy:

        1. **Grader** fills in the feedback sheet — `grade_cell` calculates
           the score
        2. **Grader** copies that value into their own grade sheet
           (`KOM.xlsx`)
        3. **Module leader** collates the grade sheets into
           `completed_grades.xlsx`, which is a filled-in `distributed.xlsx`
        4. **Module leader** reads the feedback sheets and reconciles the two

        Step 2 is a manual copy, so steps 3 and 4 exist to catch it going
        wrong. `catch_grades` and `ingest_completed_graderfiles` are the two
        halves of that audit, not alternatives.

        Two courseworks go through all of that. The **quizzes** at the end do
        not, and the reason is worth holding on to: every step above exists
        because a human copies a number by hand. Nobody marks a quiz, so
        there is no allocation, no feedback sheet, no transcription and
        nothing to reconcile — just Brightspace's own exports, read.

        Run the cells in order. Each one prints what it did.
        """
    )
    return


@app.cell
def init_the_module():
    # make_fake_module calls init_module for us, then fills the tree with a
    # class list, submission folders and blank feedback sheets.
    #
    # distributed=False so the distribution step below has something to do.
    #
    # quizzes=True puts eleven weekly quizzes where the MCQ would be, and
    # writes Brightspace's own exports into their submissions folder. The
    # weights are unchanged -- the quiz takes the MCQ's ten marks, not extra
    # ones -- so nothing about cw1 or cw2 moves.
    #
    # DELETE ROOT before re-running if you built it before the quizzes
    # existed: module.toml stops mentioning the MCQ, but `assessments/mcq/`
    # stays on disk and is confusing to find there.
    FAKE = make_fake_module(ROOT, distributed=False, marked=False, quizzes=True)

    HANDLE = ModuleFile.load(ROOT)
    MODULE = HANDLE.module

    mo.md(f"""
    **{MODULE.code} {MODULE.name}** ({MODULE.year})

    - leader: `{MODULE.leader}`
    - root: `{MODULE.root}`
    - grade sheet columns: `{MODULE.grade_sheet_columns}`
    """)
    return FAKE, HANDLE, MODULE


@app.cell
def cw1_pick_the_assessment(MODULE):
    A = MODULE.assessment(ASSESSMENT_1_ID)
    GRADERS = [person.initials for person in A.graders]

    mo.md(f"""
    ### {A.name} (`{A.id}`)

    Marked out of {A.marks_out_of}, worth {A.weight}. Graders: `{GRADERS}`.
    Mark lives in cell `{A.grade_cell}` of each feedback sheet.

    | | |
    |---|---|
    | folder | `{A.folder_path}` |
    | submissions | `{A.submissions_path}` |
    | grading output | `{A.grading_output_path}` |
    | rubric | `{A.rubric_path}` |
    """)
    return A, GRADERS


@app.cell
def ingest_class_list(MODULE):
    cl = import_brightspace_classlist(MODULE.classlist_path)
    cl
    return (cl,)


@app.cell
def cw1_check_for_resubmissions(A):
    # alphabetise_folders refuses to rename anything while a student has more
    # than one submission folder, so find them first. Resolve by deleting the
    # ones that do not count -- which one counts is a judgement call, not
    # something the tool should make for you.
    repeated = scan_multiple_subs(A.submissions_path)

    mo.md(
        f"**{len(repeated)} student(s) submitted more than once:** "
        f"`{sorted(repeated)}`"
        if repeated
        else "**No resubmissions.** Safe to alphabetise."
    )
    return (repeated,)


@app.cell
def cw1_resolve_resubmissions(A, repeated):
    # Keeps the EARLIEST submission for each repeat. Change to [:-1] to keep
    # the latest instead. Delete this cell entirely once you would rather
    # resolve them by hand.
    for student_id in repeated:
        folders = sorted(
            p for p in A.submissions_path.iterdir()
            if p.is_dir() and student_id in p.name
        )
        for extra in folders[1:]:
            for f in extra.rglob("*"):
                if f.is_file():
                    f.unlink()
            for d in sorted(extra.rglob("*"), reverse=True):
                d.rmdir()
            extra.rmdir()

    resolved = scan_multiple_subs(A.submissions_path)
    mo.md(f"Remaining resubmissions: `{sorted(resolved)}`")
    return


@app.cell
def cw1_distribute_graders(cl, GRADERS):
    # Even split, randomised. overwrite=True lets the cell re-run.
    allocation = assign_graders_individual(cl, GRADERS, overwrite=True)

    allocation["grader"].value_counts()
    return (allocation,)


@app.cell
def cw1_save_the_grader_sheets(A, GRADERS, allocation):
    # The master allocation sits at the assessment root; the per-grader
    # workbooks go in grading_output, which is safe to delete and regenerate.
    master = save_distributed_graders(allocation, A.folder_path, overwrite=True)
    workbooks = save_grader_sheets(
        allocation,
        A.grading_output_path,
        GRADERS,
        criteria=["Mark"],
        overwrite=True,
    )

    mo.md(f"""
    - master: `{master.name}` at the assessment root
    - workbooks: `{[p.name for p in workbooks.values()]}` in `grading_output/`
    """)
    return


@app.cell
def cw1_distribute_the_feedback_sheets(A):
    # A copy of the blank rubric into every student's folder, named for their
    # id. Existing sheets are skipped, never overwritten -- they may already
    # carry marks.
    distribution = distribute_feedback_sheets(A.submissions_path, A.rubric_path)

    mo.md(f"**{distribution}** — unrecognised: `{distribution.unmatched}`")
    return (distribution,)


@app.cell
def cw1_alphabetise_the_folders(A, cl):
    # Brightspace format -> UL format:
    #   "27236-46025 - 23304308 Angood - 05 March 2026 612 PM"
    #                       becomes
    #   "ANGOOD, AOIFE(23304308)"
    #
    # Returns None. The log is written into the submissions folder, and
    # brightspace_name_folders needs it to rename them back later.
    alphabetise_folders(cl, A.submissions_path)

    rename_log = pd.read_csv(A.submissions_path / "folder_rename_log.csv")
    rename_log
    return (rename_log,)


@app.cell
def cw1_graders_complete_the_feedback_sheets(A):
    """GRADER — marks each allocated student's feedback sheet."""
    # A real feedback sheet CALCULATES the grade cell from the rubric rows
    # above it; the grader fills those in and the total falls out. These
    # synthetic sheets hold a literal instead, so this writes the value
    # straight in.
    #
    # Delete this cell when running for real. The graders do this.
    marked = 0
    for sheet in sorted(A.submissions_path.glob("*/Feedback sheet *.xlsx")):
        workbook = load_workbook(sheet)
        workbook.active[A.grade_cell] = 40 + (marked * 7) % 55
        workbook.save(sheet)
        marked += 1

    mo.md(f"Marked **{marked}** feedback sheets (cell `{A.grade_cell}`).")
    return


@app.cell
def cw1_graders_copy_marks_to_their_grade_sheets(A, GRADERS):
    """GRADER — copies each calculated mark into their own grade sheet.

    This is the transcription step, and the reason the reconciliation below
    exists. In real use a grader reads the number off the feedback sheet and
    types it into their workbook, which is exactly where a mark can go
    astray.
    """
    transcribed = {}

    sheets = {
        found.stem.split(" ")[-1]: found
        for found in A.submissions_path.glob("*/Feedback sheet *.xlsx")
    }

    for grader in GRADERS:
        grader_file = A.grading_output_path / f"{grader}.xlsx"
        allocated = pd.read_excel(grader_file, dtype={"Student ID": str})

        marks = []
        for allocated_id in allocated["Student ID"]:
            their_sheet = sheets.get(allocated_id)
            # The same reader the module leader uses below, so the two see
            # the same value -- including the fallbacks for a sheet whose
            # formula has not been recalculated.
            result = (
                extract_studentid_grade(their_sheet, A.grade_cell)
                if their_sheet
                else None
            )
            marks.append(result[1] if result else None)

        allocated["Mark"] = marks
        allocated.to_excel(grader_file, index=False)
        transcribed[grader] = int(allocated["Mark"].notna().sum())

    mo.md(f"Marks copied into each grade sheet: `{transcribed}`")
    return


@app.cell
def cw1_leader_reads_the_feedback_sheets(A):
    """MODULE LEADER — what the students actually received."""
    # Walks the submissions tree, opens every feedback sheet, reads the grade
    # cell. Student id comes from the filename, not the folder.
    grades = catch_grades(A.submissions_path, A.grade_cell)

    grades
    return (grades,)


@app.cell
def cw1_leader_collates_the_grade_sheets(A, GRADERS):
    """MODULE LEADER — every grader's sheet into one collated file.

    The result is a filled-in `distributed.xlsx`: the same allocation, with
    the marks in it. Written to `grading_output/completed_grades.xlsx`.
    """
    # require_all=True (the default) refuses if any grader has not returned
    # their file -- missing files mean missing marks. Pass require_all=False
    # to proceed with whoever has, and it warns about the rest.
    #
    # Student id columns are read as text. Left to pandas they come back as
    # int64, which drops leading zeros and makes the result unmergeable with
    # the class list.
    completed = ingest_completed_graderfiles(
        A.grading_output_path,
        GRADERS,
        file_type="excel",
        save=True,
        overwrite=True,
    )

    completed
    return (completed,)


@app.cell
def cw1_reconcile_the_two_records(completed, grades):
    """MODULE LEADER — does the student's number match the recorded one?

    The student receives the feedback sheet; the department receives the
    collated file. Between them sits a manual copy, so this is the check that
    the two agree. Run it for real -- it is the point of collating and
    catching separately.
    """
    # It also proves the ids survived the Excel round trip: a merge between an
    # object column and an int64 one does not mismatch quietly, it raises.
    #
    # Not every disagreement is a fault. A student who never submitted is
    # allocated a grader from the class list, so they appear in the collated
    # file (`right_only`) but have no feedback sheet to read a mark from. Look
    # at `_merge` before assuming something went wrong:
    #
    #   right_only  collated but not submitted -- no feedback sheet exists
    #   left_only   marked, but nobody was allocated them -- worth a look
    #   both, differing marks  a transcription slip at the copy step
    comparison = grades.merge(
        completed[["Student ID", "Mark"]], on="Student ID", how="outer",
        indicator=True,
    )
    disagreements = comparison[
        (comparison["_merge"] != "both") | (comparison["grade"] != comparison["Mark"])
    ]

    mo.md(
        f"**{len(comparison)}** students compared, "
        f"**{len(disagreements)}** disagreements."
        + ("\n\nEvery mark the students received matches the collated file."
           if disagreements.empty else "")
    )
    return (disagreements,)


@app.cell
def cw1_rename_for_reupload(A, rename_log):
    # UL format back to Brightspace format, so the folders can be re-uploaded.
    #
    # It takes the rename LOG, not the class list. The names come back exactly
    # as Brightspace wrote them, case included -- matching is
    # case-insensitive, the name written is not.
    restoration = brightspace_name_folders(rename_log, A.submissions_path)

    restored = sorted(p.name for p in A.submissions_path.iterdir() if p.is_dir())
    expected = sorted(rename_log["Original Name"])
    exact = [name for name in expected if name in restored]

    mo.md(
        f"**{restoration}**\n\n"
        f"{len(exact)} of {len(expected)} folders restored to the exact name "
        "Brightspace gave.\n\n```\n" + "\n".join(restored[:4]) + "\n```"
    )
    return


@app.cell
def cw1_record_progress(HANDLE, distribution):
    # `record` reads the flag off what the step returned. A Distribution that
    # copied nothing, or left a folder unrecognised, does *not* set
    # sheets_distributed -- a green tick against a step that did nothing is
    # the same failure as a total missing a component.
    HANDLE.record(distribution, ASSESSMENT_1_ID)

    # These two are still by hand. `save_distributed_graders` returns a path
    # and `catch_grades` a frame, and neither says whether the step finished,
    # so there is no evidence for `record` to read yet. See the notes'
    # "Keeping status".
    HANDLE.set_status(
        ASSESSMENT_1_ID,
        graders_allocated=True,
        grades_collected=True,
    )

    # Read back the assessment we just wrote, not the other one.
    status = ModuleFile.load(ROOT).module.assessment(ASSESSMENT_1_ID).status
    status.model_dump()
    return


@app.cell
def cw2_pick_the_assessment(MODULE):
    A2 = MODULE.assessment(ASSESSMENT_2_ID)
    GRADERS2 = [person.initials for person in A2.graders]

    mo.md(f"""
    ### {A2.name} (`{A2.id}`)

    Marked out of {A2.marks_out_of}, worth {A2.weight}. Graders: `{GRADERS2}`.
    Mark lives in cell `{A2.grade_cell}` of each feedback sheet.

    | | |
    |---|---|
    | folder | `{A2.folder_path}` |
    | submissions | `{A2.submissions_path}` |
    | grading output | `{A2.grading_output_path}` |
    | rubric | `{A2.rubric_path}` |
    """)
    return A2, GRADERS2


@app.cell
def cw2_check_for_resubmissions(A2):
    repeated2 = scan_multiple_subs(A2.submissions_path)

    mo.md(
        f"**{len(repeated2)} student(s) submitted more than once:** "
        f"`{sorted(repeated2)}`"
        if repeated2
        else "**No resubmissions.** Safe to alphabetise."
    )
    return (repeated2,)


@app.cell
def cw2_resolve_resubmissions(A2, repeated2):
    for student_id2 in repeated2:
        folders2 = sorted(
            p for p in A2.submissions_path.iterdir()
            if p.is_dir() and student_id2 in p.name
        )
        for extra2 in folders2[1:]:
            for f2 in extra2.rglob("*"):
                if f2.is_file():
                    f2.unlink()
            for d2 in sorted(extra2.rglob("*"), reverse=True):
                d2.rmdir()
            extra2.rmdir()

    resolved2 = scan_multiple_subs(A2.submissions_path)
    mo.md(f"Remaining resubmissions: `{sorted(resolved2)}`")
    return


@app.cell
def cw2_distribute_graders(cl, GRADERS2):
    # A fresh allocation. Who marks cw2 is independent of who marked cw1.
    allocation2 = assign_graders_individual(cl, GRADERS2, overwrite=True)

    allocation2["grader"].value_counts()
    return (allocation2,)


@app.cell
def cw2_save_the_grader_sheets(A2, GRADERS2, allocation2):
    master2 = save_distributed_graders(allocation2, A2.folder_path, overwrite=True)
    workbooks2 = save_grader_sheets(
        allocation2,
        A2.grading_output_path,
        GRADERS2,
        criteria=["Mark"],
        overwrite=True,
    )

    mo.md(f"""
    - master: `{master2.name}` at the assessment root
    - workbooks: `{[p.name for p in workbooks2.values()]}` in `grading_output/`
    """)
    return


@app.cell
def cw2_distribute_the_feedback_sheets(A2):
    # A2.rubric_path, not A.rubric_path -- each assessment has its own blank
    # sheet, and cw2's rubric lives in cw2's folder.
    distribution2 = distribute_feedback_sheets(A2.submissions_path, A2.rubric_path)

    mo.md(f"**{distribution2}** — unrecognised: `{distribution2.unmatched}`")
    return


@app.cell
def cw2_alphabetise_the_folders(A2, cl):
    alphabetise_folders(cl, A2.submissions_path)

    rename_log2 = pd.read_csv(A2.submissions_path / "folder_rename_log.csv")
    rename_log2
    return (rename_log2,)


@app.cell
def cw2_graders_complete_the_feedback_sheets(A2):
    """GRADER — marks each allocated student's feedback sheet."""
    # Delete this cell when running for real. The graders do this.
    marked2 = 0
    for sheet2 in sorted(A2.submissions_path.glob("*/Feedback sheet *.xlsx")):
        workbook2 = load_workbook(sheet2)
        workbook2.active[A2.grade_cell] = 45 + (marked2 * 5) % 50
        workbook2.save(sheet2)
        marked2 += 1

    mo.md(f"Marked **{marked2}** feedback sheets (cell `{A2.grade_cell}`).")
    return


@app.cell
def cw2_graders_copy_marks_to_their_grade_sheets(A2, GRADERS2):
    """GRADER — copies each calculated mark into their own grade sheet."""
    transcribed2 = {}

    sheets2 = {
        found2.stem.split(" ")[-1]: found2
        for found2 in A2.submissions_path.glob("*/Feedback sheet *.xlsx")
    }

    for grader2 in GRADERS2:
        grader_file2 = A2.grading_output_path / f"{grader2}.xlsx"
        allocated2 = pd.read_excel(grader_file2, dtype={"Student ID": str})

        marks2 = []
        for allocated_id2 in allocated2["Student ID"]:
            their_sheet2 = sheets2.get(allocated_id2)
            result2 = (
                extract_studentid_grade(their_sheet2, A2.grade_cell)
                if their_sheet2
                else None
            )
            marks2.append(result2[1] if result2 else None)

        allocated2["Mark"] = marks2
        allocated2.to_excel(grader_file2, index=False)
        transcribed2[grader2] = int(allocated2["Mark"].notna().sum())

    mo.md(f"Marks copied into each grade sheet: `{transcribed2}`")
    return


@app.cell
def cw2_leader_reads_the_feedback_sheets(A2):
    """MODULE LEADER — what the students actually received."""
    grades2 = catch_grades(A2.submissions_path, A2.grade_cell)

    grades2
    return (grades2,)


@app.cell
def cw2_leader_collates_the_grade_sheets(A2, GRADERS2):
    """MODULE LEADER — every grader's sheet into one collated file."""
    completed2 = ingest_completed_graderfiles(
        A2.grading_output_path,
        GRADERS2,
        file_type="excel",
        save=True,
        overwrite=True,
    )

    completed2
    return (completed2,)


@app.cell
def cw2_reconcile_the_two_records(completed2, grades2):
    """MODULE LEADER — does the student's number match the recorded one?"""
    comparison2 = grades2.merge(
        completed2[["Student ID", "Mark"]], on="Student ID", how="outer",
        indicator=True,
    )
    disagreements2 = comparison2[
        (comparison2["_merge"] != "both")
        | (comparison2["grade"] != comparison2["Mark"])
    ]

    mo.md(
        f"**{len(comparison2)}** students compared, "
        f"**{len(disagreements2)}** disagreements."
        + ("\n\nEvery mark the students received matches the collated file."
           if disagreements2.empty else "")
    )
    return (disagreements2,)


@app.cell
def cw2_rename_for_reupload(A2, rename_log2):
    # As cw1: the log, not the class list, and the names come back exactly.
    restoration2 = brightspace_name_folders(rename_log2, A2.submissions_path)

    restored2 = sorted(p.name for p in A2.submissions_path.iterdir() if p.is_dir())
    expected2 = sorted(rename_log2["Original Name"])
    exact2 = [name for name in expected2 if name in restored2]

    mo.md(
        f"**{restoration2}**\n\n"
        f"{len(exact2)} of {len(expected2)} folders restored to the exact name "
        "Brightspace gave.\n\n```\n" + "\n".join(restored2[:4]) + "\n```"
    )
    return


@app.cell
def cw2_record_progress(HANDLE):
    HANDLE.set_status(
        ASSESSMENT_2_ID,
        graders_allocated=True,
        sheets_distributed=True,
        grades_collected=True,
    )

    # Read back the assessment we just wrote, not the other one.
    status2 = ModuleFile.load(ROOT).module.assessment(ASSESSMENT_2_ID).status
    status2.model_dump()
    return


@app.cell
def quiz_intro():
    mo.md(
        """
        ## The quizzes

        Everything above happens because a grader writes a number on a
        feedback sheet and then copies it somewhere else by hand. A quiz has
        neither: Brightspace scores it, and exports one CSV per quiz.

        So the whole of steps 3–7 collapses into a single read. There is
        nothing to allocate, nothing to distribute, no second record, and so
        nothing to reconcile — the audit that makes the coursework
        trustworthy has nothing to audit here.

        What is left is the module's own rules about what a pass is, and
        those live in `module.toml` rather than in this notebook.
        """
    )
    return


@app.cell
def quiz_pick_the_assessment(MODULE):
    A3 = MODULE.assessment(ASSESSMENT_3_ID)
    quiz_exports = sorted(A3.submissions_path.glob("*.csv"))

    mo.md(f"""
    ### {A3.name} (`{A3.id}`)

    Marked out of {A3.marks_out_of}, worth {A3.weight}. No graders, no
    rubric, no grade cell — nobody marks these.

    **{len(quiz_exports)} exports** for {A3.marks_out_of} marks, so one may
    be dropped.

    | | |
    |---|---|
    | pass mark | `{A3.pass_mark}` — strictly above, so exactly {A3.pass_mark} fails |
    | free passes | `{A3.free_passes}` |
    | exports | `{A3.submissions_path}` |
    | column | `{A3.raw_column}` |

    Both numbers are read off the assessment. They are in `module.toml`, so
    the module records its own rules rather than this notebook restating
    them.
    """)
    return A3, quiz_exports


@app.cell
def quiz_look_at_one_export(quiz_exports):
    # What Brightspace actually hands over, and what read_quiz does to it:
    # the '#' comes off the username (the class list has it stripped too, and
    # a join between the two forms silently matches nothing), every column is
    # read as text so a leading zero survives, and the score column is named
    # after the file.
    one_quiz = read_quiz(quiz_exports[0]).collect()

    one_quiz
    return (one_quiz,)


@app.cell
def quiz_collect_the_marks(A3, cl):
    """MODULE LEADER — the only step there is."""
    # No pass mark and no free passes are passed in. Both come off the
    # assessment, which read them from module.toml. With neither there, this
    # raises rather than picking a threshold nobody chose.
    quiz_marks = collect_quiz_marks(A3, cl)

    quiz_marks
    return (quiz_marks,)


@app.cell
def quiz_check_the_edges(A3, quiz_marks):
    # Two students are worth looking at by name, because they are where the
    # rules are visible rather than merely applied.
    by_id = quiz_marks.set_index("Student ID")[A3.raw_column]

    mo.md(f"""
    | student | mark | why |
    |---|---|---|
    | 23304309 Joyce | **{by_id["23304309"]}** | sat no quiz at all, so appears in no export. The free pass is *not* given: the departmental sheet awards NG rather than F while the total is zero, and one free mark here would quietly make him a fail |
    | 23304308 Ivers | **{by_id["23304308"]}** | passed 9 of 11, and the free pass takes her to the {A3.marks_out_of} available. It cannot take her past it |

    Full distribution: `{by_id.value_counts().sort_index().to_dict()}`

    Every student in the class list has a row, including the ones in no
    export. A missing row would take a component out of a module total, and
    a total missing a component is still a plausible number.
    """)
    return


@app.cell
def quiz_record_progress(HANDLE):
    HANDLE.set_status(ASSESSMENT_3_ID, grades_collected=True)

    status3 = ModuleFile.load(ROOT).module.assessment(ASSESSMENT_3_ID).status
    status3.model_dump()
    return


@app.cell
def the_whole_module():
    mo.md(
        """
        ## The module

        Everything above is one assessment at a time. This is the module: the
        three sets of marks in one frame, totalled, and turned into letter
        grades.

        `collate_module_marks` is the only thing here that walks a whole
        module, and it fetches each assessment's marks from wherever that
        kind of assessment keeps them -- the courseworks off their feedback
        sheets, the quizzes out of Brightspace's exports. It decides that by
        asking what the assessment *has*, not what its `type` says, because
        an MCQ can be sat in Brightspace, on a feedback sheet, or on paper in
        a lecture theatre, and all three are `type = "mcq"`.
        """
    )
    return


@app.cell
def collate_the_whole_module(MODULE, cl):
    """MODULE LEADER -- every assessment's marks, in one frame."""
    # source="feedback" reads what the students received. The other record is
    # source="collated", the completed_grades.xlsx the graders' sheets were
    # collated into. They must agree -- that is what the reconciliation above
    # checks -- and this deliberately will not fall back from one to the
    # other, because a substitution nobody asked for is invisible.
    module_marks = collate_module_marks(MODULE, cl, source="feedback")

    module_marks
    return (module_marks,)


@app.cell
def the_departmental_sheet(MODULE, module_marks):
    """MODULE LEADER -- totalled, banded, in the department's column order."""
    module_sheet = prepare_data_for_departmental_template(module_marks, MODULE)

    module_sheet
    return (module_sheet,)


@app.cell
def look_at_the_edges(module_sheet):
    graded = module_sheet.set_index("Student ID")

    mo.md(f"""
    | student | total | grade | why |
    |---|---|---|---|
    | 23304305 Egan | {graded.loc["23304305", "Total % Grade"]} | {graded.loc["23304305", "Letter Grade"]} | totals exactly 64.5. Excel rounds half away from zero and says 65 (B2); Python's own round says 64 (B3). The sheet is the source of truth, so `excel_round` is what runs |
    | 23304309 Joyce | {graded.loc["23304309", "Total % Grade"]} | {graded.loc["23304309", "Letter Grade"]} | scored zero on everything and sat no quiz. NG, not F -- no participation is a different thing from a mark of zero |
    | 23304311 Lynch | {graded.loc["23304311", "Total % Grade"]} | {graded.loc["23304311", "Letter Grade"]} | never submitted, and still on the sheet. A student missing from the sheet has no grade at all, which nobody notices |

    That frame is *our* arithmetic. The next cell puts it into the
    department's own workbook, which is the arithmetic that counts.
    """)
    return


@app.cell
def the_departmental_workbook():
    mo.md(
        """
        ## Into the department's workbook

        The frame above is our arithmetic. The workbook is the department's,
        and where the two disagree the workbook wins -- it is the source of
        truth precisely because anyone can open it and read the grades
        without running any of this.

        So two steps, and the split matters:

        - `build_departmental_sheet` lays the sheet out for the module's
          assessments and writes the department's **formulas** into it. It
          never writes a computed value, so the sheet still does its own
          arithmetic.
        - `write_departmental_sheet` puts in the name, the student id and
          each mark as awarded. Five values a row for this module, and
          nothing else -- filling in the weighted columns or the total would
          replace the department's arithmetic with ours, which is backwards.

        PS4001 is the shape the template was drawn for, so building it is a
        no-op in effect: you get the department's own file back, minus the
        sample rows. The module after this one is the interesting case.
        """
    )
    return


@app.cell
def write_the_departmental_sheet(MODULE, module_marks):
    """MODULE LEADER -- the sheet that goes to the department."""
    sheet_path = build_departmental_sheet(
        MODULE, TEMPLATE, MODULE.root / f"{MODULE.code} grades.xlsx",
        overwrite=True,
    )
    written = write_departmental_sheet(module_marks, MODULE, sheet_path)

    mo.md(f"**{written}** — `{sheet_path}`")
    return sheet_path, written


@app.cell
def check_what_landed(sheet_path):
    # Read it back the way a colleague would -- openpyxl does not evaluate
    # formulas, so this shows what is actually stored in each cell rather
    # than what Excel makes of it.
    landed = load_workbook(sheet_path)["GradeTemplate"]

    mo.md(f"""
    Row 29 (the headers): `{[landed.cell(29, i).value for i in range(1, 11)]}`

    | cell | holds | |
    |---|---|---|
    | `A30` | `{landed["A30"].value}` | ours |
    | `C30` | `{landed["C30"].value}` | ours -- the mark as awarded |
    | `D30` | `{landed["D30"].value}` | the department's |
    | `F30` | `{landed["F30"].value}` | the department's |
    | `H30` | `{landed["H30"].value}` | the department's |

    `F30` is `=E30/2`, not `=E30/100*50`. Those are the same sum on paper and
    not in floating point: `x/2` is exact, `x/100*50` is two roundings, and
    since the total is `ROUND(SUM(...),0)` the difference lands on whole
    marks. At a coursework mark of 29 the two give 15 and 14.
    """)
    return


@app.cell
def second_module_intro():
    mo.md(
        """
        # A module the template has no room for

        Everything above is **PS4001**: two courseworks and an MCQ, which is
        exactly the permutation the departmental template was drawn for. Its
        formulas hardcode that shape -- `D=C/100*40`, `F=E/2`, and a total
        summing exactly `D, F, G`.

        **PS4002** is one coursework worth 30 and *two* MCQs worth 35 each.
        Nothing exotic; simply a shape the department's file was not drawn
        for. Until now that meant the module leader reshaped the block by
        hand, and the two places that goes wrong are the two that move when
        the block changes width:

        - the **descriptives at A23** -- Mean, SD and N, one formula per
          column, and nothing complains when the last one is missing. A mean
          over two of three components is a perfectly plausible number.
        - the **Letter Grade column** and the eleven `COUNTIF`s that read it.
          Miss one and the distribution reports the whole cohort as NG.

        Same cohort, same class list -- these are the same students taking a
        second module.
        """
    )
    return


@app.cell
def init_the_second_module():
    # Lighter than PS4001's fixture on purpose. The marking pipeline is
    # demonstrated twice above, so this writes only what is needed to
    # collate a module and build its sheet: the coursework gets feedback
    # sheets, the MCQs get nothing at all.
    SECOND = make_second_module(ROOT_2)
    MODULE_2 = ModuleFile.load(ROOT_2).module

    mo.md(f"""
    **{MODULE_2.code} {MODULE_2.name}** ({MODULE_2.year})

    - leader: `{MODULE_2.leader}`
    - weights: `{ {a.name: a.weight for a in MODULE_2.assessments} }`
    - grade sheet columns: `{MODULE_2.grade_sheet_columns}`

    {len(MODULE_2.grade_sheet_columns)} columns of assessment where the
    template has five.

    The two MCQs are marked on **different scales**, which is the real
    situation rather than a contrivance: an MCQ is sometimes graded out of 100
    and then weighted, and sometimes graded out of however many questions it
    had. MCQ 1 is out of 100 worth 35; MCQ 2 is out of 10 worth 35. Both need
    a weighted column, and the sheet will end up holding one of each form.
    """)
    return MODULE_2, SECOND


@app.cell
def second_module_collate(MODULE_2, SECOND, cl):
    """MODULE LEADER -- every assessment's marks, in one frame."""
    # The coursework is read off its feedback sheets, as ever. The two MCQs
    # were sat on paper in a lecture theatre, so nothing on disk holds them
    # and they are handed in through `marks=`. collate_module_marks decides
    # per assessment by asking what it *has*, not what its type says.
    second_by_id = SECOND.expected.set_index("Student ID")

    module_2_marks = collate_module_marks(
        MODULE_2,
        cl,
        source="feedback",
        marks={
            "mcq1": second_by_id["mcq1"].to_dict(),
            "mcq2": second_by_id["mcq2"].to_dict(),
        },
    )

    module_2_marks
    return (module_2_marks,)


@app.cell
def second_module_prepare(MODULE_2, module_2_marks):
    """MODULE LEADER -- totalled, banded, in the department's column order."""
    module_2_sheet = prepare_data_for_departmental_template(
        module_2_marks, MODULE_2
    )

    module_2_sheet
    return (module_2_sheet,)


@app.cell
def second_module_build_the_sheet(MODULE_2, module_2_marks):
    """MODULE LEADER -- and this is the part that was done by hand."""
    sheet_2_path = build_departmental_sheet(
        MODULE_2, TEMPLATE, MODULE_2.root / f"{MODULE_2.code} grades.xlsx",
        overwrite=True,
    )
    written_2 = write_departmental_sheet(module_2_marks, MODULE_2, sheet_2_path)

    mo.md(f"**{written_2}** — `{sheet_2_path}`")
    return sheet_2_path, written_2


@app.cell
def second_module_what_the_builder_did(sheet_2_path):
    built = load_workbook(sheet_2_path)["GradeTemplate"]
    headers = [built.cell(29, i).value for i in range(1, 11)]

    mo.md(f"""
    ### What would otherwise have been hand-edited

    Row 29: `{headers}`

    | | | |
    |---|---|---|
    | weighting | `{built["D30"].value}` | coursework, 100 marks worth 30 |
    | weighting | `{built["F30"].value}` | MCQ 1, graded out of 100 then weighted |
    | weighting | `{built["H30"].value}` | MCQ 2, graded out of 10 -- a scale *up* |
    | **total** | `{built["I30"].value}` | three components, not the template's `D, F, G` |
    | **letter** | `{built["J30"].value[:38]}...` | reads `I30`, not the template's `H30` |

    The descriptives moved with it -- one column each, none missed:

    | row | C | E | G | I |
    |---|---|---|---|---|
    | Mean | `{built["C23"].value}` | `{built["E23"].value}` | `{built["G23"].value}` | `{built["I23"].value}` |
    | N | `{built["C25"].value}` | `{built["E25"].value}` | `{built["G25"].value}` | `{built["I25"].value}` |

    And the distribution follows the Letter Grade column wherever it landed:
    `{built["H6"].value}`

    Two details worth reading twice. The **N row counts the raw column**, not
    its own -- `D25` counts `C`, because the weighting formula sits in all 501
    rows and counting it would return 501 whatever the cohort.

    And none of these weightings simplifies to a single divisor, because
    neither 100 nor 10 divides 35 exactly. Where one does divide -- PS4001's
    `=E30/2` -- the shorter form is used, because it is the one that is exact
    in floating point.
    """)
    return


@app.cell
def third_module_intro():
    mo.md(
        """
        # A full module: four assessments, three sources

        **PS4003** is the shape a real module tends to have — a coursework,
        weekly quizzes, an MCQ and an exam — and it is here for a different
        reason from PS4002. PS4002 was about the *sheet*: a block the template
        has no room for. This one is about the *collation*.

        Its four assessments arrive by three different routes:

        | | comes from | read by |
        |---|---|---|
        | Coursework 1 | feedback sheets in the download | `catch_grades` |
        | Quizzes | Brightspace's own exports | `collect_quiz_marks` |
        | MCQ | marked on paper | handed in via `marks=` |
        | Exam | marked on paper | handed in via `marks=` |

        PS4001 covers the first two, PS4002 the first and third. Nothing until
        now has put all three in one module, and that is the thing worth
        trying: `collate_module_marks` chooses a source **per assessment**, by
        asking what each one *has* rather than what its `type` says. A module
        with one source is not evidence that it does.

        The quizzes here are **ten quizzes for ten marks with no free pass**,
        so a student's mark is simply the number they passed. PS4001 sets
        eleven for ten and forgives one. Both sets of rules live in their own
        `module.toml`, which is the point — nothing in this notebook restates
        them.
        """
    )
    return


@app.cell
def init_the_third_module():
    THIRD = make_third_module(ROOT_3)
    MODULE_3 = ModuleFile.load(ROOT_3).module
    quizzes_3 = MODULE_3.assessment("quizzes")

    mo.md(f"""
    **{MODULE_3.code} {MODULE_3.name}** ({MODULE_3.year})

    - weights: `{ {a.name: a.weight for a in MODULE_3.assessments} }`
    - grade sheet columns: `{MODULE_3.grade_sheet_columns}`

    | | |
    |---|---|
    | quiz exports written | `{len(THIRD.quiz_exports["quizzes"])}` |
    | pass mark | `{quizzes_3.pass_mark}` — strictly above |
    | free passes | `{quizzes_3.free_passes}` |

    Seven assessment columns. The quizzes need no weighted one — ten marks
    worth ten — so a raw column sits in the middle of the block with weighted
    columns either side of it, which is the case most likely to trip a
    hand-edit.
    """)
    return MODULE_3, THIRD


@app.cell
def third_module_collate(MODULE_3, THIRD, cl):
    """MODULE LEADER -- four assessments, three sources, one call."""
    # Only the MCQ and the exam are handed in. The coursework is found by its
    # grade_cell and read off the feedback sheets; the quizzes are found by
    # the exports sitting in their submissions folder. Neither is mentioned
    # here, and that is the point being demonstrated.
    third_by_id = THIRD.expected.set_index("Student ID")

    module_3_marks = collate_module_marks(
        MODULE_3,
        cl,
        source="feedback",
        marks={
            "mcq": third_by_id["mcq"].to_dict(),
            "exam": third_by_id["exam"].to_dict(),
        },
    )

    module_3_marks
    return (module_3_marks,)


@app.cell
def third_module_check_the_sources(MODULE_3, THIRD, module_3_marks):
    # Worth confirming rather than assuming: the two columns nobody handed in
    # are the two that had to be found on disk.
    found = module_3_marks.set_index("Student ID")
    expected_3 = THIRD.expected.set_index("Student ID")
    coursework = MODULE_3.assessment("cw1").raw_column
    quizzes = MODULE_3.assessment("quizzes").raw_column

    mo.md(f"""
    | column | source | agrees with the fixture |
    |---|---|---|
    | `{coursework}` | feedback sheets | `{
        found[coursework].dropna().to_dict()
        == expected_3["cw1"].dropna().to_dict()
    }` |
    | `{quizzes}` | Brightspace exports | `{
        found[quizzes].to_dict() == expected_3["quizzes"].to_dict()
    }` |

    Neither was passed in. `collate_module_marks` found the coursework by its
    `grade_cell` and the quizzes by the exports in their submissions folder,
    in the same call that took the MCQ and exam as handed-in values.
    """)
    return


@app.cell
def third_module_prepare(MODULE_3, module_3_marks):
    """MODULE LEADER -- totalled, banded, in the department's column order."""
    module_3_sheet = prepare_data_for_departmental_template(
        module_3_marks, MODULE_3
    )

    module_3_sheet
    return (module_3_sheet,)


@app.cell
def third_module_build_the_sheet(MODULE_3, module_3_marks):
    """MODULE LEADER -- seven assessment columns, laid out and filled."""
    sheet_3_path = build_departmental_sheet(
        MODULE_3, TEMPLATE, MODULE_3.root / f"{MODULE_3.code} grades.xlsx",
        overwrite=True,
    )
    written_3 = write_departmental_sheet(module_3_marks, MODULE_3, sheet_3_path)

    mo.md(f"**{written_3}** — `{sheet_3_path}`")
    return sheet_3_path, written_3


@app.cell
def third_module_what_the_builder_did(sheet_3_path):
    wide = load_workbook(sheet_3_path)["GradeTemplate"]

    mo.md(f"""
    ### Seven assessment columns

    Row 29: `{[wide.cell(29, i).value for i in range(1, 13)]}`

    | | | |
    |---|---|---|
    | weighting | `{wide["D30"].value}` | coursework, 100 marks worth 30 |
    | *(none)* | — | quizzes, 10 marks worth 10: nothing to weight |
    | weighting | `{wide["G30"].value}` | MCQ, 100 worth 20 — **divides exactly** |
    | weighting | `{wide["I30"].value}` | exam, 100 worth 40 |
    | **total** | `{wide["J30"].value}` | four components, and `E30` is a raw column |
    | **letter** | `{wide["K30"].value[:38]}...` | reads `J30` |

    `=F30/5` is the exact-divisor form: 100 goes into 20 five times, so a
    single division does it and a single division is exact. The coursework and
    exam get `/100*30` and `/100*40` because 100 does not divide 30 or 40.

    The total is the one to look at. `E30` — the quizzes — is a **raw** column
    and reaches the total directly, sitting between two weighted ones. Summing
    the weighted columns and forgetting it, or summing every other column and
    double-counting the raw marks, are both easy hand-edits and both give a
    plausible number.

    Descriptives, one per column, none missed:

    | | C | E | F | H | J |
    |---|---|---|---|---|---|
    | Mean | `{wide["C23"].value}` | `{wide["E23"].value}` | `{wide["F23"].value}` | `{wide["H23"].value}` | `{wide["J23"].value}` |
    | N | `{wide["C25"].value}` | `{wide["E25"].value}` | `{wide["F25"].value}` | `{wide["H25"].value}` | `{wide["J25"].value}` |

    And the distribution: `{wide["H6"].value}`
    """)
    return


@app.cell
def moderation_intro():
    mo.md(
        """
        # Moderation

        A second marker looks at a sample of the marking. Two things decide
        who is in it, and the second is the one the department is currently
        arguing about:

        - **a random sample per grade band** — one student from each of A1
          down to D1, so the moderator sees the range rather than whoever is
          at the top of the list;
        - **the borderline cases** — students within a point of the next
          grade up. 69 and 70 are one mark apart and a degree classification
          apart, so if a mark is wrong it costs the student more there than
          anywhere else in the range.

        Today the borderline students are *flagged* and the draw is random.
        The department is discussing moderating on the borderline alone, so
        `sample_for_moderation` takes `borderline="include"` and is ready for
        that without a rewrite.

        **The draw records its seed.** A random sample that comes out
        different every run cannot answer "why was this student moderated?"
        six months later, and re-running quietly changes the answer. The seed
        goes into the manifest with everything else.
        """
    )
    return


@app.cell
def moderation_who_is_near_a_boundary(module_3_sheet):
    """MODULE LEADER -- who a single mark would have moved."""
    near = flag_borderline(module_3_sheet)

    near[
        ["Name", "Student ID", "Total % Grade", "Letter Grade",
         "Next Grade", "Points To Next", "Borderline"]
    ].sort_values("Points To Next")
    return (near,)


@app.cell
def moderation_draw_the_sample(module_3_sheet):
    """MODULE LEADER -- one per band, plus the borderline cases."""
    # No seed passed, so one is generated and handed back. Write it down --
    # it is what makes this draw defensible later. `also=` takes anyone the
    # leader wants a second opinion on regardless of band.
    moderation = sample_for_moderation(
        module_3_sheet, n=1, borderline="include"
    )

    mo.md(f"""
    **{moderation}**

    Bands that could not fill the quota: `{moderation.short_bands or "none"}`
    """)
    return (moderation,)


@app.cell
def moderation_who_was_chosen(moderation):
    moderation.selected[
        ["Name", "Student ID", "Total % Grade", "Letter Grade",
         "Points To Next", "Selected Because"]
    ]
    return


@app.cell
def moderation_build_the_pack(MODULE_3, moderation):
    """MODULE LEADER -- the folders the second marker is handed."""
    pack = build_moderation_pack(
        MODULE_3, moderation, MODULE_3.root / "Moderation", overwrite=True
    )

    mo.md(f"""
    **{pack}** — `{pack.root}`

    | | |
    |---|---|
    | copied, per assessment | `{pack.copied}` |
    | selected but nothing submitted | `{pack.missing or "none"}` |
    | manifest | `{pack.manifest.name}` |

    Only the coursework has a download to copy from. The quizzes, the MCQ and
    the exam have no submissions folder, so they are skipped — there is
    nothing to moderate and nothing has gone wrong.

    A student who was selected but submitted nothing is **named** rather than
    left as an empty folder. An empty folder in a pack reads as work the
    moderator has already been through.
    """)
    return (pack,)


@app.cell
def moderation_the_manifest(pack):
    # The record of the draw, and the handoff to the external examiner's pack
    # later. The folders can be rebuilt from this; without it they cannot.
    read_moderation_manifest(pack.root)[
        ["Module", "Student ID", "Letter Grade", "Selected Because",
         "Seed", "N Per Band", "Missing Submissions"]
    ]
    return


@app.cell
def moderation_the_seed_reproduces_the_draw(module_3_sheet, moderation, pack):
    # `repeated` is taken by the cw1 resubmission cell -- marimo needs every
    # name in the notebook to be defined once.
    recorded_seed = int(read_moderation_manifest(pack.root)["Seed"].iloc[0])
    redrawn = sample_for_moderation(
        module_3_sheet, n=1, borderline="include", seed=recorded_seed
    )

    reproduced = (
        redrawn.selected["Student ID"].tolist()
        == moderation.selected["Student ID"].tolist()
    )

    mo.md(f"""
    Seed `{recorded_seed}` read back out of the manifest reproduces the draw
    exactly: **{reproduced}**

    That is the whole point of recording it. Anyone with the marks and the
    seed can check that the sample was what it says it was, and re-running
    reuses the draw rather than quietly making a new one.
    """)
    return


@app.cell
def moderation_a_pack_spanning_two_assessments(MODULE, module_sheet):
    """MODULE LEADER -- the same thing on PS4001, which has two marked pieces.

    PS4003 has one assessment with a download to copy from. PS4001 has two,
    which is the ordinary case: the moderator gets each sampled student's
    coursework 1 *and* coursework 2, under the one band folder.
    """
    ps4001_sample = sample_for_moderation(module_sheet, n=1, borderline="include")
    ps4001_pack = build_moderation_pack(
        MODULE, ps4001_sample, MODULE.root / "Moderation", overwrite=True
    )

    mo.md(f"""
    **{ps4001_pack}** — `{ps4001_pack.root}`

    | | |
    |---|---|
    | copied, per assessment | `{ps4001_pack.copied}` |
    | nothing submitted | `{ps4001_pack.missing or "none"}` |

    Both courseworks are in the pack. The quizzes are not: their submissions
    folder holds Brightspace's CSV exports rather than student folders, so
    there is nothing to copy and nothing has gone wrong.

    The download also contains a `__MACOSX` directory and a stray
    `index.html`, which is what a real unzipped download looks like. Neither
    is a submission folder, so neither reaches the moderator — folder names
    are parsed, and anything that is not a Brightspace submission simply does
    not parse.
    """)
    return ps4001_pack, ps4001_sample


@app.cell
def si_intro():
    mo.md(
        """
        # The SI upload

        The last step. The student information system **issues** a file with
        one row per enrolled student and three columns blank -- `Mark`,
        `Grade`, and a `CD` nobody fills -- and you send the same file back.

        So this is the departmental sheet's problem again: we fill in two
        fields of somebody else's file and change nothing else. Not "produce
        a file in SI's format".

        Three quirks make that harder than it sounds, and all three are real:

        | | |
        |---|---|
        | line endings | **bare LF**, on a file a Windows system produced |
        | quoting | none, and no field holds a comma — `KEVIN O'MALLEY`, not `O'MALLEY, KEVIN` |
        | `#` prefixes | on headers *and* values, and `#CD` holds `#07` |

        The first is the one that bites. Python's `open(path, "w")` turns
        `\n` into `\r\n` on Windows, so writing the file back the obvious
        way changes **every line in it** — a function asked to change two
        fields rewriting all forty. `write_si_marks` reads and writes bytes.

        `#SPR_Code` is `#<student id>/<attempt>`, the attempt being how many
        times they have taken the module. It is matched on and **never
        rebuilt**: that number is SI's, and nothing we hold could reproduce
        it.
        """
    )
    return


@app.cell
def si_what_si_issued(MODULE):
    # The fixture plays SI here. Nothing in the package generates one of
    # these for real -- SI sends it, we fill it in and send it back.
    issued = MODULE.si_file_path.read_bytes()

    mo.md(f"""
    `{MODULE.si_file_path.name}`, as issued:

    ```
    {issued.decode().splitlines()[0]}
    {issued.decode().splitlines()[1]}
    ```

    | | |
    |---|---|
    | CRLF | `{b"\r\n" in issued}` |
    | BOM | `{issued[:3] == b"\xef\xbb\xbf"}` |
    | quotes | `{b'"' in issued}` |
    | rows | `{len(issued.decode().splitlines()) - 1}` |
    """)
    return (issued,)


@app.cell
def si_fill_all_three_modules(
    MODULE, MODULE_2, MODULE_3, module_sheet, module_2_sheet, module_3_sheet
):
    """MODULE LEADER -- one upload per module."""
    si_results = [
        write_si_marks(sheet, module.si_file_path, module.root / f"{module.code}_upload.CSV")
        for module, sheet in (
            (MODULE, module_sheet),
            (MODULE_2, module_2_sheet),
            (MODULE_3, module_3_sheet),
        )
    ]

    mo.md("\n".join(f"- **{result}**" for result in si_results))
    return (si_results,)


@app.cell
def si_what_changed(issued, si_results):
    # The claim is that two fields moved and nothing else did. Rather than
    # take that on trust, diff the bytes field by field.
    #
    # Every name here is prefixed: marimo requires each to be defined in
    # exactly one cell across the whole notebook, and `headers` was already
    # taken by the PS4002 section.
    si_filled = si_results[0].path.read_bytes()
    si_before = issued.decode().splitlines()
    si_after = si_filled.decode().splitlines()

    si_moved = sorted(
        {
            index
            for was, now in zip(si_before[1:], si_after[1:])
            for index, (a, b) in enumerate(zip(was.split(","), now.split(",")))
            if a != b
        }
    )
    si_headers = si_before[0].split(",")

    mo.md(f"""
    ```
    was:  {si_before[1]}
    now:  {si_after[1]}
    ```

    Fields that changed: **{[si_headers[i] for i in si_moved]}**

    | | |
    |---|---|
    | line count unchanged | `{len(si_before) == len(si_after)}` |
    | CRLF introduced | `{b"\r\n" in si_filled}` |
    | header untouched | `{si_before[0] == si_after[0]}` |

    Everything else — the `#` prefixes, `#07` with its leading zero, the
    `/3` on a third attempt, the name — is copied through byte for byte.

    The non-participant goes up as `Mark = 0`, `Grade = NG`, and **SI
    accepts `NG`** — checked, not assumed. So no special case is needed
    anywhere: what the departmental sheet says they got is what SI receives.
    """)
    return

@app.cell
def status_intro():
    mo.md(
        """
        # Keeping status

        Two halves, split by a single question: **can the code honestly
        know?**

        A step can tell that it produced a file. Whether that file was then
        *sent*, *read* or *accepted* is in somebody's head and never on disk.
        So each artefact has a flag the code sets and, where a person has to
        do something with it, one beside it that only a person can set.

        | the code sets, from evidence | a person sets |
        |---|---|
        | `departmental_sheet_written` | `sent_to_department` |
        | `moderation_pack_built` | `moderated` (per assessment) |
        | `si_file_written` | `si_submitted` |

        **The evidence is the return value, not the absence of a crash.**
        `distribute_feedback_sheets` completes perfectly happily having
        matched nothing at all — forty folders, no ids recognised, no
        exception. Ticking `sheets_distributed` off that puts a green mark
        against a step that did nothing, which is this package's usual enemy
        in a different hat.
        """
    )
    return


@app.cell
def status_record_the_module_level_steps(
    HANDLE, written, pack, si_results
):
    """MODULE LEADER -- and mostly not the module leader at all."""
    # One call per artefact. `record` looks up the rule for the result's type
    # and sets the flag only if the evidence supports it.
    for evidence in (written, pack, si_results[0]):
        HANDLE.record(evidence)

    recorded_status = ModuleFile.load(ROOT).module.status

    mo.md(f"""
    ```
    {recorded_status.model_dump()}
    ```

    The three automatic flags are set. The two manual ones are still `False`,
    and correctly so: nothing on this machine can know the sheet reached the
    department or that the upload was lodged with SI.
    """)
    return (recorded_status,)


@app.cell
def status_the_manual_half(HANDLE):
    """MODULE LEADER -- the button, or this call from a notebook."""
    # In the dashboard these two are buttons. Here they are the call a
    # technical user makes.
    HANDLE.set_module_status(sent_to_department=True)

    mo.md(f"""
    `{ModuleFile.load(ROOT).module.status.model_dump()}`

    `si_submitted` is deliberately left alone — the upload has not been
    lodged, and saying it had would be the one kind of lie this whole scheme
    exists to prevent.
    """)
    return


if __name__ == "__main__":
    app.run()
